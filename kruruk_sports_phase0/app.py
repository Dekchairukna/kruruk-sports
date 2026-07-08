import os
import random
import string
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path
from io import BytesIO
import base64
import uuid
import json
import hashlib
import hmac
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from dotenv import load_dotenv
from flask import Flask, flash, redirect, render_template, request, url_for, session, send_file, make_response, jsonify, current_app
from flask_login import current_user, login_required, login_user, logout_user
from werkzeug.utils import secure_filename

load_dotenv()

from extensions import db, login_manager
from sqlalchemy import text, inspect


def normalize_database_url(url: str) -> str:
    """Return a SQLAlchemy URL that works on Railway without system libpq.

    Railway often provides postgresql://...; SQLAlchemy maps that to psycopg2
    by default. psycopg2 can fail on minimal images when libpq.so.5 is absent.
    Force psycopg v3 binary driver instead.
    """
    if not url:
        return url
    if url.startswith("postgres://"):
        url = "postgresql://" + url[len("postgres://"):]
    if url.startswith("postgresql+psycopg2://"):
        url = "postgresql+psycopg://" + url[len("postgresql+psycopg2://"):]
    elif url.startswith("postgresql://"):
        url = "postgresql+psycopg://" + url[len("postgresql://"):]
    return url


def int_env(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "kruruk-sports-dev-key")
    instance_path = Path(app.instance_path)
    instance_path.mkdir(parents=True, exist_ok=True)
    raw_database_url = os.getenv("DATABASE_URL") or f"sqlite:///{instance_path / 'kruruk_sports.db'}"
    database_url = normalize_database_url(raw_database_url)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    engine_options = {"pool_pre_ping": True}
    if database_url.startswith("postgresql"):
        engine_options.update({
            "pool_recycle": 300,
            "connect_args": {"connect_timeout": int_env("DB_CONNECT_TIMEOUT", 10)},
        })
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = engine_options
    app.config["UPLOAD_FOLDER"] = str(Path(app.root_path) / "static" / "uploads")
    app.config["SOCIAL_LOGIN_ENABLED"] = os.getenv("SOCIAL_LOGIN_ENABLED", "1") == "1"
    app.config["PAYMENT_GATEWAYS"] = [g.strip() for g in os.getenv("PAYMENT_GATEWAYS", "manual,promptpay").split(",") if g.strip()]
    app.config["PROMPTPAY_ID"] = os.getenv("PROMPTPAY_ID", "")
    app.config["PAYMENT_RETURN_BASE_URL"] = os.getenv("PAYMENT_RETURN_BASE_URL", "")
    Path(app.config["UPLOAD_FOLDER"]).mkdir(parents=True, exist_ok=True)

    @app.get("/healthz")
    def healthz():
        return jsonify({
            "ok": True,
            "service": "kruruk-sports",
            "database_url_configured": bool(os.getenv("DATABASE_URL")),
            "promptpay_configured": bool(os.getenv("PROMPTPAY_ID")),
        })

    db.init_app(app)
    login_manager.init_app(app)

    with app.app_context():
        import models  # noqa: F401
        if os.getenv("SKIP_DB_INIT", "0") == "1":
            app.logger.warning("SKIP_DB_INIT=1: skipping db.create_all/schema seed during boot")
            try:
                # Even when full DB init is skipped on Railway, keep lightweight
                # schema upgrades enabled so new columns added in models.py
                # are created automatically on existing Postgres tables.
                ensure_schema_upgrades()
            except Exception:
                app.logger.exception("Schema upgrade failed during boot")
        else:
            try:
                db.create_all()
                ensure_schema_upgrades()
                seed_default_admin()
                seed_subscription_plans()
            except Exception:
                # Do not let Railway kill the web process during boot.
                # The exact database error will still be visible in deploy logs.
                app.logger.exception("Database initialization failed during boot")

    register_routes(app)
    return app



def ensure_schema_upgrades():
    """เพิ่มคอลัมน์ใหม่ให้ฐานเดิมโดยไม่ลบข้อมูลเดิม ใช้แทน migration เบื้องต้นในช่วงพัฒนา"""
    inspector = inspect(db.engine)

    def existing_columns(table_name):
        try:
            return {col["name"] for col in inspector.get_columns(table_name)}
        except Exception:
            return set()

    def add_column(table_name, column_sql):
        try:
            db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_sql}"))
            db.session.commit()
        except Exception:
            db.session.rollback()

    cols = existing_columns("sports")
    if "result_type" not in cols:
        add_column("sports", "result_type VARCHAR(30) DEFAULT 'score_only' NOT NULL")
    if "max_sets" not in cols:
        add_column("sports", "max_sets INTEGER DEFAULT 0 NOT NULL")
    if "points_per_set" not in cols:
        add_column("sports", "points_per_set INTEGER DEFAULT 0 NOT NULL")
    if "sets_to_win" not in cols:
        add_column("sports", "sets_to_win INTEGER DEFAULT 0 NOT NULL")

    cols = existing_columns("sport_divisions")
    if "result_type" not in cols:
        add_column("sport_divisions", "result_type VARCHAR(30) DEFAULT 'score_only' NOT NULL")
    if "max_sets" not in cols:
        add_column("sport_divisions", "max_sets INTEGER DEFAULT 0 NOT NULL")
    if "points_per_set" not in cols:
        add_column("sport_divisions", "points_per_set INTEGER DEFAULT 0 NOT NULL")
    if "sets_to_win" not in cols:
        add_column("sport_divisions", "sets_to_win INTEGER DEFAULT 0 NOT NULL")

    cols = existing_columns("round_robin_matches")
    if "set_scores" not in cols:
        add_column("round_robin_matches", "set_scores TEXT")
    if "score_history" not in cols:
        add_column("round_robin_matches", "score_history TEXT")
    if "point_diff" not in cols:
        add_column("round_robin_matches", "point_diff INTEGER DEFAULT 0")

    cols = existing_columns("knockout_matches")
    if "set_scores" not in cols:
        add_column("knockout_matches", "set_scores TEXT")
    if "score_history" not in cols:
        add_column("knockout_matches", "score_history TEXT")
    if "point_diff" not in cols:
        add_column("knockout_matches", "point_diff INTEGER DEFAULT 0")

    cols = existing_columns("teams")
    if "line_user_id" not in cols:
        add_column("teams", "line_user_id VARCHAR(120)")
    if "line_contact_name" not in cols:
        add_column("teams", "line_contact_name VARCHAR(180)")
    if "line_invite_sent_at" not in cols:
        add_column("teams", "line_invite_sent_at TIMESTAMP")
    if "line_invite_error" not in cols:
        add_column("teams", "line_invite_error TEXT")

    cols = existing_columns("users")
    if "username" not in cols:
        add_column("users", "username VARCHAR(80)")
    if "avatar_url" not in cols:
        add_column("users", "avatar_url VARCHAR(500)")
    if "social_provider" not in cols:
        add_column("users", "social_provider VARCHAR(40)")
    if "social_id" not in cols:
        add_column("users", "social_id VARCHAR(255)")
    if "last_login_at" not in cols:
        add_column("users", "last_login_at DATETIME")

    cols = existing_columns("invoices")
    if "gateway_reference" not in cols:
        add_column("invoices", "gateway_reference VARCHAR(255)")
    if "payment_url" not in cols:
        add_column("invoices", "payment_url TEXT")


def seed_default_admin():
    """Create/repair the main superadmin account.

    Login ID requested by user: superadmin
    Password requested by user: yagami1225
    Email is kept as a valid internal email because the users.email column is unique/non-null.
    """
    from models import User

    username = os.getenv("DEFAULT_ADMIN_USERNAME", "superadmin").lower().strip()
    email = os.getenv("DEFAULT_ADMIN_EMAIL", "superadmin@kruruksports.local").lower().strip()
    password = os.getenv("DEFAULT_ADMIN_PASSWORD", "yagami1225")
    name = os.getenv("DEFAULT_ADMIN_NAME", "KRURUK Super Admin")

    admin = User.query.filter_by(username=username).first() or User.query.filter_by(email=email).first()
    if not admin:
        # ถ้าเคยมี admin เก่าจาก Phase ก่อน ให้ยกระดับและผูก username นี้แทนการสร้างซ้ำ
        admin = User.query.filter_by(email="admin@kruruksports.com").first()
    if not admin:
        admin = User(name=name, email=email, username=username, role="superadmin")
        db.session.add(admin)
    admin.name = admin.name or name
    admin.email = admin.email or email
    admin.username = username
    admin.role = "superadmin"
    # ตั้งรหัสผ่านตามที่กำหนดทุกครั้ง เพื่อกันลืมรหัสตอน deploy ใหม่
    admin.set_password(password)
    db.session.commit()



def seed_subscription_plans():
    """สร้างแพ็กเกจเริ่มต้น Free, Basic, Pro, Enterprise ถ้ายังไม่มี"""
    from models import SubscriptionPlan
    defaults = [
        dict(code="free", name="Free", price=0, billing_period="monthly", duration_days=3650, max_events=1, max_teams_per_event=4, max_athletes_per_event=80, allow_live_board=False, allow_certificates=False, allow_reports_pdf=False, sort_order=1, description="เริ่มต้นทดลองใช้ เหมาะกับงานเล็ก"),
        dict(code="basic", name="Basic", price=299, billing_period="monthly", duration_days=30, max_events=3, max_teams_per_event=12, max_athletes_per_event=300, allow_live_board=True, allow_certificates=False, allow_reports_pdf=False, sort_order=2, description="เหมาะกับกีฬาสีหรือกิจกรรมโรงเรียนขนาดเล็ก"),
        dict(code="pro", name="Pro", price=799, billing_period="monthly", duration_days=30, max_events=10, max_teams_per_event=32, max_athletes_per_event=1000, allow_live_board=True, allow_certificates=True, allow_reports_pdf=True, sort_order=3, description="เหมาะกับกีฬาเครือข่าย/เทศบาล/งานหลายรายการ"),
        dict(code="enterprise", name="Enterprise", price=0, billing_period="custom", duration_days=365, max_events=-1, max_teams_per_event=-1, max_athletes_per_event=-1, allow_live_board=True, allow_certificates=True, allow_reports_pdf=True, sort_order=4, description="องค์กรใหญ่ ไม่จำกัดจำนวน กำหนดราคาตามตกลง"),
    ]
    for data in defaults:
        plan = SubscriptionPlan.query.filter_by(code=data["code"]).first()
        if not plan:
            plan = SubscriptionPlan(**data)
            db.session.add(plan)
        else:
            # เติม field ใหม่โดยไม่ทับชื่อ/ราคา/ลิมิตที่แอดมินอาจแก้ไว้แล้ว
            for key, value in data.items():
                if getattr(plan, key, None) is None:
                    setattr(plan, key, value)
    db.session.commit()

    # เติมแพ็กเกจ Free ให้องค์กรเดิมที่ยังไม่มี subscription เพื่อให้ Billing แสดงครบ
    from models import Organization, OrganizationSubscription
    free_plan = SubscriptionPlan.query.filter_by(code="free").first()
    if free_plan:
        for org in Organization.query.all():
            exists = OrganizationSubscription.query.filter_by(organization_id=org.id).first()
            if not exists:
                db.session.add(OrganizationSubscription(
                    organization_id=org.id,
                    plan_id=free_plan.id,
                    status="active",
                    start_date=date.today(),
                    end_date=None,
                    manual_payment_note="ระบบกำหนดแพ็กเกจ Free ให้อัตโนมัติ",
                ))
        db.session.commit()


def get_free_plan():
    from models import SubscriptionPlan
    return SubscriptionPlan.query.filter_by(code="free").first() or SubscriptionPlan.query.order_by(SubscriptionPlan.sort_order).first()


def get_current_subscription(org):
    from models import OrganizationSubscription
    if not org:
        return None
    sub = OrganizationSubscription.query.filter(
        OrganizationSubscription.organization_id == org.id,
        OrganizationSubscription.status == "active",
    ).order_by(OrganizationSubscription.created_at.desc()).first()
    if sub and sub.end_date and sub.end_date < date.today():
        sub.status = "expired"
        db.session.commit()
        return None
    return sub


def get_current_plan(org):
    sub = get_current_subscription(org)
    if sub and sub.plan:
        return sub.plan
    return get_free_plan()


def ensure_free_subscription(org):
    from models import OrganizationSubscription
    if not org or get_current_subscription(org):
        return None
    plan = get_free_plan()
    if not plan:
        return None
    sub = OrganizationSubscription(
        organization_id=org.id,
        plan_id=plan.id,
        status="active",
        start_date=date.today(),
        end_date=None,
        manual_payment_note="ระบบกำหนดแพ็กเกจ Free ให้อัตโนมัติ",
    )
    db.session.add(sub)
    return sub


def feature_allowed(org, feature):
    plan = get_current_plan(org)
    if not plan:
        return False
    return bool(getattr(plan, feature, False))


def limit_value(plan, field):
    value = getattr(plan, field, 0)
    return 10**12 if value is None or value < 0 else value


def deny_upgrade(message, endpoint="events", **route_values):
    flash(f"{message} กรุณาอัปเกรดแพ็กเกจ", "warning")
    return redirect(url_for(endpoint, **route_values))


def check_event_limit(org):
    from models import Event
    if current_user.is_authenticated and current_user.is_superadmin:
        return True, ""
    plan = get_current_plan(org)
    if not plan:
        return False, "ยังไม่มีแพ็กเกจเริ่มต้นในระบบ"
    current = Event.query.filter_by(organization_id=org.id).count()
    limit = limit_value(plan, "max_events")
    if current >= limit:
        return False, f"แพ็กเกจ {plan.name} สร้างงานแข่งขันได้สูงสุด {plan.max_events} งาน"
    return True, ""


def check_team_limit(event):
    from models import Team
    if current_user.is_authenticated and current_user.is_superadmin:
        return True, ""
    plan = get_current_plan(event.organization)
    if not plan:
        return False, "ยังไม่มีแพ็กเกจเริ่มต้นในระบบ"
    current = Team.query.filter_by(event_id=event.id).count()
    limit = limit_value(plan, "max_teams_per_event")
    if current >= limit:
        return False, f"แพ็กเกจ {plan.name} เพิ่มทีมได้สูงสุด {plan.max_teams_per_event} ทีมต่อหนึ่งงาน"
    return True, ""


def check_athlete_limit(event, additional=1):
    from models import Athlete, Team
    if current_user.is_authenticated and current_user.is_superadmin:
        return True, ""
    plan = get_current_plan(event.organization)
    if not plan:
        return False, "ยังไม่มีแพ็กเกจเริ่มต้นในระบบ"
    current = Athlete.query.join(Team).filter(Team.event_id == event.id).count()
    limit = limit_value(plan, "max_athletes_per_event")
    if current + additional > limit:
        return False, f"แพ็กเกจ {plan.name} เพิ่มนักกีฬาได้สูงสุด {plan.max_athletes_per_event} คนต่อหนึ่งงาน"
    return True, ""


def check_feature_or_redirect(event, feature, label, fallback="event_detail"):
    if not feature_allowed(event.organization, feature):
        flash(f"แพ็กเกจปัจจุบันยังไม่มีสิทธิ์ใช้ {label} กรุณาอัปเกรดแพ็กเกจ", "warning")
        return redirect(url_for(fallback, event_id=event.id))
    return None


def make_invoice_no(org_id):
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    return f"INV-{org_id}-{stamp}-{random.randint(100,999)}"


def superadmin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_superadmin:
            flash("เมนูนี้สำหรับ Super Admin เท่านั้น", "danger")
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    return wrapper


# -----------------------------
# Phase 13B: Social Login helpers
# -----------------------------
def oauth_configured(provider):
    provider = provider.lower()
    if provider == "google":
        return bool(os.getenv("GOOGLE_CLIENT_ID") and os.getenv("GOOGLE_CLIENT_SECRET"))
    if provider == "line":
        return bool(os.getenv("LINE_CLIENT_ID") and os.getenv("LINE_CLIENT_SECRET"))
    if provider == "facebook":
        return bool(os.getenv("FACEBOOK_CLIENT_ID") and os.getenv("FACEBOOK_CLIENT_SECRET"))
    return False


def available_oauth_providers():
    providers = []
    for code, label, icon in [
        ("google", "Google", "bi-google"),
        ("line", "LINE", "bi-chat-dots"),
        ("facebook", "Facebook", "bi-facebook"),
    ]:
        providers.append({"code": code, "label": label, "icon": icon, "ready": oauth_configured(code)})
    return providers


def build_oauth_authorize_url(provider, state):
    provider = provider.lower()
    redirect_uri = url_for("social_callback", provider=provider, _external=True)
    if provider == "google":
        params = {
            "client_id": os.getenv("GOOGLE_CLIENT_ID"),
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": "openid email profile",
            "state": state,
            "access_type": "offline",
            "prompt": "select_account",
        }
        return "https://accounts.google.com/o/oauth2/v2/auth?" + urlencode(params)
    if provider == "line":
        params = {
            "response_type": "code",
            "client_id": os.getenv("LINE_CLIENT_ID"),
            "redirect_uri": redirect_uri,
            "state": state,
            "scope": "profile openid email",
        }
        return "https://access.line.me/oauth2/v2.1/authorize?" + urlencode(params)
    if provider == "facebook":
        params = {
            "client_id": os.getenv("FACEBOOK_CLIENT_ID"),
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": "email,public_profile",
            "state": state,
        }
        return "https://www.facebook.com/v19.0/dialog/oauth?" + urlencode(params)
    return None


def exchange_oauth_profile(provider, code):
    """ดึงโปรไฟล์ OAuth ผ่าน Authlib ถ้าติดตั้งแล้ว คืน dict มาตรฐาน {id,email,name,avatar,token}."""
    try:
        from authlib.integrations.requests_client import OAuth2Session
    except Exception as exc:
        raise RuntimeError("ยังไม่ได้ติดตั้ง Authlib: pip install Authlib") from exc

    provider = provider.lower()
    redirect_uri = url_for("social_callback", provider=provider, _external=True)
    if provider == "google":
        client = OAuth2Session(os.getenv("GOOGLE_CLIENT_ID"), os.getenv("GOOGLE_CLIENT_SECRET"), scope="openid email profile", redirect_uri=redirect_uri)
        token = client.fetch_token("https://oauth2.googleapis.com/token", code=code)
        profile = client.get("https://openidconnect.googleapis.com/v1/userinfo").json()
        return {"id": profile.get("sub"), "email": profile.get("email"), "name": profile.get("name"), "avatar": profile.get("picture"), "token": token, "raw": profile}
    if provider == "line":
        client = OAuth2Session(os.getenv("LINE_CLIENT_ID"), os.getenv("LINE_CLIENT_SECRET"), scope="profile openid email", redirect_uri=redirect_uri)
        token = client.fetch_token("https://api.line.me/oauth2/v2.1/token", code=code)
        profile = client.get("https://api.line.me/v2/profile").json()
        # LINE email จะอยู่ใน id_token ถ้า Channel เปิดสิทธิ์ email; เก็บได้เมื่อ Authlib decode ได้ในอนาคต
        return {"id": profile.get("userId"), "email": None, "name": profile.get("displayName"), "avatar": profile.get("pictureUrl"), "token": token, "raw": profile}
    if provider == "facebook":
        client = OAuth2Session(os.getenv("FACEBOOK_CLIENT_ID"), os.getenv("FACEBOOK_CLIENT_SECRET"), redirect_uri=redirect_uri)
        token = client.fetch_token("https://graph.facebook.com/v19.0/oauth/access_token", code=code)
        profile = client.get("https://graph.facebook.com/me?fields=id,name,email,picture.type(large)").json()
        avatar = (((profile.get("picture") or {}).get("data") or {}).get("url"))
        return {"id": profile.get("id"), "email": profile.get("email"), "name": profile.get("name"), "avatar": avatar, "token": token, "raw": profile}
    raise RuntimeError("ไม่รองรับ Provider นี้")


def upsert_social_user(provider, profile):
    from models import OAuthAccount, User
    provider_user_id = str(profile.get("id") or "").strip()
    email = (profile.get("email") or "").lower().strip() or None
    name = (profile.get("name") or email or f"{provider.title()} User").strip()
    avatar = profile.get("avatar")
    if not provider_user_id:
        raise RuntimeError("Provider ไม่ส่ง user id กลับมา")

    account = OAuthAccount.query.filter_by(provider=provider, provider_user_id=provider_user_id).first()
    if account:
        user = account.user
    else:
        user = User.query.filter_by(email=email).first() if email else None
        if not user:
            user = User(name=name, email=email or f"{provider}_{provider_user_id}@social.local", role="organization_admin")
            user.set_password(uuid.uuid4().hex)
            db.session.add(user)
            db.session.flush()
        account = OAuthAccount(user_id=user.id, provider=provider, provider_user_id=provider_user_id)
        db.session.add(account)

    user.name = user.name or name
    user.avatar_url = avatar or user.avatar_url
    user.social_provider = provider
    user.social_id = provider_user_id
    user.last_login_at = datetime.utcnow()
    account.email = email
    account.name = name
    account.avatar_url = avatar
    token = profile.get("token") or {}
    account.access_token = token.get("access_token") if isinstance(token, dict) else None
    account.refresh_token = token.get("refresh_token") if isinstance(token, dict) else None
    account.raw_profile = json.dumps(profile.get("raw") or {}, ensure_ascii=False)
    db.session.commit()
    return user


# -----------------------------
# Phase 13C: Payment helpers
# -----------------------------
def payment_gateway_enabled(gateway):
    return gateway in (current_app_payment_gateways())


def current_app_payment_gateways():
    from flask import current_app
    return current_app.config.get("PAYMENT_GATEWAYS", ["manual", "promptpay"])


def crc16_ccitt(payload: str) -> str:
    crc = 0xFFFF
    for ch in payload.encode("utf-8"):
        crc ^= ch << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = ((crc << 1) ^ 0x1021) & 0xFFFF
            else:
                crc = (crc << 1) & 0xFFFF
    return f"{crc:04X}"


def emv_field(tag, value):
    value = str(value)
    return f"{tag}{len(value):02d}{value}"


def promptpay_target_to_aid(promptpay_id):
    target = "".join(ch for ch in promptpay_id if ch.isdigit())
    if len(target) == 10:  # mobile phone
        target = "0066" + target[1:]
        return emv_field("01", target)
    if len(target) == 13:  # national id / tax id
        return emv_field("02", target)
    return emv_field("03", target)  # e-wallet/other


def build_promptpay_payload(promptpay_id, amount):
    """Build a Thai PromptPay EMVCo payload that banking apps can scan.

    จุดสำคัญ: ช่อง CRC ต้องเป็น tag 63 ความยาว 04 เสมอ (`6304`)
    แล้วค่อยคำนวณ CRC จาก payload ที่ลงท้ายด้วย `6304` ก่อนนำค่า CRC มาต่อท้าย
    ถ้าใช้ `6300` ธนาคารหลายแอปจะเห็น QR แต่สแกนจ่ายไม่ได้
    """
    merchant = emv_field("00", "A000000677010111") + promptpay_target_to_aid(promptpay_id)
    payload = ""
    payload += emv_field("00", "01")      # Payload Format Indicator
    payload += emv_field("01", "12")      # Dynamic QR
    payload += emv_field("29", merchant)  # PromptPay Merchant Account Information
    payload += emv_field("53", "764")     # THB
    amount_value = float(amount or 0)
    if amount_value > 0:
        payload += emv_field("54", f"{amount_value:.2f}")
    payload += emv_field("58", "TH")
    payload_for_crc = payload + "6304"
    return payload_for_crc + crc16_ccitt(payload_for_crc)


def is_promptpay_payload_scanable(payload):
    """Basic check for PromptPay payload CRC field. Returns False for old 6300 payloads."""
    if not payload or not isinstance(payload, str):
        return False
    payload = payload.strip()
    if len(payload) < 8 or not payload[-8:-4] == "6304":
        return False
    expected = crc16_ccitt(payload[:-4])
    return payload[-4:].upper() == expected.upper()


def ensure_promptpay_payload_for_transaction(txn):
    """Rebuild old invalid PromptPay payloads already saved in the database."""
    if not txn or txn.gateway != "promptpay":
        return False
    if is_promptpay_payload_scanable(txn.qr_payload):
        return False
    promptpay_id = current_app.config.get("PROMPTPAY_ID")
    if not promptpay_id:
        return False
    amount = txn.invoice.amount if txn.invoice else txn.amount
    txn.qr_payload = build_promptpay_payload(promptpay_id, amount)
    txn.note = "PromptPay QR payload rebuilt with valid EMV CRC 6304"
    db.session.commit()
    return True


def make_promptpay_qr_png(payload):
    """Return PNG bytes for a PromptPay QR payload.

    ถ้าเครื่องยังไม่ได้ติดตั้ง qrcode/Pillow จะคืน error กลับไปให้หน้าเว็บแสดงชัด ๆ
    แทนการกลืน exception แล้วเหลือเป็นข้อความ payload ยาว ๆ อย่างเดียว
    """
    if not payload:
        return None, "ไม่พบข้อมูล PromptPay payload"
    try:
        import qrcode
        from qrcode.constants import ERROR_CORRECT_M
    except Exception as exc:
        return None, f"ยังไม่ได้ติดตั้งไลบรารี QR: {exc}"
    try:
        qr = qrcode.QRCode(
            version=None,
            error_correction=ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(payload)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        return buffer.getvalue(), None
    except Exception as exc:
        return None, f"สร้าง QR ไม่สำเร็จ: {exc}"


def make_promptpay_qr_data_uri(payload):
    png_bytes, error = make_promptpay_qr_png(payload)
    if not png_bytes:
        return None, error
    return "data:image/png;base64," + base64.b64encode(png_bytes).decode("ascii"), None


def mark_invoice_paid(invoice, gateway="manual", reference=None):
    invoice.status = "paid"
    invoice.paid_at = invoice.paid_at or datetime.utcnow()
    invoice.payment_method = gateway
    if hasattr(invoice, "gateway_reference"):
        invoice.gateway_reference = reference or invoice.gateway_reference
    if invoice.subscription:
        invoice.subscription.status = "active"
        OrganizationSubscription.query.filter(
            OrganizationSubscription.organization_id == invoice.organization_id,
            OrganizationSubscription.id != invoice.subscription.id,
            OrganizationSubscription.status == "active",
        ).update({"status": "cancelled"})
    db.session.commit()


def create_stripe_checkout_url(invoice, provider_reference):
    try:
        import stripe
    except Exception:
        return None
    secret_key = os.getenv("STRIPE_SECRET_KEY", "")
    price_name = invoice.title or invoice.invoice_no
    if not secret_key:
        return None
    stripe.api_key = secret_key
    base = os.getenv("PAYMENT_RETURN_BASE_URL") or request.url_root.rstrip("/")
    session_obj = stripe.checkout.Session.create(
        mode="payment",
        line_items=[{
            "price_data": {
                "currency": invoice.currency.lower(),
                "product_data": {"name": price_name},
                "unit_amount": int(round(float(invoice.amount) * 100)),
            },
            "quantity": 1,
        }],
        success_url=f"{base}{url_for('organizations_billing', org_id=invoice.organization_id)}?payment=success",
        cancel_url=f"{base}{url_for('organizations_billing', org_id=invoice.organization_id)}?payment=cancel",
        metadata={"invoice_id": str(invoice.id), "provider_reference": provider_reference},
    )
    return session_obj.url


def save_upload(file):
    if not file or not file.filename:
        return None
    filename = secure_filename(file.filename)
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    final_name = f"{stamp}_{filename}"
    upload_path = Path(current_app_upload_folder()) / final_name
    file.save(upload_path)
    return final_name


def current_app_upload_folder():
    from flask import current_app
    return current_app.config["UPLOAD_FOLDER"]


def user_org_ids():
    from models import OrganizationMember
    if current_user.is_superadmin:
        return None
    return [m.organization_id for m in OrganizationMember.query.filter_by(user_id=current_user.id).all()]


def can_access_org(org_id):
    if current_user.is_superadmin:
        return True
    return org_id in (user_org_ids() or [])


def org_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        org_id = kwargs.get("org_id") or request.form.get("organization_id") or request.args.get("organization_id")
        if org_id and not can_access_org(int(org_id)):
            flash("คุณไม่มีสิทธิ์เข้าถึงองค์กรนี้", "danger")
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    return wrapper



def line_channel_access_token():
    """LINE Messaging API token สำหรับส่งข้อความไปหัวหน้าสีผ่าน Official Account"""
    return (
        os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
        or os.getenv("LINE_MESSAGING_CHANNEL_ACCESS_TOKEN")
        or os.getenv("LINE_BOT_CHANNEL_ACCESS_TOKEN")
        or ""
    ).strip()


def is_line_message_ready():
    return bool(line_channel_access_token())


def build_team_invite_message(team):
    """ข้อความมาตรฐานสำหรับส่งให้หัวหน้าสี/หัวหน้าทีมเข้ากรอกข้อมูลเอง"""
    entry_url = url_for("team_entry", _external=True)
    event = team.event
    return (
        f"แจ้งหัวหน้าสี/หัวหน้าทีม\n\n"
        f"งาน: {event.name}\n"
        f"ทีม/สี: {team.name}\n"
        f"รหัสกรอกข้อมูล: {team.access_code}\n\n"
        f"ให้เข้ากรอกข้อมูลนักกีฬา/ผู้ควบคุมทีมที่ลิงก์นี้\n"
        f"{entry_url}\n\n"
        f"วิธีเข้าใช้งาน:\n"
        f"1) กดลิงก์ด้านบน\n"
        f"2) กรอกรหัสทีม/สี: {team.access_code}\n"
        f"3) กดเข้าสู่หน้ากรอกข้อมูล\n"
        f"4) เพิ่มรายชื่อนักกีฬา ผู้ฝึกสอน และตรวจสอบข้อมูลให้ถูกต้องก่อนบันทึก\n\n"
        f"หมายเหตุ: รหัสนี้ใช้ได้เฉพาะทีม/สีของท่านเท่านั้น"
    )


def send_line_push_message(line_user_id, message):
    """ส่งข้อความ LINE ด้วย Messaging API push message.

    line_user_id ต้องเป็น userId ของ LINE OA เช่น Uxxxxxxxx ไม่ใช่ LINE ID ที่ผู้ใช้ตั้งเอง.
    """
    token = line_channel_access_token()
    if not token:
        return False, "ยังไม่ได้ตั้งค่า LINE_CHANNEL_ACCESS_TOKEN ใน .env/Railway Variables"
    line_user_id = (line_user_id or "").strip()
    if not line_user_id:
        return False, "ยังไม่ได้กรอก LINE userId ของหัวหน้าสี/ทีม"
    payload = json.dumps({
        "to": line_user_id,
        "messages": [{"type": "text", "text": message}],
    }).encode("utf-8")
    req = Request(
        "https://api.line.me/v2/bot/message/push",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}",
        },
        method="POST",
    )
    try:
        with urlopen(req, timeout=12) as resp:
            if 200 <= resp.status < 300:
                return True, "ส่งข้อความ LINE สำเร็จ"
            return False, f"LINE API ตอบกลับสถานะ {resp.status}"
    except HTTPError as exc:
        try:
            body = exc.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        return False, f"LINE API error {exc.code}: {body[:300]}"
    except URLError as exc:
        return False, f"เชื่อมต่อ LINE API ไม่สำเร็จ: {exc.reason}"
    except Exception as exc:
        return False, f"ส่งข้อความ LINE ไม่สำเร็จ: {exc}"


def register_routes(app):
    from models import Event, Organization, OrganizationMember, Team, TeamFile, TeamPerson, TeamProfile, User, Athlete, AthleteRegistration, Coach, SportCategory, Sport, SportDivision, RoundRobinCompetition, RoundRobinGroup, RoundRobinGroupTeam, RoundRobinMatch, KnockoutCompetition, KnockoutMatch, RankingCompetition, RankingResult, ContestCompetition, ContestCriterion, ContestJudge, ContestScore, ContestResult, CertificateTemplate, CertificateRecipient, LiveBoardSetting, SubscriptionPlan, OrganizationSubscription, Invoice, OAuthAccount, PaymentTransaction

    @app.context_processor
    def inject_globals():
        def active_event_id():
            """หา event ปัจจุบันให้ sidebar พาไป Teams/Sport Setup/Billing ได้ถูกหน้า"""
            try:
                args = request.view_args or {}
                endpoint = request.endpoint or ""
                if args.get("event_id"):
                    return int(args.get("event_id"))
                if args.get("team_id"):
                    team = Team.query.get(args.get("team_id"))
                    return team.event_id if team else None
                if args.get("comp_id"):
                    comp = None
                    if endpoint.startswith("rr_") or "round_robin" in endpoint:
                        comp = RoundRobinCompetition.query.get(args.get("comp_id"))
                    elif endpoint.startswith("ranking"):
                        comp = RankingCompetition.query.get(args.get("comp_id"))
                    elif endpoint.startswith("knockout"):
                        comp = KnockoutCompetition.query.get(args.get("comp_id"))
                    elif endpoint.startswith("contest"):
                        comp = ContestCompetition.query.get(args.get("comp_id"))
                    return comp.event_id if comp else None
                if args.get("match_id"):
                    match = RoundRobinMatch.query.get(args.get("match_id")) or KnockoutMatch.query.get(args.get("match_id"))
                    return match.competition.event_id if match and match.competition else None
                if args.get("division_id"):
                    div = SportDivision.query.get(args.get("division_id"))
                    return div.sport.event_id if div and div.sport else None
                if args.get("sport_id"):
                    sport = Sport.query.get(args.get("sport_id"))
                    return sport.event_id if sport else None
                if args.get("category_id"):
                    cat = SportCategory.query.get(args.get("category_id"))
                    return cat.event_id if cat else None
                # ถ้าไม่ได้อยู่ในหน้าที่มี event_id ให้ใช้ “งานล่าสุดที่เลือกไว้” ใน session
                sid = session.get("active_event_id")
                if sid:
                    ev = Event.query.get(int(sid))
                    if ev and can_access_org(ev.organization_id):
                        return ev.id
            except Exception:
                return None
            return None

        def active_org_id():
            try:
                eid = active_event_id()
                if eid:
                    event = Event.query.get(eid)
                    return event.organization_id if event else None
                args = request.view_args or {}
                if args.get("org_id"):
                    return int(args.get("org_id"))
                soid = session.get("active_org_id")
                if soid and can_access_org(int(soid)):
                    return int(soid)
            except Exception:
                return None
            return None

        return {
            "current_year": datetime.now().year,
            "render_certificate_body": render_certificate_body,
            "get_current_plan": get_current_plan,
            "feature_allowed": feature_allowed,
            "oauth_providers": available_oauth_providers(),
            "payment_gateways": current_app_payment_gateways(),
            "line_message_ready": is_line_message_ready(),
            "active_event_id": active_event_id,
            "active_org_id": active_org_id,
            "set_score_rows": set_score_rows,
            "set_point_totals": set_point_totals,
        }

    @app.route("/")
    def index():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        return redirect(url_for("login"))

    @app.route("/register", methods=["GET", "POST"])
    def register():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            email = request.form.get("email", "").lower().strip()
            password = request.form.get("password", "")
            org_name = request.form.get("org_name", "").strip()
            org_type = request.form.get("org_type", "โรงเรียน")
            if not name or not email or not password or not org_name:
                flash("กรุณากรอกข้อมูลให้ครบ", "danger")
                return render_template("auth/register.html")
            if User.query.filter_by(email=email).first():
                flash("อีเมลนี้มีผู้ใช้งานแล้ว", "danger")
                return render_template("auth/register.html")

            user = User(name=name, email=email, role="organization_admin")
            user.set_password(password)
            org = Organization(name=org_name, org_type=org_type)
            db.session.add_all([user, org])
            db.session.flush()
            db.session.add(OrganizationMember(user_id=user.id, organization_id=org.id, role="organization_admin"))
            db.session.commit()
            login_user(user)
            flash("สมัครสมาชิกและสร้างองค์กรเรียบร้อย", "success")
            return redirect(url_for("dashboard"))
        return render_template("auth/register.html")

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            identifier = request.form.get("email", "").lower().strip()
            password = request.form.get("password", "")
            user = User.query.filter((User.email == identifier) | (User.username == identifier)).first()
            if not user or not user.check_password(password):
                flash("ID/อีเมล หรือรหัสผ่านไม่ถูกต้อง", "danger")
                return render_template("auth/login.html")
            login_user(user)
            return redirect(url_for("dashboard"))
        return render_template("auth/login.html")

    @app.route("/auth/<provider>")
    def social_login(provider):
        provider = provider.lower()
        if provider not in ("google", "line", "facebook"):
            flash("ไม่รองรับ Social Login Provider นี้", "danger")
            return redirect(url_for("login"))
        if not oauth_configured(provider):
            flash(f"ยังไม่ได้ตั้งค่า {provider.title()} Login ใน .env", "warning")
            return redirect(url_for("login"))
        state = uuid.uuid4().hex
        session["oauth_state"] = state
        session["oauth_provider"] = provider
        return redirect(build_oauth_authorize_url(provider, state))

    @app.route("/auth/<provider>/callback")
    def social_callback(provider):
        provider = provider.lower()
        if request.args.get("state") != session.get("oauth_state") or provider != session.get("oauth_provider"):
            flash("Social Login state ไม่ถูกต้อง กรุณาลองใหม่", "danger")
            return redirect(url_for("login"))
        code = request.args.get("code")
        if not code:
            flash("ไม่ได้รับรหัสยืนยันจาก Provider", "danger")
            return redirect(url_for("login"))
        try:
            profile = exchange_oauth_profile(provider, code)
            user = upsert_social_user(provider, profile)
            login_user(user)
            flash(f"เข้าสู่ระบบด้วย {provider.title()} สำเร็จ", "success")
            return redirect(url_for("dashboard"))
        except Exception as exc:
            flash(f"Social Login ไม่สำเร็จ: {exc}", "danger")
            return redirect(url_for("login"))

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        flash("ออกจากระบบแล้ว", "info")
        return redirect(url_for("login"))

    @app.route("/dashboard")
    @login_required
    def dashboard():
        org_ids = user_org_ids()
        org_query = Organization.query
        event_query = Event.query
        if org_ids is not None:
            org_query = org_query.filter(Organization.id.in_(org_ids or [0]))
            event_query = event_query.filter(Event.organization_id.in_(org_ids or [0]))
        org_count = org_query.count()
        event_count = event_query.count()
        open_count = event_query.filter(Event.status == "open").count()
        latest_events = event_query.order_by(Event.created_at.desc()).limit(5).all()
        from models import Team
        team_query = Team.query.join(Event)
        if org_ids is not None:
            team_query = team_query.filter(Event.organization_id.in_(org_ids or [0]))
        team_count = team_query.count()
        return render_template("dashboard.html", org_count=org_count, event_count=event_count, open_count=open_count, team_count=team_count, latest_events=latest_events)

    @app.route("/organizations")
    @login_required
    def organizations():
        org_ids = user_org_ids()
        query = Organization.query
        if org_ids is not None:
            query = query.filter(Organization.id.in_(org_ids or [0]))
        return render_template("orgs/list.html", organizations=query.order_by(Organization.created_at.desc()).all())

    @app.route("/organizations/new", methods=["GET", "POST"])
    @login_required
    def organization_new():
        if request.method == "POST":
            org = Organization(
                name=request.form.get("name", "").strip(),
                org_type=request.form.get("org_type", "โรงเรียน"),
                logo=save_upload(request.files.get("logo")),
            )
            if not org.name:
                flash("กรุณากรอกชื่อองค์กร", "danger")
                return render_template("orgs/form.html", org=None)
            db.session.add(org)
            db.session.flush()
            db.session.add(OrganizationMember(user_id=current_user.id, organization_id=org.id, role="organization_admin"))
            ensure_free_subscription(org)
            db.session.commit()
            flash("สร้างองค์กรแล้ว", "success")
            return redirect(url_for("organizations"))
        return render_template("orgs/form.html", org=None)

    @app.route("/organizations/<int:org_id>/edit", methods=["GET", "POST"])
    @login_required
    @org_required
    def organization_edit(org_id):
        org = Organization.query.get_or_404(org_id)
        if request.method == "POST":
            org.name = request.form.get("name", "").strip()
            org.org_type = request.form.get("org_type", "โรงเรียน")
            uploaded = save_upload(request.files.get("logo"))
            if uploaded:
                org.logo = uploaded
            db.session.commit()
            flash("บันทึกข้อมูลองค์กรแล้ว", "success")
            return redirect(url_for("organizations"))
        return render_template("orgs/form.html", org=org)

    @app.route("/events")
    @login_required
    def events():
        org_ids = user_org_ids()
        query = Event.query
        if org_ids is not None:
            query = query.filter(Event.organization_id.in_(org_ids or [0]))
        return render_template("events/list.html", events=query.order_by(Event.created_at.desc()).all())

    @app.route("/events/new", methods=["GET", "POST"])
    @login_required
    def event_new():
        org_ids = user_org_ids()
        org_query = Organization.query
        if org_ids is not None:
            org_query = org_query.filter(Organization.id.in_(org_ids or [0]))
        orgs = org_query.all()
        if request.method == "POST":
            org_id = int(request.form.get("organization_id"))
            if not can_access_org(org_id):
                flash("คุณไม่มีสิทธิ์สร้างงานในองค์กรนี้", "danger")
                return redirect(url_for("events"))
            org = Organization.query.get_or_404(org_id)
            ok, msg = check_event_limit(org)
            if not ok:
                return deny_upgrade(msg, "organizations_billing", org_id=org.id)
            event = Event(
                organization_id=org_id,
                name=request.form.get("name", "").strip(),
                competition_year=request.form.get("competition_year", "").strip(),
                start_date=parse_date(request.form.get("start_date")),
                end_date=parse_date(request.form.get("end_date")),
                location=request.form.get("location", "").strip(),
                logo=save_upload(request.files.get("logo")),
                theme_color=request.form.get("theme_color", "#4f46e5"),
                status=request.form.get("status", "draft"),
            )
            if not event.name:
                flash("กรุณากรอกชื่องานแข่งขัน", "danger")
                return render_template("events/form.html", event=None, orgs=orgs)
            db.session.add(event)
            db.session.commit()
            flash("สร้างงานแข่งขันแล้ว", "success")
            return redirect(url_for("events"))
        return render_template("events/form.html", event=None, orgs=orgs)

    @app.route("/events/<int:event_id>/edit", methods=["GET", "POST"])
    @login_required
    def event_edit(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        orgs = Organization.query.filter(Organization.id.in_(user_org_ids() or [event.organization_id])).all() if not current_user.is_superadmin else Organization.query.all()
        if request.method == "POST":
            event.name = request.form.get("name", "").strip()
            event.competition_year = request.form.get("competition_year", "").strip()
            event.start_date = parse_date(request.form.get("start_date"))
            event.end_date = parse_date(request.form.get("end_date"))
            event.location = request.form.get("location", "").strip()
            event.theme_color = request.form.get("theme_color", "#4f46e5")
            event.status = request.form.get("status", "draft")
            uploaded = save_upload(request.files.get("logo"))
            if uploaded:
                event.logo = uploaded
            db.session.commit()
            flash("บันทึกงานแข่งขันแล้ว", "success")
            return redirect(url_for("events"))
        return render_template("events/form.html", event=event, orgs=orgs)

    @app.route("/events/<int:event_id>")
    @login_required
    def event_detail(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        session["active_event_id"] = event.id
        session["active_org_id"] = event.organization_id
        teams = Team.query.filter_by(event_id=event.id).order_by(Team.created_at.desc()).all()
        sport_count = Sport.query.filter_by(event_id=event.id).count()
        division_count = SportDivision.query.join(Sport).filter(Sport.event_id == event.id).count()

        competition_cards = []

        def match_done(match):
            status = getattr(match, "status", "")
            if status in ("finished", "completed"):
                return True
            winner_team_id = getattr(match, "winner_team_id", None)
            if winner_team_id:
                return True
            score_a = getattr(match, "score_a", None)
            score_b = getattr(match, "score_b", None)
            set_a = getattr(match, "set_a", None)
            set_b = getattr(match, "set_b", None)
            return (score_a is not None and score_b is not None) or (set_a is not None and set_b is not None)

        rr_competitions = RoundRobinCompetition.query.filter_by(event_id=event.id).order_by(RoundRobinCompetition.created_at.desc()).all()
        for comp in rr_competitions:
            total_matches = len(comp.matches or [])
            completed_matches = sum(1 for m in (comp.matches or []) if match_done(m))
            group_count = len(comp.groups or [])
            competition_cards.append({
                "kind": "แบ่งกลุ่ม / พบกันหมด",
                "badge": "Round Robin",
                "icon": "bi-diagram-3",
                "name": comp.name,
                "status": comp.status,
                "sport": comp.sport_division.label if comp.sport_division else "ไม่ระบุชนิดกีฬา",
                "summary": f"{group_count or 1} กลุ่ม · แข่งแล้ว {completed_matches}/{total_matches} คู่",
                "url": url_for("rr_detail", comp_id=comp.id),
                "manage_label": "จัดการผล / ตารางกลุ่ม",
                "created_at": comp.created_at,
            })

        knockout_competitions = KnockoutCompetition.query.filter_by(event_id=event.id).order_by(KnockoutCompetition.created_at.desc()).all()
        for comp in knockout_competitions:
            total_matches = len(comp.matches or [])
            completed_matches = sum(1 for m in (comp.matches or []) if match_done(m))
            competition_cards.append({
                "kind": "น็อคเอาท์",
                "badge": "Knockout",
                "icon": "bi-trophy",
                "name": comp.name,
                "status": comp.status,
                "sport": comp.sport_division.label if comp.sport_division else "ไม่ระบุชนิดกีฬา",
                "summary": f"แข่งแล้ว {completed_matches}/{total_matches} คู่",
                "url": url_for("knockout_detail", comp_id=comp.id),
                "manage_label": "จัดการสาย / บันทึกผล",
                "created_at": comp.created_at,
            })

        ranking_competitions = RankingCompetition.query.filter_by(event_id=event.id).order_by(RankingCompetition.created_at.desc()).all()
        for comp in ranking_competitions:
            result_count = len(comp.results or [])
            competition_cards.append({
                "kind": "จัดอันดับ / เก็บเวลา / เก็บคะแนน",
                "badge": "Ranking",
                "icon": "bi-list-ol",
                "name": comp.name,
                "status": comp.status,
                "sport": comp.sport_division.label if comp.sport_division else "ไม่ระบุชนิดกีฬา",
                "summary": f"บันทึกผลแล้ว {result_count} รายการ",
                "url": url_for("ranking_detail", comp_id=comp.id),
                "manage_label": "จัดอันดับ / บันทึกผล",
                "created_at": comp.created_at,
            })

        contest_competitions = ContestCompetition.query.filter_by(event_id=event.id).order_by(ContestCompetition.created_at.desc()).all()
        for comp in contest_competitions:
            result_count = len(comp.results or [])
            competition_cards.append({
                "kind": "กิจกรรมประกวด",
                "badge": "Contest",
                "icon": "bi-star",
                "name": comp.name,
                "status": comp.status,
                "sport": comp.sport_division.label if comp.sport_division else comp.activity_type or "กิจกรรมประกวด",
                "summary": f"กรรมการ {len(comp.judges or [])} คน · สรุปผลแล้ว {result_count} ทีม",
                "url": url_for("contest_detail", comp_id=comp.id),
                "manage_label": "ให้คะแนน / สรุปผล",
                "created_at": comp.created_at,
            })

        competition_cards.sort(key=lambda x: x.get("created_at") or datetime.min, reverse=True)
        competition_count = len(competition_cards)

        return render_template(
            "events/detail.html",
            event=event,
            teams=teams,
            sport_count=sport_count,
            division_count=division_count,
            competition_cards=competition_cards,
            competition_count=competition_count,
        )

    @app.route("/events/<int:event_id>/status", methods=["POST"])
    @login_required
    def event_status(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขงานนี้", "danger")
            return redirect(url_for("events"))
        event.status = request.form.get("status", event.status)
        db.session.commit()
        flash("เปลี่ยนสถานะงานแข่งขันแล้ว", "success")
        return redirect(url_for("event_detail", event_id=event.id))

    @app.route("/events/<int:event_id>/delete", methods=["POST"])
    @login_required
    def event_delete(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบงานนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(event)
        db.session.commit()
        flash("ลบงานแข่งขันแล้ว", "info")
        return redirect(url_for("events"))

    @app.route("/events/<int:event_id>/teams/new", methods=["GET", "POST"])
    @login_required
    def team_new(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดการทีมในงานนี้", "danger")
            return redirect(url_for("events"))
        if request.method == "POST":
            ok, msg = check_team_limit(event)
            if not ok:
                return deny_upgrade(msg, "organizations_billing", org_id=event.organization_id)
            team = Team(
                event_id=event.id,
                name=request.form.get("name", "").strip(),
                color_name=request.form.get("color_name", "").strip(),
                color_hex=request.form.get("color_hex", "#ef4444"),
                logo=save_upload(request.files.get("logo")),
                flag=save_upload(request.files.get("flag")),
                motto=request.form.get("motto", "").strip(),
                access_code=(request.form.get("access_code", "").strip().upper() or generate_team_code(event.id)),
                line_contact_name=request.form.get("line_contact_name", "").strip(),
                line_user_id=request.form.get("line_user_id", "").strip(),
                registration_open=bool(request.form.get("registration_open")),
            )
            if not team.name:
                flash("กรุณากรอกชื่อทีม/สี", "danger")
                return render_template("teams/form.html", event=event, team=None)
            if Team.query.filter_by(event_id=event.id, access_code=team.access_code).first():
                flash("รหัสทีมนี้ถูกใช้ในงานนี้แล้ว", "danger")
                return render_template("teams/form.html", event=event, team=team)
            db.session.add(team)
            db.session.commit()
            flash("สร้างทีม/สีเรียบร้อย", "success")
            return redirect(url_for("event_detail", event_id=event.id))
        return render_template("teams/form.html", event=event, team=None)
    
    @app.route("/events/<int:event_id>/reports")
    @login_required
    def event_reports(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงรายงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_reports_pdf", "Reports PDF")
        if denied:
            return denied

        teams = Team.query.filter_by(event_id=event.id).order_by(Team.name).all()
        athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Athlete.full_name).all()
        coaches = Coach.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Coach.full_name).all()
        rr_comps = RoundRobinCompetition.query.filter_by(event_id=event.id).order_by(RoundRobinCompetition.created_at.desc()).all()
        ranking_comps = RankingCompetition.query.filter_by(event_id=event.id).order_by(RankingCompetition.created_at.desc()).all()
        contest_comps = ContestCompetition.query.filter_by(event_id=event.id).order_by(ContestCompetition.created_at.desc()).all()

        souvenir_url = url_for("event_souvenir_public", event_id=event.id, _external=True)
        qr_data = make_qr_data_uri(souvenir_url)

        return render_template(
            "reports/index.html",
            event=event,
            teams=teams,
            athletes=athletes,
            coaches=coaches,
            rr_comps=rr_comps,
            ranking_comps=ranking_comps,
            contest_comps=contest_comps,
            souvenir_url=souvenir_url,
            qr_data=qr_data,
        )


    @app.route("/events/<int:event_id>/souvenir")
    @login_required
    def event_souvenir_admin(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงสูจิบัตรนี้", "danger")
            return redirect(url_for("events"))
        return redirect(url_for("event_souvenir_public", event_id=event.id))


    @app.route("/events/<int:event_id>/souvenir/public")
    def event_souvenir_public(event_id):
        event = Event.query.get_or_404(event_id)
        if not feature_allowed(event.organization, "allow_reports_pdf"):
            return make_response("แพ็กเกจปัจจุบันยังไม่มีสิทธิ์ใช้ Reports PDF กรุณาอัปเกรดแพ็กเกจ", 403)

        teams = Team.query.filter_by(event_id=event.id).order_by(Team.name).all()
        athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Athlete.full_name).all()
        coaches = Coach.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Coach.full_name).all()

        rr_comps = RoundRobinCompetition.query.filter_by(event_id=event.id).order_by(RoundRobinCompetition.created_at.desc()).all()
        ranking_comps = RankingCompetition.query.filter_by(event_id=event.id).order_by(RankingCompetition.created_at.desc()).all()
        contest_comps = ContestCompetition.query.filter_by(event_id=event.id).order_by(ContestCompetition.created_at.desc()).all()

        medal_rows = build_event_medal_rows(event)

        souvenir_url = url_for("event_souvenir_public", event_id=event.id, _external=True)
        qr_data = make_qr_data_uri(souvenir_url)

        return render_template(
            "reports/souvenir.html",
            event=event,
            teams=teams,
            athletes=athletes,
            coaches=coaches,
            rr_comps=rr_comps,
            ranking_comps=ranking_comps,
            contest_comps=contest_comps,
            medal_rows=medal_rows,
            souvenir_url=souvenir_url,
            qr_data=qr_data,
        )


    @app.route("/events/<int:event_id>/reports/athletes.xlsx")
    @login_required
    def report_athletes_excel(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ Export รายงานนี้", "danger")
            return redirect(url_for("events"))

        athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Athlete.full_name).all()

        rows = []
        for a in athletes:
            regs = ", ".join([f"{r.sport_name} {r.category_name or ''} {r.gender or ''}" for r in a.registrations])
            rows.append([
                a.team.name if a.team else "",
                a.full_name,
                a.gender,
                a.grade_level,
                a.classroom,
                a.student_no,
                a.phone,
                a.status,
                regs,
            ])

        output = build_report_workbook(
            "รายชื่อนักกีฬา",
            ["ทีม", "ชื่อ-สกุล", "เพศ", "ชั้น", "ห้อง", "เลขประจำตัว", "เบอร์โทร", "สถานะ", "รายการที่สมัคร"],
            rows,
        )
        return send_file(output, as_attachment=True, download_name=f"athletes_event_{event.id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    @app.route("/events/<int:event_id>/reports/teams.xlsx")
    @login_required
    def report_teams_excel(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ Export รายงานนี้", "danger")
            return redirect(url_for("events"))

        rows = []
        teams = Team.query.filter_by(event_id=event.id).order_by(Team.name).all()
        for t in teams:
            p = t.profile
            rows.append([
                t.name,
                t.color_name,
                t.motto,
                t.access_code,
                t.line_contact_name or "",
                t.line_user_id or "",
                "เปิด" if t.registration_open else "ปิด",
                p.director_name if p else "",
                p.deputy_directors if p else "",
                p.advisors if p else "",
                p.coaches_summary if p else "",
                p.parade_title if p else "",
                p.stand_member_total if p else 0,
                p.cheerleader_summary if p else "",
            ])

        output = build_report_workbook(
            "รายงานทีม",
            ["ทีม", "สี", "คำขวัญ", "รหัสทีม", "ชื่อหัวหน้าสี", "LINE userId", "สถานะกรอก", "ผอ.", "รอง ผอ.", "ครูที่ปรึกษา", "ผู้ฝึกสอน", "ขบวน", "จำนวนสแตนด์", "เชียร์ลีดเดอร์"],
            rows,
        )
        return send_file(output, as_attachment=True, download_name=f"teams_event_{event.id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    @app.route("/events/<int:event_id>/reports/medals.xlsx")
    @login_required
    def report_medals_excel(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ Export รายงานนี้", "danger")
            return redirect(url_for("events"))

        medal_rows = build_event_medal_rows(event)
        rows = [[r["team"].name, r["gold"], r["silver"], r["bronze"]] for r in medal_rows]

        output = build_report_workbook(
            "ตารางเหรียญ",
            ["ทีม", "ทอง", "เงิน", "ทองแดง"],
            rows,
        )
        return send_file(output, as_attachment=True, download_name=f"medals_event_{event.id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/teams/<int:team_id>/edit", methods=["GET", "POST"])
    @login_required
    def team_edit(team_id):
        team = Team.query.get_or_404(team_id)
        event = team.event
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขทีมนี้", "danger")
            return redirect(url_for("events"))
        if request.method == "POST":
            old_code = team.access_code
            new_code = (request.form.get("access_code", "").strip().upper() or old_code)
            exists = Team.query.filter(Team.event_id == event.id, Team.access_code == new_code, Team.id != team.id).first()
            if exists:
                flash("รหัสทีมนี้ถูกใช้ในงานนี้แล้ว", "danger")
                return render_template("teams/form.html", event=event, team=team)
            team.name = request.form.get("name", "").strip()
            team.color_name = request.form.get("color_name", "").strip()
            team.color_hex = request.form.get("color_hex", "#ef4444")
            team.motto = request.form.get("motto", "").strip()
            team.access_code = new_code
            team.line_contact_name = request.form.get("line_contact_name", "").strip()
            team.line_user_id = request.form.get("line_user_id", "").strip()
            team.registration_open = bool(request.form.get("registration_open"))
            uploaded_logo = save_upload(request.files.get("logo"))
            uploaded_flag = save_upload(request.files.get("flag"))
            if uploaded_logo:
                team.logo = uploaded_logo
            if uploaded_flag:
                team.flag = uploaded_flag
            if not team.name:
                flash("กรุณากรอกชื่อทีม/สี", "danger")
                return render_template("teams/form.html", event=event, team=team)
            db.session.commit()
            flash("บันทึกข้อมูลทีม/สีแล้ว", "success")
            return redirect(url_for("event_detail", event_id=event.id))
        return render_template("teams/form.html", event=event, team=team)

    @app.route("/teams/<int:team_id>/line-invite", methods=["POST"])
    @login_required
    def team_line_invite(team_id):
        team = Team.query.get_or_404(team_id)
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์ส่งข้อความทีมนี้", "danger")
            return redirect(url_for("events"))
        message = build_team_invite_message(team)
        ok, msg = send_line_push_message(team.line_user_id, message)
        if ok:
            team.line_invite_sent_at = datetime.utcnow()
            team.line_invite_error = None
            flash(f"ส่งข้อความ LINE ให้ {team.name} แล้ว", "success")
        else:
            team.line_invite_error = msg
            flash(msg, "danger")
        db.session.commit()
        return redirect(url_for("event_detail", event_id=team.event_id, _anchor="teams-section"))

    @app.route("/events/<int:event_id>/teams/line-invite-all", methods=["POST"])
    @login_required
    def event_line_invite_all(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ส่งข้อความในงานนี้", "danger")
            return redirect(url_for("events"))
        teams = Team.query.filter_by(event_id=event.id).order_by(Team.name).all()
        sent = 0
        failed = 0
        skipped = 0
        errors = []
        for team in teams:
            if not team.line_user_id:
                skipped += 1
                continue
            ok, msg = send_line_push_message(team.line_user_id, build_team_invite_message(team))
            if ok:
                sent += 1
                team.line_invite_sent_at = datetime.utcnow()
                team.line_invite_error = None
            else:
                failed += 1
                team.line_invite_error = msg
                errors.append(f"{team.name}: {msg}")
        db.session.commit()
        if sent:
            flash(f"ส่ง LINE สำเร็จ {sent} ทีม/สี", "success")
        if skipped:
            flash(f"ข้าม {skipped} ทีม/สี เพราะยังไม่ได้กรอก LINE userId", "warning")
        if failed:
            flash(f"ส่งไม่สำเร็จ {failed} ทีม/สี: " + " | ".join(errors[:3]), "danger")
        return redirect(url_for("event_detail", event_id=event.id, _anchor="teams-section"))

    @app.route("/teams/<int:team_id>/toggle", methods=["POST"])
    @login_required
    def team_toggle(team_id):
        team = Team.query.get_or_404(team_id)
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขทีมนี้", "danger")
            return redirect(url_for("events"))
        team.registration_open = not team.registration_open
        db.session.commit()
        flash("เปลี่ยนสถานะการกรอกข้อมูลของทีมแล้ว", "success")
        return redirect(url_for("event_detail", event_id=team.event_id))

    @app.route("/teams/<int:team_id>/delete", methods=["POST"])
    @login_required
    def team_delete(team_id):
        team = Team.query.get_or_404(team_id)
        event_id = team.event_id
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบทีมนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(team)
        db.session.commit()
        flash("ลบทีม/สีแล้ว", "info")
        return redirect(url_for("event_detail", event_id=event_id))



    @app.route("/teams/<int:team_id>/profile", methods=["GET", "POST"])
    @login_required
    def team_profile(team_id):
        team = Team.query.get_or_404(team_id)
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขข้อมูลทีมนี้", "danger")
            return redirect(url_for("events"))
        profile = get_or_create_team_profile(team)
        if request.method == "POST":
            save_team_profile_from_request(profile)
            db.session.commit()
            flash("บันทึกข้อมูลทีมสำหรับสูจิบัตรแล้ว", "success")
            return redirect(url_for("team_profile", team_id=team.id))
        return render_team_profile_page(team, profile, public_mode=False)

    @app.route("/teams/<int:team_id>/people/add", methods=["POST"])
    @login_required
    def team_person_add(team_id):
        team = Team.query.get_or_404(team_id)
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขข้อมูลทีมนี้", "danger")
            return redirect(url_for("events"))
        add_team_person(team)
        return redirect(url_for("team_profile", team_id=team.id))

    @app.route("/teams/people/<int:person_id>/delete", methods=["POST"])
    @login_required
    def team_person_delete(person_id):
        person = TeamPerson.query.get_or_404(person_id)
        team = person.team
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(person)
        db.session.commit()
        flash("ลบรายชื่อแล้ว", "info")
        return redirect(url_for("team_profile", team_id=team.id))

    @app.route("/teams/<int:team_id>/files/add", methods=["POST"])
    @login_required
    def team_file_add(team_id):
        team = Team.query.get_or_404(team_id)
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แนบไฟล์ทีมนี้", "danger")
            return redirect(url_for("events"))
        add_team_file(team)
        return redirect(url_for("team_profile", team_id=team.id))

    @app.route("/teams/files/<int:file_id>/delete", methods=["POST"])
    @login_required
    def team_file_delete(file_id):
        item = TeamFile.query.get_or_404(file_id)
        team = item.team
        if not can_access_org(team.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบไฟล์นี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(item)
        db.session.commit()
        flash("ลบไฟล์แนบแล้ว", "info")
        return redirect(url_for("team_profile", team_id=team.id))

    @app.route("/team-entry", methods=["GET", "POST"])
    def team_entry():
        if request.method == "POST":
            code = request.form.get("access_code", "").strip().upper()
            team = Team.query.filter_by(access_code=code).first()
            if not team:
                flash("ไม่พบรหัสทีมนี้", "danger")
                return render_template("team_portal/entry.html")
            if not team.registration_open:
                flash("ทีมนี้ถูกปิดสิทธิ์กรอกข้อมูลแล้ว", "warning")
                return render_template("team_portal/entry.html")
            session[f"team_access_{team.id}"] = team.access_code
            return redirect(url_for("team_portal_profile", team_id=team.id))
        return render_template("team_portal/entry.html")

    @app.route("/team-portal/<int:team_id>", methods=["GET", "POST"])
    def team_portal_profile(team_id):
        team = Team.query.get_or_404(team_id)
        if session.get(f"team_access_{team.id}") != team.access_code:
            flash("กรุณากรอกรหัสทีมก่อน", "warning")
            return redirect(url_for("team_entry"))
        if not team.registration_open:
            flash("ทีมนี้ถูกปิดสิทธิ์กรอกข้อมูลแล้ว", "warning")
            return redirect(url_for("team_entry"))
        profile = get_or_create_team_profile(team)
        if request.method == "POST":
            save_team_profile_from_request(profile)
            db.session.commit()
            flash("บันทึกข้อมูลทีมแล้ว", "success")
            return redirect(url_for("team_portal_profile", team_id=team.id))
        return render_team_profile_page(team, profile, public_mode=True)

    @app.route("/team-portal/<int:team_id>/people/add", methods=["POST"])
    def portal_person_add(team_id):
        team = require_team_portal_access(team_id)
        if not team:
            return redirect(url_for("team_entry"))
        add_team_person(team)
        return redirect(url_for("team_portal_profile", team_id=team.id))

    @app.route("/team-portal/people/<int:person_id>/delete", methods=["POST"])
    def portal_person_delete(person_id):
        person = TeamPerson.query.get_or_404(person_id)
        team = require_team_portal_access(person.team_id)
        if not team:
            return redirect(url_for("team_entry"))
        db.session.delete(person)
        db.session.commit()
        flash("ลบรายชื่อแล้ว", "info")
        return redirect(url_for("team_portal_profile", team_id=team.id))

    @app.route("/team-portal/<int:team_id>/files/add", methods=["POST"])
    def portal_file_add(team_id):
        team = require_team_portal_access(team_id)
        if not team:
            return redirect(url_for("team_entry"))
        add_team_file(team)
        return redirect(url_for("team_portal_profile", team_id=team.id))

    @app.route("/team-portal/files/<int:file_id>/delete", methods=["POST"])
    def portal_file_delete(file_id):
        item = TeamFile.query.get_or_404(file_id)
        team = require_team_portal_access(item.team_id)
        if not team:
            return redirect(url_for("team_entry"))
        db.session.delete(item)
        db.session.commit()
        flash("ลบไฟล์แนบแล้ว", "info")
        return redirect(url_for("team_portal_profile", team_id=team.id))


    @app.route("/teams")
    @login_required
    def teams_home():
        """หน้าเลือกงานสำหรับจัดการ Teams/Colors เมื่อยังไม่มี active event"""
        eid = session.get("active_event_id")
        if eid:
            ev = Event.query.get(int(eid))
            if ev and can_access_org(ev.organization_id):
                return redirect(url_for("event_detail", event_id=ev.id) + "#teams-section")
        org_ids = user_org_ids()
        query = Event.query
        if org_ids is not None:
            query = query.filter(Event.organization_id.in_(org_ids or [0]))
        events_list = query.order_by(Event.created_at.desc()).all()
        return render_template("events/select_context.html", mode="teams", events=events_list, title="เลือกงานเพื่อจัดการ Teams / Colors")

    @app.route("/sports")
    @login_required
    def sport_setup_home():
        """หน้าเลือกงานสำหรับ Sport Setup ไม่ให้กดแล้วค้างหน้า Settings"""
        eid = session.get("active_event_id")
        if eid:
            ev = Event.query.get(int(eid))
            if ev and can_access_org(ev.organization_id):
                return redirect(url_for("event_sports", event_id=ev.id))
        org_ids = user_org_ids()
        query = Event.query
        if org_ids is not None:
            query = query.filter(Event.organization_id.in_(org_ids or [0]))
        events_list = query.order_by(Event.created_at.desc()).all()
        return render_template("events/select_context.html", mode="sports", events=events_list, title="เลือกงานเพื่อตั้งค่ากีฬา")

    @app.route("/billing")
    @login_required
    def billing_home():
        oid = session.get("active_org_id")
        if oid and can_access_org(int(oid)):
            return redirect(url_for("organizations_billing", org_id=int(oid)))
        org_ids = user_org_ids()
        query = Organization.query
        if org_ids is not None:
            query = query.filter(Organization.id.in_(org_ids or [0]))
        orgs = query.order_by(Organization.created_at.desc()).all()
        return render_template("orgs/select_billing.html", organizations=orgs)

    @app.route("/events/<int:event_id>/sports")
    @login_required
    def event_sports(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        session["active_event_id"] = event.id
        session["active_org_id"] = event.organization_id
        categories = SportCategory.query.filter_by(event_id=event.id).order_by(SportCategory.sort_order, SportCategory.name).all()
        sports = Sport.query.filter_by(event_id=event.id).order_by(Sport.name).all()
        divisions = SportDivision.query.join(Sport).filter(Sport.event_id == event.id).order_by(Sport.name, SportDivision.class_name, SportDivision.gender).all()
        return render_template("sports/setup.html", event=event, categories=categories, sports=sports, divisions=divisions)

    @app.route("/events/<int:event_id>/competitions/quick-new", methods=["GET", "POST"])
    @login_required
    def event_competition_wizard(event_id):
        """สร้างรายการแข่งขันจากมุมผู้ใช้จริง
        Flow: เลือกกีฬา -> เลือกทีมที่ส่งจริง -> เลือกระบบ/กติกา -> สร้างตาราง
        ไม่บังคับให้ต้องสร้างชนิดกีฬา/รุ่นย่อยล่วงหน้า
        """
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์สร้างรายการในงานนี้", "danger")
            return redirect(url_for("events"))

        session["active_event_id"] = event.id
        session["active_org_id"] = event.organization_id

        teams = Team.query.filter_by(event_id=event.id).order_by(Team.name).all()
        sports = Sport.query.filter_by(event_id=event.id).order_by(Sport.name).all()
        divisions = SportDivision.query.join(Sport).filter(Sport.event_id == event.id).order_by(Sport.name, SportDivision.class_name, SportDivision.gender).all()

        sport_presets = [
            {"name": "ฟุตบอล", "category": "กีฬาทีม", "format": "round_robin", "result": "score_only", "max_sets": 0, "points": 0, "win": 0, "note": "IFAB Laws of the Game 2025/26: บันทึกผลเป็นประตูได้-เสีย ระยะเวลาแข่งขันปรับตามระเบียบงานได้"},
            {"name": "ฟุตซอล", "category": "กีฬาทีม", "format": "round_robin", "result": "score_only", "max_sets": 0, "points": 0, "win": 0, "note": "FIFA Futsal Laws of the Game 2025/26: ปกติ 2 ครึ่ง ครึ่งละ 20 นาที บันทึกผลเป็นประตูได้-เสีย"},
            {"name": "วอลเลย์บอล", "category": "กีฬาทีม", "format": "round_robin", "result": "set_based", "max_sets": 5, "points": 25, "win": 3, "note": "FIVB Official Volleyball Rules 2025-2028: ชนะ 3 เซต; เซต 1-4 ถึง 25 แต้ม ต้องห่าง 2; เซตตัดสินถึง 15 แต้ม"},
            {"name": "เซปักตะกร้อ", "category": "กีฬาทีม", "format": "round_robin", "result": "set_based", "max_sets": 3, "points": 15, "win": 2, "note": "ISTAF Law of the Game 2024: ชนะ 2 ใน 3 เซต; เซตละ 15 แต้ม; 14-14 เล่นถึง 17 แต้ม"},
            {"name": "เปตอง", "category": "กีฬาเฉพาะทาง", "format": "knockout", "result": "score_only", "max_sets": 0, "points": 13, "win": 0, "note": "FIPJP Official Rules: เกมปกติถึง 13 คะแนน; รอบลีก/คัดเลือกอาจกำหนด 11 คะแนนตามระเบียบงาน"},
            {"name": "แบดมินตัน", "category": "กีฬาเฉพาะทาง", "format": "knockout", "result": "set_based", "max_sets": 3, "points": 21, "win": 2, "note": "BWF Laws ปัจจุบัน: 2 ใน 3 เกม เกมละ 21 แต้ม; BWF อนุมัติ 3x15 เริ่ม 4 ม.ค. 2027"},
            {"name": "เทเบิลเทนนิส", "category": "กีฬาเฉพาะทาง", "format": "knockout", "result": "set_based", "max_sets": 5, "points": 11, "win": 3, "note": "ITTF Laws: เกมละ 11 แต้ม ต้องชนะห่าง 2; ค่าเริ่มต้นระบบใช้ 3 ใน 5 เกม"},
            {"name": "บาสเกตบอล", "category": "กีฬาทีม", "format": "round_robin", "result": "score_only", "max_sets": 0, "points": 0, "win": 0, "note": "FIBA ใช้การบันทึกคะแนนรวมของเกม; งานโรงเรียนปรับเวลาควอเตอร์ตามระเบียบงานได้"},
            {"name": "กรีฑา", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "World Athletics: ใช้ผลเวลา/ระยะ/อันดับ ไม่ใช่ระบบเซต"},
            {"name": "วิ่ง 50 เมตร", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา"},
            {"name": "วิ่ง 100 เมตร", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา"},
            {"name": "วิ่งผลัด", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา"},
            {"name": "ชักเย่อ", "category": "กีฬาพื้นบ้าน", "format": "knockout", "result": "set_based", "max_sets": 3, "points": 0, "win": 2, "note": "กีฬาโรงเรียน/กีฬาพื้นบ้าน: ค่าเริ่มต้นชนะ 2 ใน 3 เที่ยว ปรับเองได้ตามระเบียบงาน"},
            {"name": "วิ่งกระสอบ", "category": "กีฬาพื้นบ้าน", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา"},
            {"name": "วิ่งสามขา", "category": "กีฬาพื้นบ้าน", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา"},
            {"name": "ประกวดกองเชียร์", "category": "กิจกรรมประกวด", "format": "score_judging", "result": "contest", "max_sets": 0, "points": 0, "win": 0, "note": "ใช้กรรมการให้คะแนน ปรับเกณฑ์ได้ตามงาน"},
        ]

        def split_names(raw):
            raw = (raw or "").strip()
            for sep in ["\r", "\n", "，", "、", ";", "|", "\t"]:
                raw = raw.replace(sep, ",")
            names = []
            for part in raw.split(","):
                value = part.strip()
                if value and value not in names:
                    names.append(value)
            return names

        def preset_for(name):
            needle = (name or "").strip().lower()
            for item in sport_presets:
                if item["name"].lower() == needle:
                    return item
            return None

        def recommendation_for(sport_name, team_count):
            name = (sport_name or "").lower()
            if "กรีฑา" in name or "วิ่ง" in name:
                return {"format": "ranking", "groups": 1, "advance": 0, "best": 0, "text": "กีฬาวัดเวลา/ระยะ แนะนำใช้ Ranking"}
            if "ประกวด" in name or "กองเชียร์" in name:
                return {"format": "score_judging", "groups": 1, "advance": 0, "best": 0, "text": "กิจกรรมประกวด แนะนำใช้กรรมการให้คะแนน"}
            if team_count <= 1:
                return {"format": "round_robin", "groups": 1, "advance": 0, "best": 0, "text": "ยังมีทีมไม่พอ ต้องเลือกอย่างน้อย 2 ทีม"}
            if team_count == 2:
                return {"format": "knockout", "groups": 1, "advance": 1, "best": 0, "text": "2 ทีม แนะนำแข่งชิงทันทีแบบ Knockout"}
            if team_count <= 5:
                return {"format": "round_robin", "groups": 1, "advance": 1, "best": 0, "text": f"{team_count} ทีม แนะนำพบกันหมด 1 กลุ่ม รู้ผลชัดและยุติธรรม"}
            if team_count <= 8:
                return {"format": "round_robin", "groups": 2, "advance": 2, "best": 0, "text": f"{team_count} ทีม แนะนำแบ่ง 2 กลุ่ม เข้ารอบกลุ่มละ 2 แล้วค่อยสร้าง Knockout"}
            if team_count <= 12:
                return {"format": "round_robin", "groups": 4, "advance": 2, "best": 0, "text": f"{team_count} ทีม แนะนำแบ่ง 4 กลุ่ม เอาที่ 1-2 เข้ารอบ หรือเลือกเอาเฉพาะที่ 1 ได้"}
            return {"format": "round_robin", "groups": 4, "advance": 2, "best": 0, "text": f"{team_count} ทีมขึ้นไป แนะนำแบ่ง 4 กลุ่ม เข้ารอบกลุ่มละ 2"}

        def get_or_create_sport(event, sport_name, preset, result_type, competition_format, max_sets, points_per_set, sets_to_win):
            sport_name = (sport_name or "").strip()
            if not sport_name:
                return None
            category_name = (preset or {}).get("category") or "กีฬาในงาน"
            category = SportCategory.query.filter_by(event_id=event.id, name=category_name).first()
            if not category:
                category = SportCategory(event_id=event.id, name=category_name, description="สร้างจากหน้าสร้างรายการแข่งขันแบบง่าย")
                db.session.add(category)
                db.session.flush()
            sport = Sport.query.filter_by(event_id=event.id, name=sport_name).first()
            if not sport:
                sport = Sport(
                    event_id=event.id,
                    category_id=category.id,
                    name=sport_name,
                    default_format=competition_format,
                    result_type=result_type,
                    max_sets=max_sets if result_type == "set_based" else 0,
                    points_per_set=points_per_set if result_type == "set_based" else 0,
                    sets_to_win=sets_to_win if result_type == "set_based" else 0,
                    note=(preset or {}).get("note", ""),
                    is_active=True,
                )
                db.session.add(sport)
                db.session.flush()
            elif not sport.category_id:
                sport.category_id = category.id
            return sport

        def get_or_create_division(sport, class_name, gender, competition_format, result_type, max_sets, points_per_set, sets_to_win):
            class_name = (class_name or "Open").strip() or "Open"
            gender = (gender or "รวม").strip() or "รวม"
            division = SportDivision.query.filter_by(sport_id=sport.id, class_name=class_name, gender=gender).first()
            if not division:
                division = SportDivision(
                    sport_id=sport.id,
                    class_name=class_name,
                    gender=gender,
                    competition_format=competition_format,
                    result_type=result_type,
                    max_sets=max_sets,
                    points_per_set=points_per_set,
                    sets_to_win=sets_to_win,
                    is_active=True,
                )
                db.session.add(division)
                db.session.flush()
            else:
                division.competition_format = competition_format
                division.result_type = result_type
                division.max_sets = max_sets
                division.points_per_set = points_per_set
                division.sets_to_win = sets_to_win
                division.is_active = True
            return division

        def standard_tiebreakers(result_type):
            if result_type == "set_based":
                return "points,set_diff,sets_for,point_diff,head_to_head,wins,draw_lots"
            return "points,goal_diff,goals_for,head_to_head,wins,draw_lots"

        def create_competition_from_sport(sport, selected_teams, class_name, gender, competition_format, result_type, max_sets, points_per_set, sets_to_win, num_groups=None, advance_per_group=None, best_runnerup_count=None, name=None, bracket_pairing="adjacent"):
            rec = recommendation_for(sport.name, len(selected_teams))
            competition_format = competition_format or sport.default_format or rec["format"]
            result_type = result_type or sport.result_type or "score_only"
            division = get_or_create_division(sport, class_name, gender, competition_format, result_type, max_sets, points_per_set, sets_to_win)
            comp_name = (name or f"{sport.name} {division.class_name} {division.gender}").strip()

            if competition_format == "knockout":
                comp = KnockoutCompetition(
                    event_id=event.id,
                    sport_division_id=division.id,
                    name=comp_name,
                    result_type=result_type,
                    max_sets=max_sets,
                    points_per_set=points_per_set,
                    sets_to_win=sets_to_win,
                    status="scheduled",
                )
                db.session.add(comp)
                db.session.flush()
                create_knockout_first_round(comp, selected_teams, pairing_mode=bracket_pairing or "adjacent")
                return "knockout", comp

            if competition_format == "round_robin":
                comp = RoundRobinCompetition(
                    event_id=event.id,
                    sport_division_id=division.id,
                    name=comp_name,
                    num_groups=max(1, safe_int(num_groups) or rec["groups"]),
                    win_points=3,
                    draw_points=1,
                    loss_points=0,
                    advance_per_group=max(0, safe_int(advance_per_group) if advance_per_group not in (None, "") else rec["advance"]),
                    best_runnerup_count=max(0, safe_int(best_runnerup_count) if best_runnerup_count not in (None, "") else rec["best"]),
                    tiebreakers=standard_tiebreakers(result_type),
                    status="scheduled",
                )
                db.session.add(comp)
                db.session.flush()
                create_rr_groups(comp)
                assign_teams_auto(comp, [t.id for t in selected_teams])
                generate_rr_matches(comp)
                return "round_robin", comp

            if competition_format == "score_judging":
                comp = ContestCompetition(event_id=event.id, sport_division_id=division.id, name=comp_name, activity_type=sport.name, status="draft")
                db.session.add(comp)
                db.session.flush()
                for idx, cname in enumerate(["ความพร้อม", "ทักษะ/คุณภาพ", "ความคิดสร้างสรรค์", "ความพร้อมเพรียง", "ภาพรวม"], start=1):
                    db.session.add(ContestCriterion(competition_id=comp.id, name=cname, max_score=100, sort_order=idx))
                db.session.add(ContestJudge(competition_id=comp.id, name="กรรมการ 1"))
                return "score_judging", comp

            comp = RankingCompetition(event_id=event.id, sport_division_id=division.id, name=comp_name, result_mode="rank", status="draft")
            db.session.add(comp)
            return "ranking", comp

        if request.method == "POST" and request.form.get("mode") == "bulk_from_library":
            selected_sport_ids = [safe_int(x) for x in request.form.getlist("selected_sports") if safe_int(x)]
            if not selected_sport_ids:
                flash("กรุณาเลือกชนิดกีฬาที่งานนี้จะจัดอย่างน้อย 1 รายการ", "warning")
                return redirect(url_for("event_competition_wizard", event_id=event.id))

            all_team_ids = [t.id for t in teams]
            created = []
            skipped = []
            first_redirect = None

            for sport_id in selected_sport_ids:
                sport = Sport.query.filter_by(id=sport_id, event_id=event.id).first()
                if not sport:
                    continue
                team_ids = [safe_int(x) for x in request.form.getlist(f"team_ids_{sport_id}") if safe_int(x)]
                if not team_ids and request.form.get(f"use_all_teams_{sport_id}") == "1":
                    team_ids = all_team_ids
                selected_teams = Team.query.filter(Team.event_id == event.id, Team.id.in_(team_ids)).order_by(Team.name).all() if team_ids else []

                competition_format = request.form.get(f"competition_format_{sport_id}") or sport.default_format or recommendation_for(sport.name, len(selected_teams))["format"]
                result_type = request.form.get(f"result_type_{sport_id}") or sport.result_type or "score_only"
                if len(selected_teams) < 2 and competition_format not in ("ranking", "score_judging"):
                    skipped.append(f"{sport.name} (ทีมไม่พอ)")
                    continue

                if result_type == "set_based":
                    max_sets = safe_int(request.form.get(f"max_sets_{sport_id}")) or sport.max_sets or 3
                    points_per_set = safe_int(request.form.get(f"points_per_set_{sport_id}")) or sport.points_per_set or 21
                    sets_to_win = safe_int(request.form.get(f"sets_to_win_{sport_id}")) or sport.sets_to_win or ((max_sets // 2) + 1)
                else:
                    max_sets = 0
                    points_per_set = 0
                    sets_to_win = 0

                class_name = request.form.get(f"class_name_{sport_id}") or "Open"
                gender = request.form.get(f"gender_{sport_id}") or "รวม"
                name = request.form.get(f"name_{sport_id}") or None
                kind, comp = create_competition_from_sport(
                    sport=sport,
                    selected_teams=selected_teams,
                    class_name=class_name,
                    gender=gender,
                    competition_format=competition_format,
                    result_type=result_type,
                    max_sets=max_sets,
                    points_per_set=points_per_set,
                    sets_to_win=sets_to_win,
                    num_groups=request.form.get(f"num_groups_{sport_id}"),
                    advance_per_group=request.form.get(f"advance_per_group_{sport_id}"),
                    best_runnerup_count=request.form.get(f"best_runnerup_count_{sport_id}"),
                    name=name,
                    bracket_pairing=request.form.get(f"bracket_pairing_{sport_id}") or "adjacent",
                )
                created.append(f"{sport.name} ({len(selected_teams)} ทีม)")
                if first_redirect is None:
                    first_redirect = (kind, comp)

            db.session.commit()
            if created:
                flash("สร้างการแข่งขันแล้ว: " + ", ".join(created), "success")
            if skipped:
                flash("ข้ามรายการ: " + ", ".join(skipped), "warning")
            if first_redirect and len(created) == 1:
                kind, comp = first_redirect
                if kind == "knockout":
                    return redirect(url_for("knockout_detail", comp_id=comp.id))
                if kind == "round_robin":
                    return redirect(url_for("rr_detail", comp_id=comp.id))
                if kind == "score_judging":
                    return redirect(url_for("contest_detail", comp_id=comp.id))
                return redirect(url_for("ranking_detail", comp_id=comp.id))
            return redirect(url_for("event_detail", event_id=event.id))

        if request.method == "POST":
            selected_team_ids = [int(x) for x in request.form.getlist("team_ids") if str(x).isdigit()]
            selected_teams = Team.query.filter(Team.event_id == event.id, Team.id.in_(selected_team_ids)).order_by(Team.name).all() if selected_team_ids else []
            if len(selected_teams) < 2 and request.form.get("competition_format") not in ("ranking", "score_judging"):
                flash("กรุณาเลือกทีมที่ส่งแข่งขันจริงอย่างน้อย 2 ทีม", "danger")
                return redirect(url_for("event_competition_wizard", event_id=event.id))

            sport_id = safe_int_or_none(request.form.get("sport_id"))
            sport = Sport.query.filter_by(id=sport_id, event_id=event.id).first() if sport_id else None
            sport_name = (request.form.get("sport_name") or (sport.name if sport else "")).strip()
            preset = preset_for(sport_name) or {}
            class_name = (request.form.get("class_name") or "Open").strip()
            gender = (request.form.get("gender") or "ผสม").strip()
            competition_format = request.form.get("competition_format") or recommendation_for(sport_name, len(selected_teams))["format"]
            result_type = request.form.get("result_type") or (sport.result_type if sport else preset.get("result", "score_only"))

            if result_type == "set_based":
                max_sets = safe_int(request.form.get("max_sets")) or preset.get("max_sets") or 3
                points_per_set = safe_int(request.form.get("points_per_set")) or preset.get("points") or 21
                sets_to_win = safe_int(request.form.get("sets_to_win")) or preset.get("win") or ((max_sets // 2) + 1)
            else:
                max_sets = 0
                points_per_set = 0
                sets_to_win = 0

            if not sport:
                sport = get_or_create_sport(event, sport_name, preset, result_type, competition_format, max_sets, points_per_set, sets_to_win)
            if not sport:
                flash("กรุณาเลือกหรือกรอกชื่อกีฬา", "danger")
                return redirect(url_for("event_competition_wizard", event_id=event.id))

            division = SportDivision.query.filter_by(sport_id=sport.id, class_name=class_name, gender=gender).first()
            if not division:
                division = SportDivision(
                    sport_id=sport.id,
                    class_name=class_name,
                    gender=gender,
                    competition_format=competition_format,
                    result_type=result_type,
                    max_sets=max_sets,
                    points_per_set=points_per_set,
                    sets_to_win=sets_to_win,
                    is_active=True,
                )
                db.session.add(division)
                db.session.flush()
            else:
                division.competition_format = competition_format
                division.result_type = result_type
                division.max_sets = max_sets
                division.points_per_set = points_per_set
                division.sets_to_win = sets_to_win
                division.is_active = True

            default_name = f"{sport.name} {class_name} {gender}".strip()
            comp_name = (request.form.get("name") or default_name).strip()

            if competition_format == "knockout":
                comp = KnockoutCompetition(
                    event_id=event.id,
                    sport_division_id=division.id,
                    name=comp_name,
                    result_type=result_type,
                    max_sets=max_sets,
                    points_per_set=points_per_set,
                    sets_to_win=sets_to_win,
                    status="scheduled",
                )
                db.session.add(comp)
                db.session.flush()
                create_knockout_first_round(comp, selected_teams, pairing_mode=request.form.get("bracket_pairing") or "adjacent")
                db.session.commit()
                flash(f"สร้าง {comp.name} แบบ Knockout จากทีมที่เลือก {len(selected_teams)} ทีมแล้ว", "success")
                return redirect(url_for("knockout_detail", comp_id=comp.id))

            if competition_format == "round_robin":
                rec = recommendation_for(sport.name, len(selected_teams))
                selected_tiebreakers = ",".join(request.form.getlist("tiebreakers"))
                if not selected_tiebreakers:
                    selected_tiebreakers = "points,set_diff,sets_for,point_diff,head_to_head,wins,draw_lots" if result_type == "set_based" else "points,goal_diff,goals_for,head_to_head,wins,draw_lots"
                comp = RoundRobinCompetition(
                    event_id=event.id,
                    sport_division_id=division.id,
                    name=comp_name,
                    num_groups=max(1, safe_int(request.form.get("num_groups")) or rec["groups"]),
                    win_points=safe_int(request.form.get("win_points")) if request.form.get("win_points") not in (None, "") else 3,
                    draw_points=safe_int(request.form.get("draw_points")) if request.form.get("draw_points") not in (None, "") else 1,
                    loss_points=safe_int(request.form.get("loss_points")) if request.form.get("loss_points") not in (None, "") else 0,
                    advance_per_group=max(0, safe_int(request.form.get("advance_per_group")) if request.form.get("advance_per_group") not in (None, "") else rec["advance"]),
                    best_runnerup_count=max(0, safe_int(request.form.get("best_runnerup_count")) if request.form.get("best_runnerup_count") not in (None, "") else rec["best"]),
                    tiebreakers=selected_tiebreakers,
                    status="scheduled",
                )
                db.session.add(comp)
                db.session.flush()
                create_rr_groups(comp)
                assign_teams_auto(comp, [t.id for t in selected_teams])
                generate_rr_matches(comp)
                db.session.commit()
                flash(f"สร้าง {comp.name} แบบพบกันหมด/แบ่งกลุ่ม จากทีมที่เลือก {len(selected_teams)} ทีมแล้ว", "success")
                return redirect(url_for("rr_detail", comp_id=comp.id))

            if competition_format == "score_judging":
                comp = ContestCompetition(event_id=event.id, sport_division_id=division.id, name=comp_name, activity_type=sport.name, status="draft")
                db.session.add(comp)
                db.session.flush()
                for idx, cname in enumerate(["ความพร้อม", "ทักษะ/คุณภาพ", "ความคิดสร้างสรรค์", "ความพร้อมเพรียง", "ภาพรวม"], start=1):
                    db.session.add(ContestCriterion(competition_id=comp.id, name=cname, max_score=100, sort_order=idx))
                db.session.add(ContestJudge(competition_id=comp.id, name="กรรมการ 1"))
                db.session.commit()
                flash("สร้างรายการกรรมการให้คะแนนแล้ว", "success")
                return redirect(url_for("contest_detail", comp_id=comp.id))

            comp = RankingCompetition(event_id=event.id, sport_division_id=division.id, name=comp_name, result_mode=request.form.get("result_mode", "rank"), status="draft")
            db.session.add(comp)
            db.session.commit()
            flash("สร้างรายการ Ranking แล้ว", "success")
            return redirect(url_for("ranking_detail", comp_id=comp.id))

        return render_template("sports/competition_wizard.html", event=event, teams=teams, sports=sports, divisions=divisions, sport_presets=sport_presets)

    @app.route("/events/<int:event_id>/sports/categories/add", methods=["POST"])
    @login_required
    def sport_category_add(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดการกีฬาในงานนี้", "danger")
            return redirect(url_for("events"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("กรุณากรอกชื่อหมวดกีฬา", "danger")
        elif SportCategory.query.filter_by(event_id=event.id, name=name).first():
            flash("หมวดกีฬานี้มีแล้ว", "warning")
        else:
            db.session.add(SportCategory(event_id=event.id, name=name, description=request.form.get("description", "").strip(), sort_order=safe_int(request.form.get("sort_order"))))
            db.session.commit()
            flash("เพิ่มหมวดกีฬาแล้ว", "success")
        return redirect(url_for("event_sports", event_id=event.id))

    @app.route("/sports/categories/<int:category_id>/delete", methods=["POST"])
    @login_required
    def sport_category_delete(category_id):
        category = SportCategory.query.get_or_404(category_id)
        if not can_access_org(category.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบหมวดนี้", "danger")
            return redirect(url_for("events"))
        event_id = category.event_id
        db.session.delete(category)
        db.session.commit()
        flash("ลบหมวดกีฬาแล้ว", "info")
        return redirect(url_for("event_sports", event_id=event_id))

    @app.route("/sports/categories/<int:category_id>/update", methods=["POST"])
    @login_required
    def sport_category_update(category_id):
        category = SportCategory.query.get_or_404(category_id)
        if not can_access_org(category.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขหมวดนี้", "danger")
            return redirect(url_for("events"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("กรุณากรอกชื่อหมวดกีฬา", "danger")
        else:
            duplicate = SportCategory.query.filter(SportCategory.event_id == category.event_id, SportCategory.name == name, SportCategory.id != category.id).first()
            if duplicate:
                flash("หมวดกีฬานี้มีแล้ว", "warning")
            else:
                category.name = name
                category.description = request.form.get("description", "").strip()
                category.sort_order = safe_int(request.form.get("sort_order"))
                db.session.commit()
                flash("แก้ไขหมวดกีฬาแล้ว", "success")
        return redirect(url_for("event_sports", event_id=category.event_id) + "#sport-setup-categories")

    @app.route("/events/<int:event_id>/sports/add", methods=["POST"])
    @login_required
    def sport_add(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดการกีฬาในงานนี้", "danger")
            return redirect(url_for("events"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("กรุณากรอกชื่อชนิดกีฬา", "danger")
        elif Sport.query.filter_by(event_id=event.id, name=name).first():
            flash("ชนิดกีฬานี้มีแล้ว", "warning")
        else:
            category_id = request.form.get("category_id") or None
            result_type = request.form.get("result_type", "score_only")
            sport = Sport(
                event_id=event.id,
                category_id=category_id,
                name=name,
                default_format=request.form.get("default_format", "ranking"),
                result_type=result_type,
                max_sets=safe_int(request.form.get("max_sets")) if result_type == "set_based" else 0,
                points_per_set=safe_int(request.form.get("points_per_set")) if result_type == "set_based" else 0,
                sets_to_win=safe_int(request.form.get("sets_to_win")) if result_type == "set_based" else 0,
                note=request.form.get("note", "").strip(),
                is_active=bool(request.form.get("is_active", "1")),
            )
            db.session.add(sport)
            db.session.commit()
            flash("เพิ่มชนิดกีฬาแล้ว", "success")
        return redirect(url_for("event_sports", event_id=event.id))

    @app.route("/events/<int:event_id>/sports/quick-add", methods=["POST"])
    @login_required
    def sport_quick_add(event_id):
        """เพิ่มกีฬาเข้าอีเว้นท์แบบผู้ใช้ทั่วไป: จบในฟอร์มเดียว
        - เลือก/สร้างหมวดกีฬา
        - เลือกกีฬาเดิมหรือสร้างชนิดกีฬาใหม่
        - สร้างรุ่นแข่งขัน + เพศหลายรายการพร้อมกัน
        """
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดการกีฬาในงานนี้", "danger")
            return redirect(url_for("events"))

        def split_lines(raw):
            raw = (raw or "").strip()
            for sep in ["\r", "\n", "，", "、", ";", "|", "\t"]:
                raw = raw.replace(sep, ",")
            items = []
            for part in raw.split(","):
                value = part.strip()
                if value and value not in items:
                    items.append(value)
            return items

        sport_id = safe_int_or_none(request.form.get("sport_id"))
        sport = Sport.query.filter_by(id=sport_id, event_id=event.id).first() if sport_id else None
        sport_name = (request.form.get("sport_name") or "").strip()
        category_name = (request.form.get("category_name") or request.form.get("category_preset") or "").strip()
        category_id = safe_int_or_none(request.form.get("category_id"))
        category = SportCategory.query.filter_by(id=category_id, event_id=event.id).first() if category_id else None

        if not category and category_name:
            category = SportCategory.query.filter_by(event_id=event.id, name=category_name).first()
            if not category:
                category = SportCategory(event_id=event.id, name=category_name, description="สร้างจากฟอร์มเพิ่มกีฬาแบบง่าย")
                db.session.add(category)
                db.session.flush()

        competition_format = request.form.get("competition_format") or (sport.default_format if sport else "ranking")
        result_type = request.form.get("result_type") or (sport.result_type if sport else "score_only")
        if result_type == "set_based":
            max_sets = safe_int(request.form.get("max_sets")) or (sport.max_sets if sport else 0) or 3
            points_per_set = safe_int(request.form.get("points_per_set")) or (sport.points_per_set if sport else 0) or 21
            sets_to_win = safe_int(request.form.get("sets_to_win")) or (sport.sets_to_win if sport else 0) or ((max_sets // 2) + 1)
        else:
            max_sets = 0
            points_per_set = 0
            sets_to_win = 0
        max_athletes_per_team = safe_int_or_none(request.form.get("max_athletes_per_team"))
        note = (request.form.get("note") or "").strip()

        if not sport:
            if not sport_name:
                flash("กรุณาเลือกกีฬาเดิม หรือกรอกชื่อชนิดกีฬาใหม่", "danger")
                return redirect(url_for("event_sports", event_id=event.id) + "#sport-wizard")
            sport = Sport.query.filter_by(event_id=event.id, name=sport_name).first()
            if not sport:
                sport = Sport(
                    event_id=event.id,
                    category_id=category.id if category else None,
                    name=sport_name,
                    default_format=competition_format,
                    result_type=result_type,
                    max_sets=max_sets,
                    points_per_set=points_per_set,
                    sets_to_win=sets_to_win,
                    note=note,
                    is_active=True,
                )
                db.session.add(sport)
                db.session.flush()
            elif category and not sport.category_id:
                sport.category_id = category.id
        elif category:
            sport.category_id = category.id

        class_names = split_lines(request.form.get("class_name")) or ["Open"]
        genders = request.form.getlist("gender") or ["ผสม"]

        added = 0
        skipped = 0
        for class_name in class_names:
            for gender in genders:
                exists = SportDivision.query.filter_by(sport_id=sport.id, class_name=class_name, gender=gender).first()
                if exists:
                    skipped += 1
                    continue
                db.session.add(SportDivision(
                    sport_id=sport.id,
                    class_name=class_name,
                    gender=gender,
                    competition_format=competition_format,
                    result_type=result_type,
                    max_sets=max_sets or sport.max_sets or 0,
                    points_per_set=points_per_set or sport.points_per_set or 0,
                    sets_to_win=sets_to_win or sport.sets_to_win or 0,
                    max_athletes_per_team=max_athletes_per_team,
                    is_active=True,
                ))
                added += 1

        db.session.commit()
        if added:
            flash(f"เพิ่ม {sport.name} เข้าอีเว้นท์แล้ว {added} รายการ" + (f" · ข้ามรายการซ้ำ {skipped}" if skipped else ""), "success")
        else:
            flash(f"ยังไม่มีรายการใหม่สำหรับ {sport.name} เพราะรายการที่เลือกมีอยู่แล้ว", "warning")
        return redirect(url_for("event_sports", event_id=event.id) + "#sport-wizard")

    @app.route("/sports/<int:sport_id>/delete", methods=["POST"])
    @login_required
    def sport_delete(sport_id):
        sport = Sport.query.get_or_404(sport_id)
        if not can_access_org(sport.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบกีฬานี้", "danger")
            return redirect(url_for("events"))
        event_id = sport.event_id
        db.session.delete(sport)
        db.session.commit()
        flash("ลบชนิดกีฬาแล้ว", "info")
        return redirect(url_for("event_sports", event_id=event_id))

    @app.route("/sports/<int:sport_id>/update", methods=["POST"])
    @login_required
    def sport_update(sport_id):
        sport = Sport.query.get_or_404(sport_id)
        if not can_access_org(sport.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขกีฬานี้", "danger")
            return redirect(url_for("events"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("กรุณากรอกชื่อชนิดกีฬา", "danger")
        else:
            duplicate = Sport.query.filter(Sport.event_id == sport.event_id, Sport.name == name, Sport.id != sport.id).first()
            if duplicate:
                flash("ชนิดกีฬานี้มีแล้ว", "warning")
            else:
                result_type = request.form.get("result_type", "score_only")
                sport.name = name
                sport.category_id = safe_int_or_none(request.form.get("category_id"))
                sport.default_format = request.form.get("default_format", "ranking")
                sport.result_type = result_type
                sport.max_sets = safe_int(request.form.get("max_sets")) if result_type == "set_based" else 0
                sport.points_per_set = safe_int(request.form.get("points_per_set")) if result_type == "set_based" else 0
                sport.sets_to_win = safe_int(request.form.get("sets_to_win")) if result_type == "set_based" else 0
                sport.note = request.form.get("note", "").strip()
                sport.is_active = bool(request.form.get("is_active"))
                db.session.commit()
                flash("แก้ไขชนิดกีฬาและรูปแบบบันทึกผลแล้ว", "success")
        return redirect(url_for("event_sports", event_id=sport.event_id) + f"#sport-{sport.id}")

    @app.route("/events/<int:event_id>/sports/divisions/add", methods=["POST"])
    @login_required
    def sport_division_add(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดการรุ่นแข่งขันในงานนี้", "danger")
            return redirect(url_for("events"))
        sport_id = request.form.get("sport_id")
        sport = Sport.query.filter_by(id=sport_id, event_id=event.id).first()
        if not sport:
            flash("กรุณาเลือกชนิดกีฬา", "danger")
            return redirect(url_for("event_sports", event_id=event.id))
        class_names = [x.strip() for x in request.form.get("class_name", "").replace("\n", ",").split(",") if x.strip()]
        genders = request.form.getlist("gender") or ["ชาย"]
        if not class_names:
            flash("กรุณากรอกรุ่นแข่งขัน", "danger")
            return redirect(url_for("event_sports", event_id=event.id))
        added = 0
        for class_name in class_names:
            for gender in genders:
                if SportDivision.query.filter_by(sport_id=sport.id, class_name=class_name, gender=gender).first():
                    continue
                db.session.add(SportDivision(
                    sport_id=sport.id,
                    class_name=class_name,
                    gender=gender,
                    competition_format=request.form.get("competition_format") or sport.default_format or "ranking",
                    result_type=request.form.get("result_type") or sport.result_type or "score_only",
                    max_sets=safe_int(request.form.get("max_sets")) or sport.max_sets or 0,
                    points_per_set=safe_int(request.form.get("points_per_set")) or sport.points_per_set or 0,
                    sets_to_win=safe_int(request.form.get("sets_to_win")) or sport.sets_to_win or 0,
                    max_athletes_per_team=safe_int_or_none(request.form.get("max_athletes_per_team")),
                    is_active=True,
                ))
                added += 1
        db.session.commit()
        flash(f"เพิ่มรายการย่อยแล้ว {added} รายการ", "success")
        return redirect(url_for("event_sports", event_id=event.id))

    @app.route("/sports/divisions/<int:division_id>/delete", methods=["POST"])
    @login_required
    def sport_division_delete(division_id):
        division = SportDivision.query.get_or_404(division_id)
        event = division.sport.event
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบรายการนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(division)
        db.session.commit()
        flash("ลบรายการย่อยแล้ว", "info")
        return redirect(url_for("event_sports", event_id=event.id))

    @app.route("/sports/divisions/<int:division_id>/update", methods=["POST"])
    @login_required
    def sport_division_update(division_id):
        division = SportDivision.query.get_or_404(division_id)
        event = division.sport.event
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขรายการนี้", "danger")
            return redirect(url_for("events"))
        sport = Sport.query.filter_by(id=safe_int_or_none(request.form.get("sport_id")), event_id=event.id).first() or division.sport
        class_name = request.form.get("class_name", "").strip()
        gender = request.form.get("gender", division.gender).strip() or division.gender
        if not class_name:
            flash("กรุณากรอกรุ่นแข่งขัน", "danger")
        else:
            duplicate = SportDivision.query.filter(
                SportDivision.sport_id == sport.id,
                SportDivision.class_name == class_name,
                SportDivision.gender == gender,
                SportDivision.id != division.id,
            ).first()
            if duplicate:
                flash("รายการย่อยนี้มีแล้ว", "warning")
            else:
                result_type = request.form.get("result_type") or sport.result_type or "score_only"
                division.sport_id = sport.id
                division.class_name = class_name
                division.gender = gender
                division.competition_format = request.form.get("competition_format") or sport.default_format or "ranking"
                division.result_type = result_type
                division.max_sets = safe_int(request.form.get("max_sets")) if result_type == "set_based" else 0
                division.points_per_set = safe_int(request.form.get("points_per_set")) if result_type == "set_based" else 0
                division.sets_to_win = safe_int(request.form.get("sets_to_win")) if result_type == "set_based" else 0
                division.max_athletes_per_team = safe_int_or_none(request.form.get("max_athletes_per_team"))
                division.is_active = bool(request.form.get("is_active"))
                db.session.commit()
                flash("แก้ไขรายการย่อยและวิธีบันทึกผลแล้ว", "success")
        return redirect(url_for("event_sports", event_id=event.id) + f"#division-{division.id}")

    @app.route("/events/<int:event_id>/sports/seed-defaults", methods=["POST"])
    @login_required
    def sport_seed_defaults(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดการกีฬาในงานนี้", "danger")
            return redirect(url_for("events"))
        seed_default_sports(event)
        flash("ล้างคลังกีฬาเดิมและสร้างชุดมาตรฐานล่าสุดแล้ว ขั้นต่อไปเลือกกีฬาที่งานนี้จัดและสร้างการแข่งขันจริง", "success")
        return redirect(url_for("event_competition_wizard", event_id=event.id))


    @app.route("/events/<int:event_id>/round-robin")
    @login_required
    def event_round_robin(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        competitions = RoundRobinCompetition.query.filter_by(event_id=event.id).order_by(RoundRobinCompetition.created_at.desc()).all()
        return render_template("round_robin/list.html", event=event, competitions=competitions)

    @app.route("/events/<int:event_id>/round-robin/new", methods=["GET", "POST"])
    @login_required
    def rr_new(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์สร้างรายการในงานนี้", "danger")
            return redirect(url_for("events"))
        divisions = SportDivision.query.join(Sport).filter(Sport.event_id == event.id).order_by(Sport.name, SportDivision.class_name, SportDivision.gender).all()
        teams = Team.query.filter_by(event_id=event.id).order_by(Team.name).all()
        if request.method == "POST":
            selected_team_ids = [int(x) for x in request.form.getlist("team_ids") if str(x).isdigit()]
            if not selected_team_ids:
                selected_team_ids = [t.id for t in teams]
            comp = RoundRobinCompetition(
                event_id=event.id,
                sport_division_id=safe_int_or_none(request.form.get("sport_division_id")),
                name=request.form.get("name", "").strip() or "Round Robin",
                num_groups=max(1, safe_int(request.form.get("num_groups")) or 1),
                win_points=safe_int(request.form.get("win_points")) if request.form.get("win_points") not in (None, "") else 3,
                draw_points=safe_int(request.form.get("draw_points")) if request.form.get("draw_points") not in (None, "") else 1,
                loss_points=safe_int(request.form.get("loss_points")) if request.form.get("loss_points") not in (None, "") else 0,
                advance_per_group=max(0, safe_int(request.form.get("advance_per_group")) or 0),
                best_runnerup_count=max(0, safe_int(request.form.get("best_runnerup_count")) or 0),
                tiebreakers=",".join(request.form.getlist("tiebreakers")) or "points,goal_diff,goals_for,head_to_head,wins",
                status="draft",
            )
            db.session.add(comp)
            db.session.flush()
            create_rr_groups(comp)
            if request.form.get("assignment_mode") == "auto":
                assign_teams_auto(comp, selected_team_ids)
                generate_rr_matches(comp)
                comp.status = "scheduled"
                db.session.commit()
                flash("สร้างรายการ แบ่งกลุ่มอัตโนมัติ และสร้างตารางพบกันหมดแล้ว", "success")
                return redirect(url_for("rr_detail", comp_id=comp.id))
            db.session.commit()
            flash("สร้างรายการแล้ว กรุณาลากทีมลงกลุ่ม", "success")
            return redirect(url_for("rr_assign", comp_id=comp.id))
        return render_template("round_robin/form.html", event=event, divisions=divisions, teams=teams)

    @app.route("/round-robin/<int:comp_id>")
    @login_required
    def rr_detail(comp_id):
        comp = RoundRobinCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงรายการนี้", "danger")
            return redirect(url_for("events"))
        standings = calculate_rr_standings(comp)
        qualifiers = calculate_rr_qualifiers(comp, standings)
        return render_template("round_robin/detail.html", comp=comp, standings=standings, qualifiers=qualifiers)

    @app.route("/round-robin/<int:comp_id>/assign", methods=["GET", "POST"])
    @login_required
    def rr_assign(comp_id):
        comp = RoundRobinCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดกลุ่มรายการนี้", "danger")
            return redirect(url_for("events"))
        teams = Team.query.filter_by(event_id=comp.event_id).order_by(Team.name).all()
        if request.method == "POST":
            RoundRobinMatch.query.filter_by(competition_id=comp.id).delete()
            group_ids = [g.id for g in comp.groups]
            if group_ids:
                RoundRobinGroupTeam.query.filter(RoundRobinGroupTeam.group_id.in_(group_ids)).delete(synchronize_session=False)
            for group in comp.groups:
                ids = [x for x in request.form.get(f"group_{group.id}", "").split(",") if x.strip().isdigit()]
                for order, team_id in enumerate(ids, start=1):
                    db.session.add(RoundRobinGroupTeam(group_id=group.id, team_id=int(team_id), sort_order=order))
            db.session.flush()
            generate_rr_matches(comp)
            comp.status = "scheduled"
            db.session.commit()
            flash("บันทึกกลุ่มและสร้างตารางพบกันหมดแล้ว", "success")
            return redirect(url_for("rr_detail", comp_id=comp.id))
        assigned_ids = {gt.team_id for g in comp.groups for gt in g.group_teams}
        unassigned = [t for t in teams if t.id not in assigned_ids]
        return render_template("round_robin/assign.html", comp=comp, teams=teams, unassigned=unassigned)

    @app.route("/round-robin/<int:comp_id>/auto-assign", methods=["POST"])
    @login_required
    def rr_auto_assign(comp_id):
        comp = RoundRobinCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์จัดกลุ่มรายการนี้", "danger")
            return redirect(url_for("events"))
        team_ids = [t.id for t in Team.query.filter_by(event_id=comp.event_id).all()]
        RoundRobinMatch.query.filter_by(competition_id=comp.id).delete()
        group_ids = [g.id for g in comp.groups]
        if group_ids:
            RoundRobinGroupTeam.query.filter(RoundRobinGroupTeam.group_id.in_(group_ids)).delete(synchronize_session=False)
        assign_teams_auto(comp, team_ids)
        generate_rr_matches(comp)
        comp.status = "scheduled"
        db.session.commit()
        flash("สุ่มแบ่งกลุ่มและสร้างตารางใหม่แล้ว", "success")
        return redirect(url_for("rr_detail", comp_id=comp.id))

    @app.route("/round-robin/matches/<int:match_id>/score", methods=["POST"])
    @login_required
    def rr_match_score(match_id):
        match = RoundRobinMatch.query.get_or_404(match_id)
        comp = match.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์บันทึกผลรายการนี้", "danger")
            return redirect(url_for("events"))
        if rr_result_type(comp) == "set_based":
            return redirect(url_for("rr_match_result", match_id=match.id))
        match.score_a = safe_int_or_none(request.form.get("score_a"))
        match.score_b = safe_int_or_none(request.form.get("score_b"))
        match.set_a = None
        match.set_b = None
        match.set_scores = None
        match.point_diff = 0
        match.note = request.form.get("note", "").strip()
        match.status = "completed" if match.score_a is not None and match.score_b is not None else "scheduled"
        comp.status = "in_progress"
        db.session.commit()
        flash("บันทึกผลแล้ว", "success")
        return redirect(url_for("rr_detail", comp_id=comp.id) + f"#match-{match.id}")

    @app.route("/round-robin/matches/<int:match_id>/result", methods=["GET", "POST"])
    @login_required
    def rr_match_result(match_id):
        match = RoundRobinMatch.query.get_or_404(match_id)
        comp = match.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์บันทึกผลรายการนี้", "danger")
            return redirect(url_for("events"))
        if rr_result_type(comp) != "set_based":
            return redirect(url_for("rr_detail", comp_id=comp.id) + f"#match-{match.id}")
        cfg = rr_set_config(comp)
        current_sets = parse_set_scores(match)
        if request.method == "POST":
            set_scores = []
            sets_a = 0
            sets_b = 0
            total_a = 0
            total_b = 0
            for i in range(1, cfg["max_sets"] + 1):
                a = safe_int_or_none(request.form.get(f"set_{i}_a"))
                b = safe_int_or_none(request.form.get(f"set_{i}_b"))
                if a is None and b is None:
                    continue
                a = a or 0
                b = b or 0
                set_scores.append({"set": i, "a": a, "b": b})
                total_a += a
                total_b += b
                if a > b:
                    sets_a += 1
                elif b > a:
                    sets_b += 1
            match.set_scores = json.dumps(set_scores, ensure_ascii=False) if set_scores else None
            match.score_history = normalize_score_history_payload(request.form.get("score_history"))
            match.set_a = sets_a if set_scores else None
            match.set_b = sets_b if set_scores else None
            # สำหรับ Set Based: score_a/score_b = ผลเซต, ส่วนแต้มรวมเก็บใน set_scores/point_diff
            match.score_a = sets_a if set_scores else None
            match.score_b = sets_b if set_scores else None
            match.point_diff = total_a - total_b if set_scores else 0
            match.note = request.form.get("note", "").strip()
            match.status = "completed" if set_scores and (sets_a >= cfg["sets_to_win"] or sets_b >= cfg["sets_to_win"] or len(set_scores) >= cfg["max_sets"]) else "scheduled"
            comp.status = "in_progress"
            db.session.commit()
            flash("บันทึกคะแนนรายเซตแล้ว", "success")
            return redirect(url_for("rr_detail", comp_id=comp.id) + f"#match-{match.id}")
        group_standings = calculate_rr_standings(comp).get(match.group_id, [])
        return render_template("round_robin/match_result.html", match=match, comp=comp, cfg=cfg, current_sets=current_sets, score_history=parse_score_history(match), group_standings=group_standings)

    @app.route("/round-robin/<int:comp_id>/delete", methods=["POST"])
    @login_required
    def rr_delete(comp_id):
        comp = RoundRobinCompetition.query.get_or_404(comp_id)
        event_id = comp.event_id
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบรายการนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(comp)
        db.session.commit()
        flash("ลบรายการ Round Robin แล้ว", "info")
        return redirect(url_for("event_round_robin", event_id=event_id))


    @app.route("/round-robin/<int:comp_id>/create-knockout", methods=["POST"])
    @login_required
    def rr_create_knockout(comp_id):
        comp = RoundRobinCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์สร้างรอบ Knockout", "danger")
            return redirect(url_for("events"))
        standings = calculate_rr_standings(comp)
        qualifiers = calculate_rr_qualifiers(comp, standings)
        teams = build_rr_playoff_seed_order(comp, standings)
        if len(teams) < 2:
            flash("ยังมีทีมเข้ารอบไม่พอสำหรับสร้าง Knockout", "warning")
            return redirect(url_for("rr_detail", comp_id=comp.id))
        existing = KnockoutCompetition.query.filter_by(source_round_robin_id=comp.id).first()
        if existing:
            flash("รายการ Knockout จากรอบนี้มีอยู่แล้ว", "info")
            return redirect(url_for("knockout_detail", comp_id=existing.id))
        div = comp.sport_division
        ko = KnockoutCompetition(
            event_id=comp.event_id,
            sport_division_id=comp.sport_division_id,
            source_round_robin_id=comp.id,
            name=f"{comp.name} · Knockout",
            result_type=(div.result_type if div else rr_result_type(comp)) or "score_only",
            max_sets=(div.max_sets if div else 0) or 0,
            points_per_set=(div.points_per_set if div else 0) or 0,
            sets_to_win=(div.sets_to_win if div else 0) or 0,
            status="scheduled",
        )
        db.session.add(ko)
        db.session.flush()
        create_knockout_first_round(ko, teams, pairing_mode="adjacent")
        db.session.commit()
        flash("สร้างรอบ Knockout จากทีมเข้ารอบแล้ว (ไขว้สายตามสูตรที่กำหนด)", "success")
        return redirect(url_for("knockout_detail", comp_id=ko.id))

    @app.route("/events/<int:event_id>/knockout")
    @login_required
    def event_knockout(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        competitions = KnockoutCompetition.query.filter_by(event_id=event.id).order_by(KnockoutCompetition.created_at.desc()).all()
        return render_template("knockout/list.html", event=event, competitions=competitions)

    @app.route("/knockout/<int:comp_id>")
    @login_required
    def knockout_detail(comp_id):
        comp = KnockoutCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงรายการนี้", "danger")
            return redirect(url_for("events"))
        rounds = {}
        for m in comp.matches:
            rounds.setdefault((m.round_no, m.round_name), []).append(m)
        return render_template("knockout/detail.html", comp=comp, rounds=rounds)

    @app.route("/knockout/matches/<int:match_id>/score", methods=["POST"])
    @login_required
    def knockout_match_score(match_id):
        match = KnockoutMatch.query.get_or_404(match_id)
        comp = match.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์บันทึกผลรายการนี้", "danger")
            return redirect(url_for("events"))
        if comp.result_type == "set_based":
            return redirect(url_for("knockout_match_result", match_id=match.id))
        else:
            match.score_a = safe_int_or_none(request.form.get("score_a"))
            match.score_b = safe_int_or_none(request.form.get("score_b"))
            match.set_a = None
            match.set_b = None
        match.note = request.form.get("note", "").strip()
        if match.team_a_id and not match.team_b_id:
            match.winner_team_id = match.team_a_id
            match.status = "completed"
        elif match.team_b_id and not match.team_a_id:
            match.winner_team_id = match.team_b_id
            match.status = "completed"
        elif match.score_a is not None and match.score_b is not None and match.score_a != match.score_b:
            match.winner_team_id = match.team_a_id if match.score_a > match.score_b else match.team_b_id
            match.status = "completed"
        else:
            match.winner_team_id = None
            match.status = "scheduled"
        db.session.commit()
        advance_knockout_if_ready(comp)
        flash("บันทึกผล Knockout แล้ว", "success")
        return redirect(url_for("knockout_detail", comp_id=comp.id) + f"#match-{match.id}")

    @app.route("/knockout/matches/<int:match_id>/result", methods=["GET", "POST"])
    @login_required
    def knockout_match_result(match_id):
        match = KnockoutMatch.query.get_or_404(match_id)
        comp = match.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์บันทึกผลรายการนี้", "danger")
            return redirect(url_for("events"))
        if comp.result_type != "set_based":
            return redirect(url_for("knockout_detail", comp_id=comp.id) + f"#match-{match.id}")
        cfg = knockout_set_config(comp)
        current_sets = parse_set_scores(match)
        if request.method == "POST":
            set_scores = []
            sets_a = 0
            sets_b = 0
            total_a = 0
            total_b = 0
            for i in range(1, cfg["max_sets"] + 1):
                a = safe_int_or_none(request.form.get(f"set_{i}_a"))
                b = safe_int_or_none(request.form.get(f"set_{i}_b"))
                if a is None and b is None:
                    continue
                a = a or 0
                b = b or 0
                set_scores.append({"set": i, "a": a, "b": b})
                total_a += a
                total_b += b
                if a > b:
                    sets_a += 1
                elif b > a:
                    sets_b += 1
            match.set_scores = json.dumps(set_scores, ensure_ascii=False) if set_scores else None
            match.score_history = normalize_score_history_payload(request.form.get("score_history"))
            match.set_a = sets_a if set_scores else None
            match.set_b = sets_b if set_scores else None
            match.score_a = sets_a if set_scores else None
            match.score_b = sets_b if set_scores else None
            match.point_diff = total_a - total_b if set_scores else 0
            match.note = request.form.get("note", "").strip()
            if match.team_a_id and not match.team_b_id:
                match.winner_team_id = match.team_a_id
                match.status = "completed"
            elif match.team_b_id and not match.team_a_id:
                match.winner_team_id = match.team_b_id
                match.status = "completed"
            elif set_scores and sets_a != sets_b and (sets_a >= cfg["sets_to_win"] or sets_b >= cfg["sets_to_win"] or len(set_scores) >= cfg["max_sets"]):
                match.winner_team_id = match.team_a_id if sets_a > sets_b else match.team_b_id
                match.status = "completed"
            else:
                match.winner_team_id = None
                match.status = "scheduled"
            db.session.commit()
            advance_knockout_if_ready(comp)
            flash("บันทึกคะแนนรายเซต Knockout แล้ว", "success")
            return redirect(url_for("knockout_detail", comp_id=comp.id) + f"#match-{match.id}")
        return render_template("knockout/match_result.html", match=match, comp=comp, cfg=cfg, current_sets=current_sets, score_history=parse_score_history(match))


    @app.route("/knockout/<int:comp_id>/delete", methods=["POST"])
    @login_required
    def knockout_delete(comp_id):
        comp = KnockoutCompetition.query.get_or_404(comp_id)
        event_id = comp.event_id
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบรายการนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(comp)
        db.session.commit()
        flash("ลบรายการ Knockout แล้ว", "info")
        return redirect(url_for("event_knockout", event_id=event_id))


    @app.route("/events/<int:event_id>/ranking")
    @login_required
    def event_ranking(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        competitions = RankingCompetition.query.filter_by(event_id=event.id).order_by(RankingCompetition.created_at.desc()).all()
        return render_template("ranking/list.html", event=event, competitions=competitions)

    @app.route("/events/<int:event_id>/ranking/new", methods=["GET", "POST"])
    @login_required
    def ranking_new(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์สร้างรายการในงานนี้", "danger")
            return redirect(url_for("events"))
        divisions = SportDivision.query.join(Sport).filter(Sport.event_id == event.id).order_by(Sport.name, SportDivision.class_name, SportDivision.gender).all()
        if request.method == "POST":
            division_id = safe_int_or_none(request.form.get("sport_division_id"))
            division = SportDivision.query.get(division_id) if division_id else None
            default_name = division.label if division else "Ranking Competition"
            comp = RankingCompetition(
                event_id=event.id,
                sport_division_id=division_id,
                name=request.form.get("name", "").strip() or default_name,
                result_mode=request.form.get("result_mode", "rank"),
                status="draft",
            )
            db.session.add(comp)
            db.session.commit()
            flash("สร้างรายการ Ranking แล้ว", "success")
            return redirect(url_for("ranking_detail", comp_id=comp.id))
        return render_template("ranking/form.html", event=event, divisions=divisions)

    @app.route("/ranking/<int:comp_id>")
    @login_required
    def ranking_detail(comp_id):
        comp = RankingCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงรายการนี้", "danger")
            return redirect(url_for("events"))
        teams = Team.query.filter_by(event_id=comp.event_id).order_by(Team.name).all()
        athletes = Athlete.query.join(Team).filter(Team.event_id == comp.event_id).order_by(Team.name, Athlete.full_name).all()
        results = RankingResult.query.filter_by(competition_id=comp.id).order_by(RankingResult.rank.asc().nullslast(), RankingResult.created_at.asc()).all()
        return render_template("ranking/detail.html", comp=comp, teams=teams, athletes=athletes, results=results)

    @app.route("/ranking/<int:comp_id>/results/add", methods=["POST"])
    @login_required
    def ranking_result_add(comp_id):
        comp = RankingCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์บันทึกผลรายการนี้", "danger")
            return redirect(url_for("events"))
        result = RankingResult(competition_id=comp.id)
        save_ranking_result_from_request(result, comp)
        db.session.add(result)
        comp.status = "completed"
        db.session.commit()
        refresh_ranking_medals(comp)
        flash("บันทึกผล Ranking แล้ว", "success")
        return redirect(url_for("ranking_detail", comp_id=comp.id))

    @app.route("/ranking/results/<int:result_id>/update", methods=["POST"])
    @login_required
    def ranking_result_update(result_id):
        result = RankingResult.query.get_or_404(result_id)
        comp = result.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขผลรายการนี้", "danger")
            return redirect(url_for("events"))
        save_ranking_result_from_request(result, comp)
        db.session.commit()
        refresh_ranking_medals(comp)
        flash("แก้ไขผลแล้ว", "success")
        return redirect(url_for("ranking_detail", comp_id=comp.id))

    @app.route("/ranking/results/<int:result_id>/delete", methods=["POST"])
    @login_required
    def ranking_result_delete(result_id):
        result = RankingResult.query.get_or_404(result_id)
        comp = result.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบผลรายการนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(result)
        db.session.commit()
        refresh_ranking_medals(comp)
        flash("ลบผลแล้ว", "info")
        return redirect(url_for("ranking_detail", comp_id=comp.id))

    @app.route("/ranking/<int:comp_id>/print")
    @login_required
    def ranking_print(comp_id):
        comp = RankingCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์พิมพ์รายการนี้", "danger")
            return redirect(url_for("events"))
        results = RankingResult.query.filter_by(competition_id=comp.id).order_by(RankingResult.rank.asc().nullslast()).all()
        return render_template("ranking/print.html", comp=comp, results=results)

    @app.route("/ranking/<int:comp_id>/export.xlsx")
    @login_required
    def ranking_export_excel(comp_id):
        comp = RankingCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ Export รายการนี้", "danger")
            return redirect(url_for("events"))
        output = build_ranking_workbook(comp)
        safe_name = secure_filename(comp.name) or f"ranking_{comp.id}"
        return send_file(output, as_attachment=True, download_name=f"{safe_name}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/ranking/<int:comp_id>/delete", methods=["POST"])
    @login_required
    def ranking_delete(comp_id):
        comp = RankingCompetition.query.get_or_404(comp_id)
        event_id = comp.event_id
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบรายการนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(comp)
        db.session.commit()
        flash("ลบรายการ Ranking แล้ว", "info")
        return redirect(url_for("event_ranking", event_id=event_id))



    @app.route("/events/<int:event_id>/contests")
    @login_required
    def event_contests(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        competitions = ContestCompetition.query.filter_by(event_id=event.id).order_by(ContestCompetition.created_at.desc()).all()
        return render_template("contests/list.html", event=event, competitions=competitions)

    @app.route("/events/<int:event_id>/contests/new", methods=["GET", "POST"])
    @login_required
    def contest_new(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์สร้างกิจกรรมในงานนี้", "danger")
            return redirect(url_for("events"))
        divisions = SportDivision.query.join(Sport).filter(Sport.event_id == event.id).order_by(Sport.name, SportDivision.class_name, SportDivision.gender).all()
        if request.method == "POST":
            division_id = safe_int_or_none(request.form.get("sport_division_id"))
            division = SportDivision.query.get(division_id) if division_id else None
            name = request.form.get("name", "").strip() or (division.label if division else "กิจกรรมประกวด")
            comp = ContestCompetition(
                event_id=event.id,
                sport_division_id=division_id,
                name=name,
                activity_type=request.form.get("activity_type", "กิจกรรมประกวด").strip() or "กิจกรรมประกวด",
                status="draft",
            )
            db.session.add(comp)
            db.session.flush()
            raw_criteria = request.form.get("criteria", "").strip()
            if raw_criteria:
                for idx, line in enumerate(raw_criteria.splitlines(), start=1):
                    parts = [x.strip() for x in line.split("|")]
                    cname = parts[0]
                    max_score = float(parts[1]) if len(parts) > 1 and parts[1] else 100
                    if cname:
                        db.session.add(ContestCriterion(competition_id=comp.id, name=cname, max_score=max_score, sort_order=idx))
            else:
                defaults = ["ความคิดสร้างสรรค์", "ความสวยงาม", "ความพร้อมเพรียง", "การนำเสนอ", "ความประทับใจ"]
                for idx, cname in enumerate(defaults, start=1):
                    db.session.add(ContestCriterion(competition_id=comp.id, name=cname, max_score=100, sort_order=idx))
            judges = [x.strip() for x in request.form.get("judges", "").splitlines() if x.strip()]
            for j in judges:
                db.session.add(ContestJudge(competition_id=comp.id, name=j))
            if not judges:
                db.session.add(ContestJudge(competition_id=comp.id, name="กรรมการ 1"))
            db.session.commit()
            flash("สร้างกิจกรรมประกวดแล้ว", "success")
            return redirect(url_for("contest_detail", comp_id=comp.id))
        return render_template("contests/form.html", event=event, divisions=divisions)

    @app.route("/contests/<int:comp_id>")
    @login_required
    def contest_detail(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงกิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        teams = Team.query.filter_by(event_id=comp.event_id).order_by(Team.name).all()
        selected_judge_id = safe_int_or_none(request.args.get("judge_id")) or (comp.judges[0].id if comp.judges else None)
        score_map = {(s.team_id, s.criterion_id, s.judge_id): s for s in comp.scores}
        results = ContestResult.query.filter_by(competition_id=comp.id).order_by(ContestResult.rank.asc().nullslast(), ContestResult.total_score.desc()).all()
        return render_template("contests/detail.html", comp=comp, teams=teams, selected_judge_id=selected_judge_id, score_map=score_map, results=results)

    @app.route("/contests/<int:comp_id>/criteria/add", methods=["POST"])
    @login_required
    def contest_criterion_add(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขกิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        name = request.form.get("name", "").strip()
        if name:
            db.session.add(ContestCriterion(competition_id=comp.id, name=name, max_score=float(request.form.get("max_score") or 100), sort_order=safe_int(request.form.get("sort_order"))))
            db.session.commit()
            flash("เพิ่มเกณฑ์คะแนนแล้ว", "success")
        return redirect(url_for("contest_detail", comp_id=comp.id))

    @app.route("/contests/criteria/<int:criterion_id>/delete", methods=["POST"])
    @login_required
    def contest_criterion_delete(criterion_id):
        item = ContestCriterion.query.get_or_404(criterion_id)
        comp = item.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบเกณฑ์นี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(item)
        db.session.commit()
        refresh_contest_results(comp)
        flash("ลบเกณฑ์แล้ว", "info")
        return redirect(url_for("contest_detail", comp_id=comp.id))

    @app.route("/contests/<int:comp_id>/judges/add", methods=["POST"])
    @login_required
    def contest_judge_add(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขกิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        name = request.form.get("name", "").strip()
        if name:
            db.session.add(ContestJudge(competition_id=comp.id, name=name, position=request.form.get("position", "").strip()))
            db.session.commit()
            flash("เพิ่มกรรมการแล้ว", "success")
        return redirect(url_for("contest_detail", comp_id=comp.id))

    @app.route("/contests/judges/<int:judge_id>/delete", methods=["POST"])
    @login_required
    def contest_judge_delete(judge_id):
        item = ContestJudge.query.get_or_404(judge_id)
        comp = item.competition
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบกรรมการนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(item)
        db.session.commit()
        refresh_contest_results(comp)
        flash("ลบกรรมการแล้ว", "info")
        return redirect(url_for("contest_detail", comp_id=comp.id))

    @app.route("/contests/<int:comp_id>/scores/save", methods=["POST"])
    @login_required
    def contest_scores_save(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์บันทึกคะแนนกิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        judge_id = safe_int_or_none(request.form.get("judge_id"))
        judge = ContestJudge.query.get(judge_id) if judge_id else None
        if not judge or judge.competition_id != comp.id:
            flash("กรุณาเลือกกรรมการ", "danger")
            return redirect(url_for("contest_detail", comp_id=comp.id))
        teams = Team.query.filter_by(event_id=comp.event_id).all()
        criteria = list(comp.criteria)
        for team in teams:
            for criterion in criteria:
                key = f"score_{team.id}_{criterion.id}"
                if key not in request.form:
                    continue
                value = float(request.form.get(key) or 0)
                if value < 0:
                    value = 0
                if value > criterion.max_score:
                    value = criterion.max_score
                score = ContestScore.query.filter_by(competition_id=comp.id, team_id=team.id, criterion_id=criterion.id, judge_id=judge.id).first()
                if not score:
                    score = ContestScore(competition_id=comp.id, team_id=team.id, criterion_id=criterion.id, judge_id=judge.id)
                    db.session.add(score)
                score.score = value
        comp.status = "completed"
        db.session.commit()
        refresh_contest_results(comp)
        flash("บันทึกคะแนนกรรมการแล้ว", "success")
        return redirect(url_for("contest_detail", comp_id=comp.id, judge_id=judge.id))

    @app.route("/contests/<int:comp_id>/print")
    @login_required
    def contest_print(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์พิมพ์กิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        refresh_contest_results(comp)
        results = ContestResult.query.filter_by(competition_id=comp.id).order_by(ContestResult.rank.asc().nullslast(), ContestResult.total_score.desc()).all()
        return render_template("contests/print.html", comp=comp, results=results)

    @app.route("/contests/<int:comp_id>/export.xlsx")
    @login_required
    def contest_export_excel(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ Export กิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        refresh_contest_results(comp)
        output = build_contest_workbook(comp)
        safe_name = secure_filename(comp.name) or f"contest_{comp.id}"
        return send_file(output, as_attachment=True, download_name=f"{safe_name}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/contests/<int:comp_id>/delete", methods=["POST"])
    @login_required
    def contest_delete(comp_id):
        comp = ContestCompetition.query.get_or_404(comp_id)
        event_id = comp.event_id
        if not can_access_org(comp.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบกิจกรรมนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(comp)
        db.session.commit()
        flash("ลบกิจกรรมประกวดแล้ว", "info")
        return redirect(url_for("event_contests", event_id=event_id))


    @app.route("/events/<int:event_id>/certificates")
    @login_required
    def event_certificates(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_certificates", "Certificate")
        if denied:
            return denied
        templates = CertificateTemplate.query.filter_by(event_id=event.id).order_by(CertificateTemplate.created_at.desc()).all()
        recipients = CertificateRecipient.query.filter_by(event_id=event.id).order_by(CertificateRecipient.issued_at.desc()).limit(80).all()
        return render_template("certificates/list.html", event=event, templates=templates, recipients=recipients)

    @app.route("/events/<int:event_id>/certificates/templates/new", methods=["GET", "POST"])
    @login_required
    def certificate_template_new(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_certificates", "Certificate")
        if denied:
            return denied
        if request.method == "POST":
            tpl = CertificateTemplate(event_id=event.id)
            fill_certificate_template_from_form(tpl)
            db.session.add(tpl)
            db.session.commit()
            flash("สร้าง Template เกียรติบัตรแล้ว", "success")
            return redirect(url_for("event_certificates", event_id=event.id))
        return render_template("certificates/template_form.html", event=event, template=None)

    @app.route("/certificates/templates/<int:template_id>/edit", methods=["GET", "POST"])
    @login_required
    def certificate_template_edit(template_id):
        tpl = CertificateTemplate.query.get_or_404(template_id)
        event = tpl.event
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_certificates", "Certificate")
        if denied:
            return denied
        if request.method == "POST":
            fill_certificate_template_from_form(tpl)
            db.session.commit()
            flash("บันทึก Template เกียรติบัตรแล้ว", "success")
            return redirect(url_for("event_certificates", event_id=event.id))
        return render_template("certificates/template_form.html", event=event, template=tpl)

    @app.route("/certificates/templates/<int:template_id>/delete", methods=["POST"])
    @login_required
    def certificate_template_delete(template_id):
        tpl = CertificateTemplate.query.get_or_404(template_id)
        event_id = tpl.event_id
        if not can_access_org(tpl.event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(tpl.event, "allow_certificates", "Certificate")
        if denied:
            return denied
        db.session.delete(tpl)
        db.session.commit()
        flash("ลบ Template แล้ว", "info")
        return redirect(url_for("event_certificates", event_id=event_id))

    @app.route("/certificates/templates/<int:template_id>/generate", methods=["POST"])
    @login_required
    def certificate_generate(template_id):
        tpl = CertificateTemplate.query.get_or_404(template_id)
        event = tpl.event
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_certificates", "Certificate")
        if denied:
            return denied
        mode = request.form.get("mode", tpl.cert_type or "participant")
        created = generate_certificates_for_template(tpl, mode)
        flash(f"สร้างรายชื่อเกียรติบัตร {created} รายการ", "success")
        return redirect(url_for("event_certificates", event_id=event.id))

    @app.route("/events/<int:event_id>/certificates/manual", methods=["POST"])
    @login_required
    def certificate_manual(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_certificates", "Certificate")
        if denied:
            return denied
        template_id = int(request.form.get("template_id") or 0)
        tpl = CertificateTemplate.query.filter_by(id=template_id, event_id=event.id).first_or_404()
        names = [x.strip() for x in (request.form.get("names") or "").splitlines() if x.strip()]
        role_text = request.form.get("role_text", "").strip()
        award_text = request.form.get("award_text", "").strip()
        sport_text = request.form.get("sport_text", "").strip()
        created = 0
        for name in names:
            if create_certificate_recipient(tpl, full_name=name, recipient_type="manual", role_text=role_text, award_text=award_text, sport_text=sport_text):
                created += 1
        db.session.commit()
        flash(f"เพิ่มรายชื่อเกียรติบัตรเอง {created} รายการ", "success")
        return redirect(url_for("event_certificates", event_id=event.id))

    @app.route("/certificates/recipients/<int:recipient_id>/delete", methods=["POST"])
    @login_required
    def certificate_recipient_delete(recipient_id):
        cert = CertificateRecipient.query.get_or_404(recipient_id)
        event_id = cert.event_id
        if not can_access_org(cert.event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(cert.event, "allow_certificates", "Certificate")
        if denied:
            return denied
        db.session.delete(cert)
        db.session.commit()
        flash("ลบรายชื่อเกียรติบัตรแล้ว", "info")
        return redirect(url_for("event_certificates", event_id=event_id))

    @app.route("/certificates/search")
    def certificate_search():
        q = (request.args.get("q") or "").strip()
        results = []
        if q:
            results = CertificateRecipient.query.filter(CertificateRecipient.full_name.ilike(f"%{q}%"), CertificateRecipient.is_revoked == False).order_by(CertificateRecipient.issued_at.desc()).limit(50).all()
        return render_template("certificates/search.html", q=q, results=results)

    @app.route("/certificates/verify/<verify_code>")
    def certificate_verify(verify_code):
        cert = CertificateRecipient.query.filter_by(verify_code=verify_code).first_or_404()
        verify_url = url_for("certificate_verify", verify_code=cert.verify_code, _external=True)
        qr_data = make_qr_data_uri(verify_url)
        return render_template("certificates/verify.html", cert=cert, verify_url=verify_url, qr_data=qr_data)

    @app.route("/certificates/print/<int:recipient_id>")
    def certificate_print(recipient_id):
        cert = CertificateRecipient.query.get_or_404(recipient_id)
        verify_url = url_for("certificate_verify", verify_code=cert.verify_code, _external=True)
        qr_data = make_qr_data_uri(verify_url)
        return render_template("certificates/print.html", cert=cert, verify_url=verify_url, qr_data=qr_data)

    @app.route("/events/<int:event_id>/live-board/settings", methods=["GET", "POST"])
    @login_required
    def event_live_board_settings(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ตั้งค่า Live Board", "danger")
            return redirect(url_for("events"))
        denied = check_feature_or_redirect(event, "allow_live_board", "Live Board")
        if denied:
            return denied
        setting = get_live_board_setting(event)
        if request.method == "POST":
            setting.marquee_text = request.form.get("marquee_text", "").strip()
            setting.theme = request.form.get("theme", "stadium")
            setting.refresh_seconds = safe_int(request.form.get("refresh_seconds")) or 10
            setting.show_medals = bool(request.form.get("show_medals"))
            setting.show_schedule = bool(request.form.get("show_schedule"))
            setting.show_results = bool(request.form.get("show_results"))
            setting.show_rr_standings = bool(request.form.get("show_rr_standings"))
            db.session.commit()
            flash("บันทึกการตั้งค่า Live Board แล้ว", "success")
            return redirect(url_for("event_live_board_settings", event_id=event.id))
        return render_template("live_board/settings.html", event=event, setting=setting)

    @app.route("/events/<int:event_id>/live-board")
    @login_required
    def event_live_board(event_id):
        event = Event.query.get_or_404(event_id)
        denied = check_feature_or_redirect(event, "allow_live_board", "Live Board")
        if denied:
            return denied
        setting = get_live_board_setting(event)
        return render_template("live_board/display.html", event=event, setting=setting)

    @app.route("/events/<int:event_id>/live-board/data")
    @login_required
    def event_live_board_data(event_id):
        event = Event.query.get_or_404(event_id)
        denied = check_feature_or_redirect(event, "allow_live_board", "Live Board")
        if denied:
            return denied
        setting = get_live_board_setting(event)
        data = build_live_board_data(event, setting)
        return jsonify(data)





    @app.route("/events/<int:event_id>/medals")
    @login_required
    def event_medals(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        sport_id = safe_int_or_none(request.args.get("sport_id"))
        class_name = request.args.get("class_name", "").strip()
        gender = request.args.get("gender", "").strip()
        entries = collect_medal_entries(event, sport_id=sport_id, class_name=class_name, gender=gender)
        table = build_medal_table(event, entries)
        sports = Sport.query.filter_by(event_id=event.id).order_by(Sport.name).all()
        classes = sorted({d.class_name for sp in sports for d in sp.divisions if d.class_name})
        genders = sorted({d.gender for sp in sports for d in sp.divisions if d.gender})
        return render_template("medals/table.html", event=event, table=table, entries=entries, sports=sports, classes=classes, genders=genders, selected={"sport_id": sport_id, "class_name": class_name, "gender": gender})

    @app.route("/events/<int:event_id>/medals/export.xlsx")
    @login_required
    def event_medals_export(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์ Export ตารางเหรียญนี้", "danger")
            return redirect(url_for("events"))
        sport_id = safe_int_or_none(request.args.get("sport_id"))
        class_name = request.args.get("class_name", "").strip()
        gender = request.args.get("gender", "").strip()
        entries = collect_medal_entries(event, sport_id=sport_id, class_name=class_name, gender=gender)
        table = build_medal_table(event, entries)
        output = build_medal_workbook(event, table, entries)
        safe_name = secure_filename(event.name) or f"event_{event.id}"
        return send_file(output, as_attachment=True, download_name=f"{safe_name}_medal_table.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/events/<int:event_id>/registrations")
    @login_required
    def event_registrations(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Athlete.full_name).all()
        coaches = Coach.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Coach.full_name).all()
        return render_template("registrations/admin.html", event=event, athletes=athletes, coaches=coaches)

    @app.route("/athletes/<int:athlete_id>/status", methods=["POST"])
    @login_required
    def athlete_status(athlete_id):
        athlete = Athlete.query.get_or_404(athlete_id)
        if not can_access_org(athlete.team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        athlete.status = request.form.get("status", "pending")
        for reg in athlete.registrations:
            reg.status = athlete.status
        db.session.commit()
        flash("อัปเดตสถานะนักกีฬาแล้ว", "success")
        return redirect(url_for("event_registrations", event_id=athlete.team.event_id))

    @app.route("/coaches/<int:coach_id>/status", methods=["POST"])
    @login_required
    def coach_status(coach_id):
        coach = Coach.query.get_or_404(coach_id)
        if not can_access_org(coach.team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        coach.status = request.form.get("status", "pending")
        db.session.commit()
        flash("อัปเดตสถานะผู้ฝึกสอนแล้ว", "success")
        return redirect(url_for("event_registrations", event_id=coach.team.event_id))

    @app.route("/athletes/<int:athlete_id>/edit", methods=["GET", "POST"])
    @login_required
    def athlete_edit(athlete_id):
        athlete = Athlete.query.get_or_404(athlete_id)
        if not can_access_org(athlete.team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        if request.method == "POST":
            save_athlete_from_request(athlete)
            db.session.commit()
            flash("บันทึกข้อมูลนักกีฬาแล้ว", "success")
            return redirect(url_for("event_registrations", event_id=athlete.team.event_id))
        return render_template("registrations/athlete_form.html", athlete=athlete, team=athlete.team, public_mode=False, sport_divisions=get_event_sport_divisions(athlete.team.event_id))

    @app.route("/coaches/<int:coach_id>/edit", methods=["GET", "POST"])
    @login_required
    def coach_edit(coach_id):
        coach = Coach.query.get_or_404(coach_id)
        if not can_access_org(coach.team.event.organization_id):
            flash("คุณไม่มีสิทธิ์แก้ไขข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        if request.method == "POST":
            save_coach_from_request(coach)
            db.session.commit()
            flash("บันทึกข้อมูลผู้ฝึกสอนแล้ว", "success")
            return redirect(url_for("event_registrations", event_id=coach.team.event_id))
        return render_template("registrations/coach_form.html", coach=coach, team=coach.team, public_mode=False)

    @app.route("/athletes/<int:athlete_id>/delete", methods=["POST"])
    @login_required
    def athlete_delete(athlete_id):
        athlete = Athlete.query.get_or_404(athlete_id)
        event_id = athlete.team.event_id
        if not can_access_org(athlete.team.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(athlete)
        db.session.commit()
        flash("ลบนักกีฬาแล้ว", "info")
        return redirect(url_for("event_registrations", event_id=event_id))

    @app.route("/coaches/<int:coach_id>/delete", methods=["POST"])
    @login_required
    def coach_delete(coach_id):
        coach = Coach.query.get_or_404(coach_id)
        event_id = coach.team.event_id
        if not can_access_org(coach.team.event.organization_id):
            flash("คุณไม่มีสิทธิ์ลบข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        db.session.delete(coach)
        db.session.commit()
        flash("ลบผู้ฝึกสอนแล้ว", "info")
        return redirect(url_for("event_registrations", event_id=event_id))

    @app.route("/events/<int:event_id>/registrations/export.xlsx")
    @login_required
    def registrations_export_excel(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        output = build_registration_workbook(event)
        return send_file(output, as_attachment=True, download_name=f"registrations_event_{event.id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/events/<int:event_id>/registrations/print")
    @login_required
    def registrations_print(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("events"))
        athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Athlete.full_name).all()
        coaches = Coach.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Coach.full_name).all()
        return render_template("registrations/print.html", event=event, athletes=athletes, coaches=coaches)

    @app.route("/events/<int:event_id>/registrations/import", methods=["GET", "POST"])
    @login_required
    def registrations_import(event_id):
        event = Event.query.get_or_404(event_id)
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์นำเข้าข้อมูลนี้", "danger")
            return redirect(url_for("events"))
        if request.method == "POST":
            file = request.files.get("file")
            if not file or not file.filename:
                flash("กรุณาเลือกไฟล์ Excel", "danger")
                return render_template("registrations/import.html", event=event)
            count = import_registrations_from_excel(event, file)
            flash(f"นำเข้าข้อมูลแล้ว {count} รายการ", "success")
            return redirect(url_for("event_registrations", event_id=event.id))
        return render_template("registrations/import.html", event=event)

    @app.route("/team-portal/<int:team_id>/athletes")
    def portal_athletes(team_id):
        team = require_team_portal_access(team_id)
        if not team:
            return redirect(url_for("team_entry"))
        athletes = Athlete.query.filter_by(team_id=team.id).order_by(Athlete.created_at.desc()).all()
        coaches = Coach.query.filter_by(team_id=team.id).order_by(Coach.created_at.desc()).all()
        return render_template("registrations/portal_list.html", team=team, athletes=athletes, coaches=coaches)

    @app.route("/team-portal/<int:team_id>/athletes/new", methods=["GET", "POST"])
    def portal_athlete_new(team_id):
        team = require_team_portal_access(team_id)
        if not team:
            return redirect(url_for("team_entry"))
        if request.method == "POST":
            ok, msg = check_athlete_limit(team.event)
            if not ok:
                return deny_upgrade(msg, "team_entry")
            athlete = Athlete(team_id=team.id)
            save_athlete_from_request(athlete)
            db.session.add(athlete)
            db.session.flush()
            save_athlete_registrations(athlete)
            db.session.commit()
            flash("เพิ่มนักกีฬาแล้ว รอแอดมินตรวจสอบ", "success")
            return redirect(url_for("portal_athletes", team_id=team.id))
        return render_template("registrations/athlete_form.html", athlete=None, team=team, public_mode=True, sport_divisions=get_event_sport_divisions(team.event_id))

    @app.route("/team-portal/athletes/<int:athlete_id>/edit", methods=["GET", "POST"])
    def portal_athlete_edit(athlete_id):
        athlete = Athlete.query.get_or_404(athlete_id)
        team = require_team_portal_access(athlete.team_id)
        if not team:
            return redirect(url_for("team_entry"))
        if request.method == "POST":
            save_athlete_from_request(athlete)
            athlete.status = "pending"
            athlete.registrations.clear()
            db.session.flush()
            save_athlete_registrations(athlete)
            db.session.commit()
            flash("แก้ไขนักกีฬาแล้ว รอแอดมินตรวจสอบใหม่", "success")
            return redirect(url_for("portal_athletes", team_id=team.id))
        return render_template("registrations/athlete_form.html", athlete=athlete, team=team, public_mode=True, sport_divisions=get_event_sport_divisions(team.event_id))

    @app.route("/team-portal/athletes/<int:athlete_id>/delete", methods=["POST"])
    def portal_athlete_delete(athlete_id):
        athlete = Athlete.query.get_or_404(athlete_id)
        team = require_team_portal_access(athlete.team_id)
        if not team:
            return redirect(url_for("team_entry"))
        db.session.delete(athlete)
        db.session.commit()
        flash("ลบนักกีฬาแล้ว", "info")
        return redirect(url_for("portal_athletes", team_id=team.id))

    @app.route("/team-portal/<int:team_id>/coaches/new", methods=["GET", "POST"])
    def portal_coach_new(team_id):
        team = require_team_portal_access(team_id)
        if not team:
            return redirect(url_for("team_entry"))
        if request.method == "POST":
            coach = Coach(team_id=team.id)
            save_coach_from_request(coach)
            db.session.add(coach)
            db.session.commit()
            flash("เพิ่มผู้ฝึกสอนแล้ว รอแอดมินตรวจสอบ", "success")
            return redirect(url_for("portal_athletes", team_id=team.id))
        return render_template("registrations/coach_form.html", coach=None, team=team, public_mode=True)

    @app.route("/team-portal/coaches/<int:coach_id>/edit", methods=["GET", "POST"])
    def portal_coach_edit(coach_id):
        coach = Coach.query.get_or_404(coach_id)
        team = require_team_portal_access(coach.team_id)
        if not team:
            return redirect(url_for("team_entry"))
        if request.method == "POST":
            save_coach_from_request(coach)
            coach.status = "pending"
            db.session.commit()
            flash("แก้ไขผู้ฝึกสอนแล้ว รอแอดมินตรวจสอบใหม่", "success")
            return redirect(url_for("portal_athletes", team_id=team.id))
        return render_template("registrations/coach_form.html", coach=coach, team=team, public_mode=True)

    @app.route("/team-portal/coaches/<int:coach_id>/delete", methods=["POST"])
    def portal_coach_delete(coach_id):
        coach = Coach.query.get_or_404(coach_id)
        team = require_team_portal_access(coach.team_id)
        if not team:
            return redirect(url_for("team_entry"))
        db.session.delete(coach)
        db.session.commit()
        flash("ลบผู้ฝึกสอนแล้ว", "info")
        return redirect(url_for("portal_athletes", team_id=team.id))


    @app.route("/admin/subscription-plans", methods=["GET", "POST"])
    @login_required
    @superadmin_required
    def admin_subscription_plans():
        if request.method == "POST":
            plan = SubscriptionPlan(
                code=request.form.get("code", "").strip().lower(),
                name=request.form.get("name", "").strip(),
                description=request.form.get("description", "").strip(),
                price=float(request.form.get("price") or 0),
                billing_period=request.form.get("billing_period", "monthly"),
                duration_days=int(request.form.get("duration_days") or 30),
                max_events=int(request.form.get("max_events") or 0),
                max_teams_per_event=int(request.form.get("max_teams_per_event") or 0),
                max_athletes_per_event=int(request.form.get("max_athletes_per_event") or 0),
                allow_live_board=bool(request.form.get("allow_live_board")),
                allow_certificates=bool(request.form.get("allow_certificates")),
                allow_reports_pdf=bool(request.form.get("allow_reports_pdf")),
                is_active=bool(request.form.get("is_active", "1")),
                sort_order=int(request.form.get("sort_order") or 0),
            )
            if not plan.code or not plan.name:
                flash("กรุณากรอกรหัสและชื่อแพ็กเกจ", "danger")
            elif SubscriptionPlan.query.filter_by(code=plan.code).first():
                flash("รหัสแพ็กเกจนี้มีแล้ว", "danger")
            else:
                db.session.add(plan)
                db.session.commit()
                flash("เพิ่มแพ็กเกจแล้ว", "success")
            return redirect(url_for("admin_subscription_plans"))
        plans = SubscriptionPlan.query.order_by(SubscriptionPlan.sort_order, SubscriptionPlan.id).all()
        return render_template("billing/admin_plans.html", plans=plans)

    @app.route("/admin/subscription-plans/<int:plan_id>/edit", methods=["GET", "POST"])
    @login_required
    @superadmin_required
    def admin_subscription_plan_edit(plan_id):
        plan = SubscriptionPlan.query.get_or_404(plan_id)
        if request.method == "POST":
            plan.name = request.form.get("name", "").strip()
            plan.description = request.form.get("description", "").strip()
            plan.price = float(request.form.get("price") or 0)
            plan.billing_period = request.form.get("billing_period", "monthly")
            plan.duration_days = int(request.form.get("duration_days") or 30)
            plan.max_events = int(request.form.get("max_events") or 0)
            plan.max_teams_per_event = int(request.form.get("max_teams_per_event") or 0)
            plan.max_athletes_per_event = int(request.form.get("max_athletes_per_event") or 0)
            plan.allow_live_board = bool(request.form.get("allow_live_board"))
            plan.allow_certificates = bool(request.form.get("allow_certificates"))
            plan.allow_reports_pdf = bool(request.form.get("allow_reports_pdf"))
            plan.is_active = bool(request.form.get("is_active"))
            plan.sort_order = int(request.form.get("sort_order") or 0)
            db.session.commit()
            flash("บันทึกแพ็กเกจแล้ว", "success")
            return redirect(url_for("admin_subscription_plans"))
        return render_template("billing/plan_form.html", plan=plan)

    @app.route("/organizations/<int:org_id>/billing", methods=["GET", "POST"])
    @login_required
    def organizations_billing(org_id):
        org = Organization.query.get_or_404(org_id)
        if not can_access_org(org.id):
            flash("คุณไม่มีสิทธิ์เข้าถึง Billing ขององค์กรนี้", "danger")
            return redirect(url_for("organizations"))
        if ensure_free_subscription(org):
            db.session.commit()
        if request.method == "POST":
            action = request.form.get("action")
            if action == "request_plan":
                plan = SubscriptionPlan.query.get_or_404(int(request.form.get("plan_id") or 0))
                sub = OrganizationSubscription(
                    organization_id=org.id,
                    plan_id=plan.id,
                    status="pending_payment",
                    start_date=date.today(),
                    end_date=date.today() + timedelta(days=plan.duration_days),
                    manual_payment_note=request.form.get("manual_payment_note", "").strip(),
                )
                db.session.add(sub)
                db.session.flush()
                invoice = Invoice(
                    organization_id=org.id,
                    subscription_id=sub.id,
                    invoice_no=make_invoice_no(org.id),
                    title=f"ค่าแพ็กเกจ {plan.name}",
                    amount=plan.price,
                    currency=plan.currency,
                    due_date=date.today() + timedelta(days=7),
                    status="paid" if plan.price <= 0 else "unpaid",
                    payment_method=request.form.get("payment_gateway", "manual"),
                    note="Payment Gateway / รอตรวจสอบการชำระเงิน",
                )
                if plan.price <= 0:
                    invoice.paid_at = datetime.utcnow()
                    sub.status = "active"
                db.session.add(invoice)
                db.session.commit()
                flash("สร้างใบแจ้งชำระเงินแล้ว" if plan.price > 0 else "เปลี่ยนแพ็กเกจแล้ว", "success")
                return redirect(url_for("organizations_billing", org_id=org.id))
            if action in ("activate_subscription", "mark_invoice_paid", "cancel_subscription") and not current_user.is_superadmin:
                flash("รายการนี้ต้องให้ Super Admin ดำเนินการ", "danger")
                return redirect(url_for("organizations_billing", org_id=org.id))
            if action == "activate_subscription":
                sub = OrganizationSubscription.query.get_or_404(int(request.form.get("subscription_id") or 0))
                if sub.organization_id != org.id:
                    flash("รายการไม่ตรงกับองค์กร", "danger")
                    return redirect(url_for("organizations_billing", org_id=org.id))
                OrganizationSubscription.query.filter(OrganizationSubscription.organization_id == org.id, OrganizationSubscription.id != sub.id, OrganizationSubscription.status == "active").update({"status": "cancelled"})
                sub.status = "active"
                if not sub.start_date:
                    sub.start_date = date.today()
                if not sub.end_date:
                    sub.end_date = date.today() + timedelta(days=sub.plan.duration_days)
                for inv in sub.invoices:
                    inv.status = "paid"
                    inv.paid_at = inv.paid_at or datetime.utcnow()
                db.session.commit()
                flash("เปิดใช้งานแพ็กเกจให้องค์กรแล้ว", "success")
                return redirect(url_for("organizations_billing", org_id=org.id))
            if action == "mark_invoice_paid":
                invoice = Invoice.query.get_or_404(int(request.form.get("invoice_id") or 0))
                if invoice.organization_id != org.id:
                    flash("ใบแจ้งหนี้ไม่ตรงกับองค์กร", "danger")
                    return redirect(url_for("organizations_billing", org_id=org.id))
                invoice.status = "paid"
                invoice.paid_at = datetime.utcnow()
                if invoice.subscription:
                    invoice.subscription.status = "active"
                    OrganizationSubscription.query.filter(OrganizationSubscription.organization_id == org.id, OrganizationSubscription.id != invoice.subscription.id, OrganizationSubscription.status == "active").update({"status": "cancelled"})
                db.session.commit()
                flash("บันทึกรับชำระเงินแล้ว", "success")
                return redirect(url_for("organizations_billing", org_id=org.id))
            if action == "cancel_subscription":
                sub = OrganizationSubscription.query.get_or_404(int(request.form.get("subscription_id") or 0))
                if sub.organization_id == org.id:
                    sub.status = "cancelled"
                    db.session.commit()
                    flash("ยกเลิกแพ็กเกจแล้ว", "info")
                return redirect(url_for("organizations_billing", org_id=org.id))
        current_sub = get_current_subscription(org)
        current_plan = get_current_plan(org)
        plans = SubscriptionPlan.query.filter_by(is_active=True).order_by(SubscriptionPlan.sort_order, SubscriptionPlan.id).all()
        subscriptions = OrganizationSubscription.query.filter_by(organization_id=org.id).order_by(OrganizationSubscription.created_at.desc()).all()
        invoices = Invoice.query.filter_by(organization_id=org.id).order_by(Invoice.created_at.desc()).all()
        return render_template("billing/organization_billing.html", org=org, current_sub=current_sub, current_plan=current_plan, plans=plans, subscriptions=subscriptions, invoices=invoices)

    @app.route("/invoices/<int:invoice_id>/print")
    @login_required
    def invoice_print(invoice_id):
        invoice = Invoice.query.get_or_404(invoice_id)
        if not can_access_org(invoice.organization_id):
            flash("คุณไม่มีสิทธิ์ดูใบแจ้งชำระเงินนี้", "danger")
            return redirect(url_for("organizations"))
        return render_template("billing/invoice_print.html", invoice=invoice)

    @app.route("/invoices/<int:invoice_id>/pay", methods=["GET", "POST"])
    @login_required
    def invoice_pay(invoice_id):
        invoice = Invoice.query.get_or_404(invoice_id)
        if not can_access_org(invoice.organization_id):
            flash("คุณไม่มีสิทธิ์ชำระใบแจ้งหนี้นี้", "danger")
            return redirect(url_for("organizations"))
        if invoice.status == "paid":
            flash("ใบแจ้งหนี้นี้ชำระแล้ว", "info")
            return redirect(url_for("organizations_billing", org_id=invoice.organization_id))
        gateway = request.form.get("gateway") if request.method == "POST" else (request.args.get("gateway") or invoice.payment_method or "manual")
        if gateway not in current_app_payment_gateways():
            flash("ยังไม่ได้เปิด Payment Gateway นี้", "warning")
            return redirect(url_for("organizations_billing", org_id=invoice.organization_id))

        txn = PaymentTransaction(
            organization_id=invoice.organization_id,
            invoice_id=invoice.id,
            gateway=gateway,
            amount=invoice.amount,
            currency=invoice.currency,
            status="pending",
            provider_reference=f"KS-{invoice.invoice_no}-{uuid.uuid4().hex[:8]}",
        )
        if gateway == "promptpay":
            promptpay_id = current_app.config.get("PROMPTPAY_ID")
            if not promptpay_id:
                flash("ยังไม่ได้ตั้งค่า PROMPTPAY_ID ใน .env", "danger")
                return redirect(url_for("organizations_billing", org_id=invoice.organization_id))
            txn.qr_payload = build_promptpay_payload(promptpay_id, invoice.amount)
            txn.note = "สแกนจ่าย PromptPay แล้วให้แอดมินตรวจหลักฐาน/กดรับชำระ"
        elif gateway == "manual":
            txn.note = "โอนเงิน/ชำระเงินนอกระบบ แล้วให้ Super Admin กดรับชำระ"
        elif gateway == "stripe":
            txn.checkout_url = create_stripe_checkout_url(invoice, txn.provider_reference)
        elif gateway == "omise":
            txn.note = "เตรียมข้อมูล Transaction สำหรับ Omise / ต้องต่อ Secret Key และ Webhook เพิ่ม"
        db.session.add(txn)
        invoice.payment_method = gateway
        db.session.commit()

        if gateway == "stripe" and txn.checkout_url:
            return redirect(txn.checkout_url)
        return redirect(url_for("payment_transaction", txn_id=txn.id))

    @app.route("/payments/<int:txn_id>")
    @login_required
    def payment_transaction(txn_id):
        txn = PaymentTransaction.query.get_or_404(txn_id)
        if not can_access_org(txn.organization_id):
            flash("คุณไม่มีสิทธิ์ดูรายการชำระเงินนี้", "danger")
            return redirect(url_for("organizations"))
        rebuilt = ensure_promptpay_payload_for_transaction(txn)
        qr_data, qr_error = make_promptpay_qr_data_uri(txn.qr_payload) if txn.qr_payload else (None, "ไม่พบข้อมูล PromptPay payload")
        if rebuilt:
            flash("ระบบซ่อม PromptPay QR เดิมให้เป็นรูปแบบที่แอปธนาคารสแกนได้แล้ว", "success")
        return render_template("billing/payment_transaction.html", txn=txn, qr_data=qr_data, qr_error=qr_error)

    @app.route("/payments/<int:txn_id>/promptpay-qr.png")
    @login_required
    def payment_promptpay_qr_png(txn_id):
        txn = PaymentTransaction.query.get_or_404(txn_id)
        if not can_access_org(txn.organization_id):
            flash("คุณไม่มีสิทธิ์ดูรายการชำระเงินนี้", "danger")
            return redirect(url_for("organizations"))
        ensure_promptpay_payload_for_transaction(txn)
        png_bytes, error = make_promptpay_qr_png(txn.qr_payload)
        if not png_bytes:
            return error or "สร้าง QR ไม่สำเร็จ", 500
        return send_file(BytesIO(png_bytes), mimetype="image/png", download_name=f"promptpay-{txn.provider_reference or txn.id}.png")

    @app.route("/payments/<int:txn_id>/mark-paid", methods=["POST"])
    @login_required
    @superadmin_required
    def payment_mark_paid(txn_id):
        txn = PaymentTransaction.query.get_or_404(txn_id)
        txn.status = "paid"
        txn.paid_at = datetime.utcnow()
        mark_invoice_paid(txn.invoice, gateway=txn.gateway, reference=txn.provider_reference)
        flash("ยืนยันรับชำระเงินแล้ว", "success")
        return redirect(url_for("organizations_billing", org_id=txn.organization_id))

    @app.route("/payments/webhook/stripe", methods=["POST"])
    def stripe_webhook():
        payload = request.get_data(as_text=True)
        sig_header = request.headers.get("Stripe-Signature", "")
        secret = os.getenv("STRIPE_WEBHOOK_SECRET", "")
        try:
            import stripe
            if secret:
                event = stripe.Webhook.construct_event(payload, sig_header, secret)
            else:
                event = json.loads(payload)
            if event.get("type") == "checkout.session.completed":
                session_obj = event.get("data", {}).get("object", {})
                ref = (session_obj.get("metadata") or {}).get("provider_reference")
                txn = PaymentTransaction.query.filter_by(provider_reference=ref).first() if ref else None
                if txn:
                    txn.status = "paid"
                    txn.paid_at = datetime.utcnow()
                    txn.raw_response = payload
                    mark_invoice_paid(txn.invoice, gateway="stripe", reference=ref)
            return jsonify({"ok": True})
        except Exception as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

    @app.route("/settings")
    @login_required
    def settings():
        org_ids = user_org_ids()
        event_query = Event.query
        org_query = Organization.query
        if org_ids is not None:
            event_query = event_query.filter(Event.organization_id.in_(org_ids or [0]))
            org_query = org_query.filter(Organization.id.in_(org_ids or [0]))
        events_list = event_query.order_by(Event.created_at.desc()).all()
        orgs = org_query.order_by(Organization.created_at.desc()).all()
        active_event = Event.query.get(session.get("active_event_id")) if session.get("active_event_id") else None
        active_org = Organization.query.get(session.get("active_org_id")) if session.get("active_org_id") else None
        return render_template("settings.html", events=events_list, organizations=orgs, active_event=active_event, active_org=active_org)

    @app.route("/settings/active-event", methods=["POST"])
    @login_required
    def settings_active_event():
        event_id = request.form.get("event_id")
        if not event_id:
            session.pop("active_event_id", None)
            flash("ล้างงานที่เลือกแล้ว", "info")
            return redirect(url_for("settings"))
        event = Event.query.get_or_404(int(event_id))
        if not can_access_org(event.organization_id):
            flash("คุณไม่มีสิทธิ์เข้าถึงงานนี้", "danger")
            return redirect(url_for("settings"))
        session["active_event_id"] = event.id
        session["active_org_id"] = event.organization_id
        flash(f"เลือกงานปัจจุบัน: {event.name}", "success")
        next_page = request.form.get("next") or "settings"
        if next_page == "sports":
            return redirect(url_for("event_sports", event_id=event.id))
        if next_page == "teams":
            return redirect(url_for("event_detail", event_id=event.id) + "#teams-section")
        return redirect(url_for("settings"))



def get_or_create_team_profile(team):
    from models import TeamProfile
    profile = TeamProfile.query.filter_by(team_id=team.id).first()
    if not profile:
        profile = TeamProfile(team_id=team.id)
        db.session.add(profile)
        db.session.flush()
    return profile

def build_report_workbook(title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = 12
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_event_medal_rows(event):
    from models import Team, RankingCompetition, RankingResult, ContestCompetition, ContestResult, RoundRobinCompetition

    table = {}
    teams = Team.query.filter_by(event_id=event.id).all()
    for team in teams:
        table[team.id] = {
            "team": team,
            "gold": 0,
            "silver": 0,
            "bronze": 0,
        }

    def add_medal(team_id, rank):
        if not team_id or team_id not in table:
            return
        if rank == 1:
            table[team_id]["gold"] += 1
        elif rank == 2:
            table[team_id]["silver"] += 1
        elif rank == 3:
            table[team_id]["bronze"] += 1

    ranking_comps = RankingCompetition.query.filter_by(event_id=event.id).all()
    for comp in ranking_comps:
        results = RankingResult.query.filter_by(competition_id=comp.id).all()
        for r in results:
            add_medal(getattr(r, "team_id", None), getattr(r, "rank", None))

    contest_comps = ContestCompetition.query.filter_by(event_id=event.id).all()
    for comp in contest_comps:
        results = ContestResult.query.filter_by(competition_id=comp.id).all()
        for r in results:
            add_medal(getattr(r, "team_id", None), getattr(r, "rank", None))

    rr_comps = RoundRobinCompetition.query.filter_by(event_id=event.id).all()
    for comp in rr_comps:
        try:
            standings = calculate_rr_standings(comp)
            flat = []
            for group_name, rows in standings.items():
                flat.extend(rows)
            flat = sorted(
                flat,
                key=lambda x: (
                    -x.get("points", 0),
                    -x.get("goal_diff", 0),
                    -x.get("goals_for", 0),
                    x.get("team").name if x.get("team") else "",
                ),
            )
            for idx, row in enumerate(flat[:3], start=1):
                team = row.get("team")
                add_medal(team.id if team else None, idx)
        except Exception:
            pass

    rows = list(table.values())
    rows.sort(key=lambda x: (-x["gold"], -x["silver"], -x["bronze"], x["team"].name))
    return rows

def save_team_profile_from_request(profile):
    profile.director_name = request.form.get("director_name", "").strip()
    uploaded_director = save_upload(request.files.get("director_photo"))
    if uploaded_director:
        profile.director_photo = uploaded_director
    profile.deputy_directors = request.form.get("deputy_directors", "").strip()
    profile.advisors = request.form.get("advisors", "").strip()
    profile.coaches_summary = request.form.get("coaches_summary", "").strip()
    profile.parade_title = request.form.get("parade_title", "").strip()
    profile.parade_concept = request.form.get("parade_concept", "").strip()
    profile.parade_description = request.form.get("parade_description", "").strip()
    profile.stand_leaders = request.form.get("stand_leaders", "").strip()
    profile.stand_member_total = safe_int(request.form.get("stand_member_total"))
    profile.stand_member_male = safe_int(request.form.get("stand_member_male"))
    profile.stand_member_female = safe_int(request.form.get("stand_member_female"))
    profile.cheerleader_summary = request.form.get("cheerleader_summary", "").strip()


def add_team_person(team):
    from models import TeamPerson
    section = request.form.get("section", "parade")
    name = request.form.get("name", "").strip()
    if not name:
        flash("กรุณากรอกชื่อ", "danger")
        return
    person = TeamPerson(
        team_id=team.id,
        section=section,
        name=name,
        role=request.form.get("role", "").strip(),
        phone=request.form.get("phone", "").strip(),
        note=request.form.get("note", "").strip(),
        sort_order=safe_int(request.form.get("sort_order")),
    )
    db.session.add(person)
    db.session.commit()
    flash("เพิ่มรายชื่อแล้ว", "success")


def add_team_file(team):
    from models import TeamFile
    file = request.files.get("file")
    filename = save_upload(file)
    if not filename:
        flash("กรุณาเลือกไฟล์", "danger")
        return
    item = TeamFile(
        team_id=team.id,
        section=request.form.get("section", "general"),
        title=request.form.get("title", "").strip() or file.filename,
        filename=filename,
        original_filename=file.filename,
        file_type=(file.mimetype or ""),
    )
    db.session.add(item)
    db.session.commit()
    flash("แนบไฟล์แล้ว", "success")



def save_athlete_from_request(athlete):
    athlete.full_name = request.form.get("full_name", "").strip()
    athlete.gender = request.form.get("gender", "ชาย")
    athlete.grade_level = request.form.get("grade_level", "").strip()
    athlete.classroom = request.form.get("classroom", "").strip()
    athlete.student_no = request.form.get("student_no", "").strip()
    athlete.phone = request.form.get("phone", "").strip()
    athlete.note = request.form.get("note", "").strip()
    uploaded = save_upload(request.files.get("photo"))
    if uploaded:
        athlete.photo = uploaded


def save_athlete_registrations(athlete):
    from models import AthleteRegistration, SportDivision
    division_ids = request.form.getlist("sport_division_id[]")
    if division_ids:
        for division_id in division_ids:
            if not division_id:
                continue
            division = SportDivision.query.get(int(division_id))
            if not division or division.sport.event_id != athlete.team.event_id:
                continue
            reg = AthleteRegistration(
                athlete_id=athlete.id,
                sport_name=division.sport.name,
                category_name=division.class_name,
                gender=division.gender,
                status=athlete.status,
            )
            db.session.add(reg)
        return

    # fallback: รองรับไฟล์/ฟอร์มเก่าที่กรอกชื่อกีฬาเอง
    sport_names = request.form.getlist("sport_name[]")
    category_names = request.form.getlist("category_name[]")
    genders = request.form.getlist("reg_gender[]")
    for idx, sport_name in enumerate(sport_names):
        sport_name = (sport_name or "").strip()
        if not sport_name:
            continue
        reg = AthleteRegistration(
            athlete_id=athlete.id,
            sport_name=sport_name,
            category_name=(category_names[idx] if idx < len(category_names) else "").strip(),
            gender=(genders[idx] if idx < len(genders) else athlete.gender) or athlete.gender,
            status=athlete.status,
        )
        db.session.add(reg)


def save_coach_from_request(coach):
    coach.full_name = request.form.get("full_name", "").strip()
    coach.phone = request.form.get("phone", "").strip()
    coach.sport_responsibility = request.form.get("sport_responsibility", "").strip()
    coach.note = request.form.get("note", "").strip()


def build_registration_workbook(event):
    from openpyxl import Workbook
    from models import Athlete, Coach, Team
    wb = Workbook()
    ws = wb.active
    ws.title = "Athletes"
    ws.append(["ทีม", "ชื่อ-สกุล", "เพศ", "ชั้น", "ห้อง", "เลขประจำตัว", "เบอร์โทร", "สถานะ", "กีฬา/รุ่น/เพศ"])
    athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Athlete.full_name).all()
    for a in athletes:
        regs = "; ".join([f"{r.sport_name} / {r.category_name or '-'} / {r.gender}" for r in a.registrations])
        ws.append([a.team.name, a.full_name, a.gender, a.grade_level, a.classroom, a.student_no, a.phone, a.status, regs])
    ws2 = wb.create_sheet("Coaches")
    ws2.append(["ทีม", "ชื่อ-สกุล", "เบอร์โทร", "กีฬาที่รับผิดชอบ", "สถานะ"])
    coaches = Coach.query.join(Team).filter(Team.event_id == event.id).order_by(Team.name, Coach.full_name).all()
    for c in coaches:
        ws2.append([c.team.name, c.full_name, c.phone, c.sport_responsibility, c.status])
    ws3 = wb.create_sheet("Import_Template")
    ws3.append(["type", "team_code", "full_name", "gender", "grade_level", "classroom", "student_no", "phone", "sport_name", "category_name", "reg_gender", "sport_responsibility"])
    ws3.append(["athlete", "TEAMCODE", "เด็กชายตัวอย่าง", "ชาย", "ป.6", "1", "12345", "", "ฟุตซอล", "ประถมปลาย", "ชาย", ""])
    ws3.append(["coach", "TEAMCODE", "ครูตัวอย่าง", "", "", "", "", "0812345678", "", "", "", "ฟุตซอล"])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_registrations_from_excel(event, file):
    from openpyxl import load_workbook
    from models import Athlete, AthleteRegistration, Coach, Team
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    count = 0
    def cell(row, name):
        pos = idx.get(name)
        if pos is None or pos >= len(row):
            return ""
        return str(row[pos].value).strip() if row[pos].value is not None else ""
    for row in ws.iter_rows(min_row=2):
        row_type = (cell(row, "type") or "athlete").lower()
        team_code = cell(row, "team_code").upper()
        team = Team.query.filter_by(event_id=event.id, access_code=team_code).first()
        if not team:
            continue
        full_name = cell(row, "full_name")
        if not full_name:
            continue
        if row_type == "coach":
            coach = Coach(team_id=team.id, full_name=full_name, phone=cell(row, "phone"), sport_responsibility=cell(row, "sport_responsibility"), status="pending")
            db.session.add(coach)
            count += 1
        else:
            ok, msg = check_athlete_limit(event)
            if not ok:
                continue
            athlete = Athlete(
                team_id=team.id,
                full_name=full_name,
                gender=cell(row, "gender") or "ชาย",
                grade_level=cell(row, "grade_level"),
                classroom=cell(row, "classroom"),
                student_no=cell(row, "student_no"),
                phone=cell(row, "phone"),
                status="pending",
            )
            db.session.add(athlete)
            db.session.flush()
            sport_name = cell(row, "sport_name")
            if sport_name:
                db.session.add(AthleteRegistration(
                    athlete_id=athlete.id,
                    sport_name=sport_name,
                    category_name=cell(row, "category_name"),
                    gender=cell(row, "reg_gender") or athlete.gender,
                    status="pending",
                ))
            count += 1
    db.session.commit()
    return count

def require_team_portal_access(team_id):
    from models import Team
    team = Team.query.get_or_404(team_id)
    if session.get(f"team_access_{team.id}") != team.access_code:
        flash("กรุณากรอกรหัสทีมก่อน", "warning")
        return None
    if not team.registration_open:
        flash("ทีมนี้ถูกปิดสิทธิ์กรอกข้อมูลแล้ว", "warning")
        return None
    return team


def render_team_profile_page(team, profile, public_mode=False):
    people = {key: [] for key in ["executive", "advisor", "coach", "parade", "stand", "cheerleader"]}
    for person in sorted(team.people, key=lambda x: (x.section, x.sort_order or 0, x.id)):
        people.setdefault(person.section, []).append(person)
    files = {key: [] for key in ["general", "parade", "stand", "cheerleader"]}
    for item in sorted(team.files, key=lambda x: (x.section, x.id)):
        files.setdefault(item.section, []).append(item)
    return render_template(
        "teams/profile.html",
        team=team,
        event=team.event,
        profile=profile,
        people=people,
        files=files,
        public_mode=public_mode,
    )


def safe_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0

def generate_team_code(event_id):
    from models import Team
    alphabet = string.ascii_uppercase + string.digits
    for _ in range(50):
        code = "TEAM" + "".join(random.choice(alphabet) for _ in range(5))
        if not Team.query.filter_by(event_id=event_id, access_code=code).first():
            return code
    return "TEAM" + datetime.utcnow().strftime("%H%M%S")


def get_event_sport_divisions(event_id):
    from models import Sport, SportDivision
    return (SportDivision.query.join(Sport)
            .filter(Sport.event_id == event_id, Sport.is_active == True, SportDivision.is_active == True)
            .order_by(Sport.name, SportDivision.class_name, SportDivision.gender)
            .all())


def safe_int_or_none(value):
    try:
        if value in (None, ""):
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def seed_default_sports(event, force_update=True, reset_catalog=True):
    """ล้างคลังกีฬาเดิมของอีเว้นท์ แล้วสร้างชุดมาตรฐานใหม่ที่ตรวจจากกติกากลางล่าสุด

    แนวคิดใหม่: คลังมีเฉพาะ "ชนิดกีฬา + ค่าเริ่มต้นกติกา" ไม่สร้างรายการย่อยจำนวนมากตั้งแต่ต้น
    การแข่งขันจริงให้ไปสร้างที่หน้า /competitions/quick-new แล้วเลือกรุ่น/ทีม/ระบบแข่งขันตอนนั้น
    """
    from models import (
        SportCategory, Sport, SportDivision,
        RoundRobinCompetition, KnockoutCompetition, RankingCompetition, ContestCompetition,
    )

    if reset_catalog:
        # ป้องกัน FK ค้าง หากเคยสร้างการแข่งขันผูกกับ division เก่าไว้
        RoundRobinCompetition.query.filter_by(event_id=event.id).update({RoundRobinCompetition.sport_division_id: None}, synchronize_session=False)
        KnockoutCompetition.query.filter_by(event_id=event.id).update({KnockoutCompetition.sport_division_id: None}, synchronize_session=False)
        RankingCompetition.query.filter_by(event_id=event.id).update({RankingCompetition.sport_division_id: None}, synchronize_session=False)
        ContestCompetition.query.filter_by(event_id=event.id).update({ContestCompetition.sport_division_id: None}, synchronize_session=False)
        SportCategory.query.filter_by(event_id=event.id).delete(synchronize_session=False)
        db.session.flush()

    category_names = ["กีฬาทีม", "กีฬาเฉพาะทาง", "กรีฑา", "กีฬาพื้นบ้าน", "กิจกรรมประกวด"]
    categories = {}
    for i, name in enumerate(category_names, start=1):
        cat = SportCategory.query.filter_by(event_id=event.id, name=name).first()
        if not cat:
            cat = SportCategory(event_id=event.id, name=name, sort_order=i)
            db.session.add(cat)
            db.session.flush()
        categories[name] = cat

    # ค่ากติกากลางตรวจล่าสุด: 2026-06-19
    # หมายเหตุ: schema ปัจจุบันยังไม่มีช่อง decisive_set_points/deuce_cap จึงเก็บรายละเอียดไว้ใน note ก่อน
    presets = [
        {"name": "ฟุตบอล", "category": "กีฬาทีม", "format": "round_robin", "result": "score_only", "max_sets": 0, "points": 0, "win": 0, "note": "IFAB Laws of the Game 2025/26: บันทึกผลเป็นประตูได้-เสีย ระยะเวลาแข่งขันปรับตามระเบียบงานได้ | checked 2026-06-19"},
        {"name": "ฟุตซอล", "category": "กีฬาทีม", "format": "round_robin", "result": "score_only", "max_sets": 0, "points": 0, "win": 0, "note": "FIFA Futsal Laws of the Game 2025/26: ปกติ 2 ครึ่ง ครึ่งละ 20 นาที บันทึกผลเป็นประตูได้-เสีย | checked 2026-06-19"},
        {"name": "วอลเลย์บอล", "category": "กีฬาทีม", "format": "round_robin", "result": "set_based", "max_sets": 5, "points": 25, "win": 3, "note": "FIVB Official Volleyball Rules 2025-2028: ชนะ 3 เซต; เซต 1-4 ถึง 25 แต้ม ต้องห่าง 2; เซตตัดสินถึง 15 แต้ม | checked 2026-06-19"},
        {"name": "เซปักตะกร้อ", "category": "กีฬาทีม", "format": "round_robin", "result": "set_based", "max_sets": 3, "points": 15, "win": 2, "note": "ISTAF Law of the Game 2024: ชนะ 2 ใน 3 เซต; เซตละ 15 แต้ม; 14-14 เล่นถึง 17 แต้ม | checked 2026-06-19"},
        {"name": "เปตอง", "category": "กีฬาเฉพาะทาง", "format": "knockout", "result": "score_only", "max_sets": 0, "points": 13, "win": 0, "note": "FIPJP Official Rules: เกมปกติถึง 13 คะแนน; ลีก/รอบคัดเลือกอาจกำหนด 11 คะแนนตามระเบียบงาน | checked 2026-06-19"},
        {"name": "แบดมินตัน", "category": "กีฬาเฉพาะทาง", "format": "knockout", "result": "set_based", "max_sets": 3, "points": 21, "win": 2, "note": "BWF Laws ปัจจุบัน: 2 ใน 3 เกม เกมละ 21 แต้ม; BWF อนุมัติ 3x15 เริ่ม 4 ม.ค. 2027 | checked 2026-06-19"},
        {"name": "เทเบิลเทนนิส", "category": "กีฬาเฉพาะทาง", "format": "knockout", "result": "set_based", "max_sets": 5, "points": 11, "win": 3, "note": "ITTF Laws: เกมละ 11 แต้ม ต้องชนะห่าง 2; ค่าเริ่มต้นระบบใช้ 3 ใน 5 เกม | checked 2026-06-19"},
        {"name": "บาสเกตบอล", "category": "กีฬาทีม", "format": "round_robin", "result": "score_only", "max_sets": 0, "points": 0, "win": 0, "note": "FIBA: บันทึกคะแนนรวมของเกม; งานโรงเรียนปรับเวลาควอเตอร์ตามระเบียบงานได้ | checked 2026-06-19"},
        {"name": "กรีฑา", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "World Athletics: ใช้ผลเวลา/ระยะ/อันดับ ไม่ใช่ระบบเซต | checked 2026-06-19"},
        {"name": "วิ่ง 50 เมตร", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา | checked 2026-06-19"},
        {"name": "วิ่ง 100 เมตร", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา | checked 2026-06-19"},
        {"name": "วิ่งผลัด", "category": "กรีฑา", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา | checked 2026-06-19"},
        {"name": "ชักเย่อ", "category": "กีฬาพื้นบ้าน", "format": "knockout", "result": "set_based", "max_sets": 3, "points": 0, "win": 2, "note": "กีฬาโรงเรียน/กีฬาพื้นบ้าน: ค่าเริ่มต้นชนะ 2 ใน 3 เที่ยว ปรับเองได้ตามระเบียบงาน | checked 2026-06-19"},
        {"name": "วิ่งกระสอบ", "category": "กีฬาพื้นบ้าน", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา | checked 2026-06-19"},
        {"name": "วิ่งสามขา", "category": "กีฬาพื้นบ้าน", "format": "ranking", "result": "ranking", "max_sets": 0, "points": 0, "win": 0, "note": "บันทึกอันดับหรือเวลา | checked 2026-06-19"},
        {"name": "ประกวดกองเชียร์", "category": "กิจกรรมประกวด", "format": "score_judging", "result": "contest", "max_sets": 0, "points": 0, "win": 0, "note": "ใช้กรรมการให้คะแนน ปรับเกณฑ์ได้ตามงาน | checked 2026-06-19"},
    ]

    for index, rec in enumerate(presets, start=1):
        sport = Sport.query.filter_by(event_id=event.id, name=rec["name"]).first()
        data = dict(
            category_id=categories[rec["category"]].id,
            default_format=rec["format"],
            result_type=rec["result"],
            max_sets=rec["max_sets"] if rec["result"] == "set_based" else 0,
            points_per_set=rec["points"] if rec["result"] == "set_based" else 0,
            sets_to_win=rec["win"] if rec["result"] == "set_based" else 0,
            note=rec["note"],
            is_active=True,
        )
        if not sport:
            sport = Sport(event_id=event.id, name=rec["name"], **data)
            db.session.add(sport)
            db.session.flush()
        else:
            for k, v in data.items():
                setattr(sport, k, v)

        # สร้างรายการย่อยเริ่มต้นให้น้อยที่สุดเท่านั้น เพื่อไม่ให้หน้าอีเว้นท์รกเป็นร้อยรายการ
        div = SportDivision.query.filter_by(sport_id=sport.id, class_name="Open", gender="รวม").first()
        if not div:
            div = SportDivision(sport_id=sport.id, class_name="Open", gender="รวม")
            db.session.add(div)
        div.competition_format = rec["format"]
        div.result_type = rec["result"]
        div.max_sets = data["max_sets"]
        div.points_per_set = data["points_per_set"]
        div.sets_to_win = data["sets_to_win"]
        div.max_athletes_per_team = None
        div.is_active = True

    db.session.commit()


def normalize_event_sport_settings(event):
    """ซ่อม settings ของกีฬาที่มีอยู่แล้วจากข้อมูลเก่าให้ตรงชนิดกีฬา"""
    seed_default_sports(event, force_update=True)


def knockout_round_name(team_count):
    if team_count <= 2:
        return "ชิงชนะเลิศ"
    if team_count <= 4:
        return "รอบรองชนะเลิศ"
    if team_count <= 8:
        return "รอบ 8 ทีม"
    if team_count <= 16:
        return "รอบ 16 ทีม"
    return f"รอบ {team_count} ทีม"


def create_knockout_first_round(comp, teams, pairing_mode="balanced"):
    """สร้างคู่ Knockout รอบแรก
    - balanced: 1 พบท้ายสุด / 2 พบรองท้ายสุด เหมาะกับการ seed
    - adjacent: ทีมที่อยู่ติดกันเจอกัน 1-2, 3-4 แล้วผู้ชนะสายติดกันเจอกันต่อ ใช้ง่ายกับกีฬาสี/อบต.
    """
    from models import KnockoutMatch
    ordered = list(teams)
    round_name = knockout_round_name(len(ordered))

    pairs = []
    if pairing_mode == "adjacent":
        for i in range(0, len(ordered), 2):
            team_a = ordered[i] if i < len(ordered) else None
            team_b = ordered[i + 1] if i + 1 < len(ordered) else None
            pairs.append((team_a, team_b))
    else:
        pair_count = (len(ordered) + 1) // 2
        for i in range(pair_count):
            team_a = ordered[i]
            team_b = ordered[-(i + 1)] if i != len(ordered) - (i + 1) else None
            pairs.append((team_a, team_b))

    for i, (team_a, team_b) in enumerate(pairs, start=1):
        match = KnockoutMatch(
            competition_id=comp.id,
            round_no=1,
            round_name=round_name,
            match_no=i,
            team_a_id=team_a.id if team_a else None,
            team_b_id=team_b.id if team_b else None,
            status="scheduled",
        )
        if team_a and not team_b:
            match.winner_team_id = team_a.id
            match.status = "completed"
        db.session.add(match)
    db.session.flush()


def build_rr_playoff_seed_order(comp, standings):
    """จัดลำดับทีมเข้ารอบให้ไขว้สายแบบกีฬาทั่วไป ก่อนส่งเข้า Knockout แบบ adjacent
    ตัวอย่าง 2 กลุ่ม อันดับ 1-2: A1-B2, B1-A2
    ตัวอย่าง 4 กลุ่ม อันดับ 1-2: A1-B2, B1-A2, C1-D2, D1-C2
    ถ้าเป็นสูตรอื่นหรือมี best runner-up ให้ fallback ตามอันดับที่คำนวณไว้
    """
    groups = sorted(comp.groups, key=lambda g: g.sort_order)
    rows_by_group = {g.id: standings.get(g.id, []) for g in groups}

    if comp.best_runnerup_count == 0 and comp.advance_per_group == 2 and len(groups) == 2:
        a, b = groups[0], groups[1]
        return [rows_by_group[a.id][0]["team"], rows_by_group[b.id][1]["team"], rows_by_group[b.id][0]["team"], rows_by_group[a.id][1]["team"]]

    if comp.best_runnerup_count == 0 and comp.advance_per_group == 2 and len(groups) == 4:
        a, b, c, d = groups[:4]
        return [
            rows_by_group[a.id][0]["team"], rows_by_group[b.id][1]["team"],
            rows_by_group[b.id][0]["team"], rows_by_group[a.id][1]["team"],
            rows_by_group[c.id][0]["team"], rows_by_group[d.id][1]["team"],
            rows_by_group[d.id][0]["team"], rows_by_group[c.id][1]["team"],
        ]

    if comp.best_runnerup_count == 0 and comp.advance_per_group == 1 and len(groups) == 4:
        a, b, c, d = groups[:4]
        return [rows_by_group[a.id][0]["team"], rows_by_group[d.id][0]["team"], rows_by_group[b.id][0]["team"], rows_by_group[c.id][0]["team"]]

    qualifiers = calculate_rr_qualifiers(comp, standings)
    return [q["team"] for q in qualifiers]


def advance_knockout_if_ready(comp):
    """ถ้ารอบล่าสุดแข่งครบ สร้างรอบถัดไปอัตโนมัติ"""
    from models import KnockoutMatch
    rounds = sorted({m.round_no for m in comp.matches})
    if not rounds:
        return
    latest_round = rounds[-1]
    latest_matches = [m for m in comp.matches if m.round_no == latest_round]
    if any(not m.winner_team_id for m in latest_matches):
        return
    if len(latest_matches) <= 1:
        comp.status = "completed"
        db.session.commit()
        return
    if any(m.round_no == latest_round + 1 for m in comp.matches):
        return
    winners = [m.winner_team for m in sorted(latest_matches, key=lambda x: x.match_no) if m.winner_team]
    round_name = knockout_round_name(len(winners))
    pair_count = (len(winners) + 1) // 2
    for i in range(pair_count):
        team_a = winners[i * 2] if i * 2 < len(winners) else None
        team_b = winners[i * 2 + 1] if i * 2 + 1 < len(winners) else None
        match = KnockoutMatch(
            competition_id=comp.id,
            round_no=latest_round + 1,
            round_name=round_name,
            match_no=i + 1,
            team_a_id=team_a.id if team_a else None,
            team_b_id=team_b.id if team_b else None,
            status="scheduled",
        )
        if team_a and not team_b:
            match.winner_team_id = team_a.id
            match.status = "completed"
        db.session.add(match)
    comp.status = "in_progress"
    db.session.commit()


def create_rr_groups(comp):
    from models import RoundRobinGroup
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i in range(comp.num_groups):
        db.session.add(RoundRobinGroup(competition_id=comp.id, name=letters[i] if i < len(letters) else str(i+1), sort_order=i+1))
    db.session.flush()


def assign_teams_auto(comp, team_ids):
    from models import RoundRobinGroupTeam
    groups = list(comp.groups)
    random.shuffle(team_ids)
    for idx, team_id in enumerate(team_ids):
        group = groups[idx % len(groups)]
        db.session.add(RoundRobinGroupTeam(group_id=group.id, team_id=team_id, sort_order=(idx // len(groups)) + 1))
    db.session.flush()


def generate_rr_matches(comp):
    from models import RoundRobinMatch
    match_no = 1
    for group in comp.groups:
        teams = [gt.team for gt in group.group_teams]
        round_no = 1
        for i in range(len(teams)):
            for j in range(i + 1, len(teams)):
                db.session.add(RoundRobinMatch(
                    competition_id=comp.id,
                    group_id=group.id,
                    round_no=round_no,
                    match_no=match_no,
                    team_a_id=teams[i].id,
                    team_b_id=teams[j].id,
                ))
                match_no += 1
                round_no += 1
    db.session.flush()



def get_live_board_setting(event):
    from models import LiveBoardSetting
    setting = LiveBoardSetting.query.filter_by(event_id=event.id).first()
    if not setting:
        setting = LiveBoardSetting(
            event_id=event.id,
            marquee_text=f"ยินดีต้อนรับสู่ {event.name} · KRURUK SPORTS",
            theme="stadium",
            refresh_seconds=10,
            show_medals=True,
            show_schedule=True,
            show_results=True,
            show_rr_standings=True,
        )
        db.session.add(setting)
        db.session.commit()
    return setting


def team_logo_url(team):
    if team and team.logo:
        return url_for("static", filename=f"uploads/{team.logo}")
    return ""


def build_live_board_data(event, setting):
    from models import RoundRobinCompetition, RoundRobinMatch, KnockoutCompetition, KnockoutMatch, RankingCompetition, RankingResult, ContestCompetition, ContestResult
    entries = collect_medal_entries(event)
    medal_table = build_medal_table(event, entries)[:8]
    medals = []
    for idx, row in enumerate(medal_table, start=1):
        team = row["team"]
        medals.append({
            "rank": idx,
            "team": team.name,
            "color": team.color_hex or "#64748b",
            "logo": team_logo_url(team),
            "gold": row["gold"],
            "silver": row["silver"],
            "bronze": row["bronze"],
            "total": row["total"],
        })

    latest_results = []
    rr_matches = (RoundRobinMatch.query
        .join(RoundRobinCompetition)
        .filter(RoundRobinCompetition.event_id == event.id, RoundRobinMatch.status == "completed")
        .order_by(RoundRobinMatch.id.desc()).limit(12).all())
    for m in rr_matches:
        latest_results.append({
            "type": "Round Robin",
            "competition": m.competition.name,
            "title": f"{m.team_a.name} {m.score_a} - {m.score_b} {m.team_b.name}",
            "detail": f"กลุ่ม {m.group.name} · รอบ {m.round_no}",
            "team_color": m.team_a.color_hex or "#0f172a",
        })

    ranking_results = (RankingResult.query
        .join(RankingCompetition)
        .filter(RankingCompetition.event_id == event.id, RankingResult.rank.in_([1,2,3]))
        .order_by(RankingResult.id.desc()).limit(12).all())
    for r in ranking_results:
        latest_results.append({
            "type": "Ranking",
            "competition": r.competition.name,
            "title": f"อันดับ {r.rank} {r.team.name if r.team else ''}",
            "detail": r.competitor_name or (r.athlete.full_name if r.athlete else ""),
            "team_color": r.team.color_hex if r.team else "#0f172a",
        })

    contest_results = (ContestResult.query
        .join(ContestCompetition)
        .filter(ContestCompetition.event_id == event.id, ContestResult.rank.in_([1,2,3]))
        .order_by(ContestResult.id.desc()).limit(12).all())
    for r in contest_results:
        latest_results.append({
            "type": "Contest",
            "competition": r.competition.name,
            "title": f"อันดับ {r.rank} {r.team.name if r.team else ''}",
            "detail": f"{r.total_score:g} คะแนน",
            "team_color": r.team.color_hex if r.team else "#0f172a",
        })
    latest_results = latest_results[:16]

    schedule = []
    upcoming = (RoundRobinMatch.query
        .join(RoundRobinCompetition)
        .filter(RoundRobinCompetition.event_id == event.id, RoundRobinMatch.status != "completed")
        .order_by(RoundRobinMatch.round_no.asc(), RoundRobinMatch.match_no.asc()).limit(12).all())
    for m in upcoming:
        schedule.append({
            "competition": m.competition.name,
            "round": m.round_no,
            "match": m.match_no,
            "group": m.group.name,
            "team_a": m.team_a.name,
            "team_b": m.team_b.name,
            "color_a": m.team_a.color_hex or "#0f172a",
            "color_b": m.team_b.color_hex or "#0f172a",
        })

    standings = []
    rr_comps = RoundRobinCompetition.query.filter_by(event_id=event.id).order_by(RoundRobinCompetition.created_at.desc()).limit(4).all()
    for comp in rr_comps:
        comp_standings = calculate_rr_standings(comp)
        for group in comp.groups:
            rows = comp_standings.get(group.id, [])[:4]
            standings.append({
                "competition": comp.name,
                "group": group.name,
                "rows": [{
                    "rank": r["rank"],
                    "team": r["team"].name,
                    "color": r["team"].color_hex or "#64748b",
                    "played": r["played"],
                    "wins": r["wins"],
                    "draws": r["draws"],
                    "losses": r["losses"],
                    "points": r["points"],
                    "goal_diff": r["goal_diff"],
                } for r in rows]
            })

    champion = medals[0] if medals and medals[0]["total"] > 0 else None
    return {
        "event": {
            "name": event.name,
            "year": event.competition_year or "",
            "location": event.location or "",
            "logo": url_for("static", filename=f"uploads/{event.logo}") if event.logo else "",
            "theme_color": event.theme_color or "#2563eb",
        },
        "setting": {
            "marquee_text": setting.marquee_text or "",
            "refresh_seconds": setting.refresh_seconds or 10,
            "show_medals": setting.show_medals,
            "show_schedule": setting.show_schedule,
            "show_results": setting.show_results,
            "show_rr_standings": setting.show_rr_standings,
        },
        "champion": champion,
        "medals": medals,
        "latest_results": latest_results,
        "schedule": schedule,
        "standings": standings,
        "updated_at": datetime.now().strftime("%H:%M:%S"),
    }



def rr_result_type(comp):
    division = comp.sport_division
    if division:
        return division.result_type or (division.sport.result_type if division.sport else "score_only") or "score_only"
    return "score_only"


def rr_set_config(comp):
    division = comp.sport_division
    sport = division.sport if division else None
    max_sets = (division.max_sets if division and division.max_sets else 0) or (sport.max_sets if sport else 0) or 3
    points_per_set = (division.points_per_set if division and division.points_per_set else 0) or (sport.points_per_set if sport else 0) or 21
    sets_to_win = (division.sets_to_win if division and division.sets_to_win else 0) or (sport.sets_to_win if sport else 0) or ((max_sets // 2) + 1)
    return {"max_sets": max_sets, "points_per_set": points_per_set, "sets_to_win": sets_to_win}


def parse_set_scores(match):
    if not match.set_scores:
        return []
    try:
        data = json.loads(match.set_scores)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def set_score_rows(match):
    """คืนคะแนนรายเซตในรูปแบบที่ template อ่านง่าย ไม่โชว์ JSON ดิบ"""
    rows = []
    for idx, row in enumerate(parse_set_scores(match), start=1):
        if not isinstance(row, dict):
            continue
        set_no = safe_int(row.get("set")) or idx
        a = safe_int(row.get("a"))
        b = safe_int(row.get("b"))
        if a > b:
            winner = "a"
        elif b > a:
            winner = "b"
        else:
            winner = "draw"
        rows.append({"set": set_no, "a": a, "b": b, "winner": winner})
    return rows


def parse_score_history(match):
    if not getattr(match, "score_history", None):
        return []
    try:
        data = json.loads(match.score_history)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def normalize_score_history_payload(raw):
    if not raw:
        return None
    try:
        data = json.loads(raw)
        if not isinstance(data, list):
            return None
        cleaned = []
        for idx, row in enumerate(data[-500:], start=1):
            if not isinstance(row, dict):
                continue
            cleaned.append({
                "no": safe_int(row.get("no")) or idx,
                "set": safe_int(row.get("set")),
                "team": str(row.get("team", ""))[:200],
                "side": str(row.get("side", ""))[:1],
                "a": safe_int(row.get("a")),
                "b": safe_int(row.get("b")),
                "at": str(row.get("at", ""))[:30],
            })
        return json.dumps(cleaned, ensure_ascii=False) if cleaned else None
    except Exception:
        return None


def knockout_set_config(comp):
    max_sets = comp.max_sets or 3
    points_per_set = comp.points_per_set or 21
    sets_to_win = comp.sets_to_win or ((max_sets // 2) + 1)
    return {"max_sets": max_sets, "points_per_set": points_per_set, "sets_to_win": sets_to_win}


def set_point_totals(match):
    total_a = 0
    total_b = 0
    for row in parse_set_scores(match):
        total_a += safe_int(row.get("a"))
        total_b += safe_int(row.get("b"))
    return total_a, total_b

def calculate_rr_standings(comp):
    data = {}
    for group in comp.groups:
        rows = {}
        for gt in group.group_teams:
            rows[gt.team_id] = {
                "team": gt.team,
                "group": group,
                "played": 0,
                "wins": 0,
                "draws": 0,
                "losses": 0,
                "points": 0,
                "goals_for": 0,
                "goals_against": 0,
                "goal_diff": 0,
                "sets_for": 0,
                "sets_against": 0,
                "set_diff": 0,
                "rank": 0,
            }
        matches = [m for m in comp.matches if m.group_id == group.id and m.status == "completed" and m.score_a is not None and m.score_b is not None]
        for m in matches:
            if m.team_a_id not in rows or m.team_b_id not in rows:
                continue
            a = rows[m.team_a_id]
            b = rows[m.team_b_id]
            a["played"] += 1
            b["played"] += 1
            if rr_result_type(comp) == "set_based":
                point_a, point_b = set_point_totals(m)
                # ช่องได้/เสีย ใช้เป็นแต้มรวมจากทุกเซต เพื่อให้ tie-break แบบผลต่างแต้มทำงานได้
                a["goals_for"] += point_a
                a["goals_against"] += point_b
                b["goals_for"] += point_b
                b["goals_against"] += point_a
                a["sets_for"] += m.set_a or 0
                a["sets_against"] += m.set_b or 0
                b["sets_for"] += m.set_b or 0
                b["sets_against"] += m.set_a or 0
                result_a = m.set_a or 0
                result_b = m.set_b or 0
            else:
                a["goals_for"] += m.score_a
                a["goals_against"] += m.score_b
                b["goals_for"] += m.score_b
                b["goals_against"] += m.score_a
                result_a = m.score_a
                result_b = m.score_b
            if result_a > result_b:
                a["wins"] += 1; b["losses"] += 1
                a["points"] += comp.win_points; b["points"] += comp.loss_points
            elif result_a < result_b:
                b["wins"] += 1; a["losses"] += 1
                b["points"] += comp.win_points; a["points"] += comp.loss_points
            else:
                a["draws"] += 1; b["draws"] += 1
                a["points"] += comp.draw_points; b["points"] += comp.draw_points
        for row in rows.values():
            row["goal_diff"] = row["goals_for"] - row["goals_against"]
            row["set_diff"] = row["sets_for"] - row["sets_against"]
        ordered = sorted(rows.values(), key=lambda r: rr_sort_key(comp, r), reverse=True)
        for idx, row in enumerate(ordered, start=1):
            row["rank"] = idx
        data[group.id] = ordered
    return data


def rr_sort_key(comp, row):
    mapping = {
        "points": row["points"],
        "goal_diff": row["goal_diff"],
        "goals_for": row["goals_for"],
        "wins": row["wins"],
        "set_diff": row["set_diff"],
        "sets_for": row["sets_for"],
        "head_to_head": 0,
        "draw_lots": 0,
    }
    return tuple(mapping.get(k, 0) for k in comp.tiebreaker_list) + (row["team"].name,)


def calculate_rr_qualifiers(comp, standings):
    qualifiers = []
    runnerups = []
    for group in comp.groups:
        rows = standings.get(group.id, [])
        for row in rows:
            if row["rank"] <= comp.advance_per_group:
                qualifiers.append({**row, "reason": f"อันดับ {row['rank']} กลุ่ม {group.name}"})
            elif row["rank"] == comp.advance_per_group + 1 and comp.best_runnerup_count > 0:
                runnerups.append(row)
    runnerups = sorted(runnerups, key=lambda r: rr_sort_key(comp, r), reverse=True)
    for row in runnerups[:comp.best_runnerup_count]:
        qualifiers.append({**row, "reason": f"Best Runner-up กลุ่ม {row['group'].name}"})
    return qualifiers


def save_ranking_result_from_request(result, comp):
    from models import Athlete, Team
    team_id = safe_int_or_none(request.form.get("team_id"))
    athlete_id = safe_int_or_none(request.form.get("athlete_id"))
    athlete = Athlete.query.get(athlete_id) if athlete_id else None
    if athlete and athlete.team.event_id == comp.event_id:
        result.athlete_id = athlete.id
        result.team_id = athlete.team_id
        result.competitor_name = athlete.full_name
    else:
        team = Team.query.get(team_id) if team_id else None
        if team and team.event_id == comp.event_id:
            result.team_id = team.id
        result.athlete_id = None
        result.competitor_name = request.form.get("competitor_name", "").strip()
    result.rank = safe_int_or_none(request.form.get("rank"))
    result.time_value = request.form.get("time_value", "").strip()
    result.distance_value = request.form.get("distance_value", "").strip()
    result.score_value = request.form.get("score_value", "").strip()
    result.note = request.form.get("note", "").strip()
    result.medal = medal_for_rank(result.rank)


def medal_for_rank(rank):
    if rank == 1:
        return "gold"
    if rank == 2:
        return "silver"
    if rank == 3:
        return "bronze"
    return None


def refresh_ranking_medals(comp):
    from models import RankingResult
    for r in RankingResult.query.filter_by(competition_id=comp.id).all():
        r.medal = medal_for_rank(r.rank)
    db.session.commit()


def build_ranking_workbook(comp):
    from openpyxl import Workbook
    from models import RankingResult
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking Results"
    ws.append(["รายการ", comp.name])
    ws.append(["งาน", comp.event.name])
    ws.append([])
    ws.append(["อันดับ", "เหรียญ", "ทีม", "ผู้แข่งขัน", "เวลา", "ระยะ", "คะแนน", "หมายเหตุ"])
    results = RankingResult.query.filter_by(competition_id=comp.id).order_by(RankingResult.rank.asc().nullslast(), RankingResult.created_at.asc()).all()
    medal_map = {"gold": "ทอง", "silver": "เงิน", "bronze": "ทองแดง", None: ""}
    for r in results:
        ws.append([r.rank, medal_map.get(r.medal, r.medal or ""), r.team.name if r.team else "", r.competitor_name or (r.athlete.full_name if r.athlete else ""), r.time_value, r.distance_value, r.score_value, r.note])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def division_matches_filter(division, sport_id=None, class_name="", gender=""):
    if sport_id and (not division or division.sport_id != sport_id):
        return False
    if class_name and (not division or division.class_name != class_name):
        return False
    if gender and (not division or division.gender != gender):
        return False
    return True


def medal_label(medal):
    return {"gold": "ทอง", "silver": "เงิน", "bronze": "ทองแดง"}.get(medal, "")


def medal_for_position(position):
    if position == 1:
        return "gold"
    if position == 2:
        return "silver"
    if position == 3:
        return "bronze"
    return None


def collect_medal_entries(event, sport_id=None, class_name="", gender=""):
    from models import RankingCompetition, RankingResult, RoundRobinCompetition, ContestCompetition, ContestResult
    entries = []

    ranking_comps = RankingCompetition.query.filter_by(event_id=event.id).all()
    for comp in ranking_comps:
        if not division_matches_filter(comp.sport_division, sport_id, class_name, gender):
            continue
        results = RankingResult.query.filter_by(competition_id=comp.id).order_by(RankingResult.rank.asc().nullslast(), RankingResult.created_at.asc()).all()
        for result in results:
            medal = result.medal or medal_for_rank(result.rank)
            if medal not in ("gold", "silver", "bronze"):
                continue
            entries.append({
                "source": "Ranking",
                "competition": comp.name,
                "sport": comp.sport_division.sport.name if comp.sport_division else "-",
                "class_name": comp.sport_division.class_name if comp.sport_division else "-",
                "gender": comp.sport_division.gender if comp.sport_division else "-",
                "team": result.team,
                "team_id": result.team_id,
                "medal": medal,
                "rank": result.rank,
                "detail": result.competitor_name or (result.athlete.full_name if result.athlete else ""),
            })


    contest_comps = ContestCompetition.query.filter_by(event_id=event.id).all()
    for comp in contest_comps:
        if not division_matches_filter(comp.sport_division, sport_id, class_name, gender):
            continue
        results = ContestResult.query.filter_by(competition_id=comp.id).order_by(ContestResult.rank.asc().nullslast(), ContestResult.total_score.desc()).all()
        for result in results:
            medal = result.medal or medal_for_rank(result.rank)
            if medal not in ("gold", "silver", "bronze"):
                continue
            entries.append({
                "source": "Contest",
                "competition": comp.name,
                "sport": comp.sport_division.sport.name if comp.sport_division else comp.activity_type or "กิจกรรมประกวด",
                "class_name": comp.sport_division.class_name if comp.sport_division else "-",
                "gender": comp.sport_division.gender if comp.sport_division else "-",
                "team": result.team,
                "team_id": result.team_id,
                "medal": medal,
                "rank": result.rank,
                "detail": f"{result.total_score:g} คะแนน",
            })

    rr_comps = RoundRobinCompetition.query.filter_by(event_id=event.id).all()
    for comp in rr_comps:
        if not division_matches_filter(comp.sport_division, sport_id, class_name, gender):
            continue
        standings = calculate_rr_standings(comp)
        all_rows = []
        for rows in standings.values():
            all_rows.extend(rows)
        if not all_rows:
            continue
        ordered = sorted(all_rows, key=lambda r: rr_sort_key(comp, r), reverse=True)
        for idx, row in enumerate(ordered[:3], start=1):
            entries.append({
                "source": "Round Robin",
                "competition": comp.name,
                "sport": comp.sport_division.sport.name if comp.sport_division else "-",
                "class_name": comp.sport_division.class_name if comp.sport_division else "-",
                "gender": comp.sport_division.gender if comp.sport_division else "-",
                "team": row["team"],
                "team_id": row["team"].id,
                "medal": medal_for_position(idx),
                "rank": idx,
                "detail": f"{row['points']} คะแนน / ได้เสีย {row['goal_diff']}",
            })
    return entries


def build_medal_table(event, entries):
    rows = {}
    for team in event.teams:
        rows[team.id] = {"team": team, "gold": 0, "silver": 0, "bronze": 0, "total": 0}
    for e in entries:
        team_id = e.get("team_id")
        if team_id not in rows and e.get("team"):
            rows[team_id] = {"team": e["team"], "gold": 0, "silver": 0, "bronze": 0, "total": 0}
        if team_id in rows and e.get("medal") in ("gold", "silver", "bronze"):
            rows[team_id][e["medal"]] += 1
            rows[team_id]["total"] += 1
    return sorted(rows.values(), key=lambda r: (r["gold"], r["silver"], r["bronze"], r["total"], r["team"].name), reverse=True)


def build_medal_workbook(event, table, entries):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Medal Table"
    ws.append(["งาน", event.name])
    ws.append([])
    ws.append(["อันดับ", "ทีม", "ทอง", "เงิน", "ทองแดง", "รวม"])
    for idx, row in enumerate(table, start=1):
        ws.append([idx, row["team"].name, row["gold"], row["silver"], row["bronze"], row["total"]])
    ws2 = wb.create_sheet("Medal Details")
    ws2.append(["เหรียญ", "อันดับ", "ทีม", "รายการ", "กีฬา", "รุ่น", "เพศ", "ประเภท", "รายละเอียด"])
    for e in entries:
        ws2.append([medal_label(e["medal"]), e["rank"], e["team"].name if e.get("team") else "", e["competition"], e["sport"], e["class_name"], e["gender"], e["source"], e["detail"]])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def refresh_contest_results(comp):
    from models import Team, ContestScore, ContestResult
    teams = Team.query.filter_by(event_id=comp.event_id).all()
    totals = []
    for team in teams:
        total = db.session.query(db.func.coalesce(db.func.sum(ContestScore.score), 0)).filter_by(competition_id=comp.id, team_id=team.id).scalar() or 0
        totals.append((team, float(total)))
    totals.sort(key=lambda x: (-x[1], x[0].name))
    ContestResult.query.filter_by(competition_id=comp.id).delete()
    rank = 0
    prev_score = None
    seen = 0
    for team, total in totals:
        seen += 1
        if prev_score is None or total != prev_score:
            rank = seen
            prev_score = total
        db.session.add(ContestResult(competition_id=comp.id, team_id=team.id, total_score=total, rank=rank, medal=medal_for_rank(rank)))
    db.session.commit()


def build_contest_workbook(comp):
    from openpyxl import Workbook
    from models import ContestResult, ContestScore
    wb = Workbook()
    ws = wb.active
    ws.title = "Contest Results"
    ws.append(["กิจกรรม", comp.name])
    ws.append(["ประเภท", comp.activity_type or ""])
    ws.append([])
    ws.append(["อันดับ", "เหรียญ", "ทีม", "คะแนนรวม"])
    results = ContestResult.query.filter_by(competition_id=comp.id).order_by(ContestResult.rank.asc().nullslast(), ContestResult.total_score.desc()).all()
    for r in results:
        ws.append([r.rank, medal_label(r.medal), r.team.name if r.team else "", r.total_score])
    ws2 = wb.create_sheet("Scores")
    ws2.append(["กรรมการ", "ทีม", "เกณฑ์", "คะแนน", "คะแนนเต็ม"])
    scores = ContestScore.query.filter_by(competition_id=comp.id).all()
    for s in scores:
        ws2.append([s.judge.name if s.judge else "", s.team.name if s.team else "", s.criterion.name if s.criterion else "", s.score, s.criterion.max_score if s.criterion else ""])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def fill_certificate_template_from_form(tpl):
    tpl.name = request.form.get("name", "").strip() or "Template เกียรติบัตร"
    tpl.cert_type = request.form.get("cert_type", "participant")
    tpl.title = request.form.get("title", "เกียรติบัตร").strip() or "เกียรติบัตร"
    tpl.subtitle = request.form.get("subtitle", "").strip()
    tpl.body = request.form.get("body", "").strip()
    tpl.footer_text = request.form.get("footer_text", "").strip()
    tpl.signature_left_name = request.form.get("signature_left_name", "").strip()
    tpl.signature_left_position = request.form.get("signature_left_position", "").strip()
    tpl.signature_right_name = request.form.get("signature_right_name", "").strip()
    tpl.signature_right_position = request.form.get("signature_right_position", "").strip()
    tpl.background_color = request.form.get("background_color", "#ffffff") or "#ffffff"
    tpl.accent_color = request.form.get("accent_color", "#1d4ed8") or "#1d4ed8"
    tpl.is_active = bool(request.form.get("is_active", "1"))
    for field in ["logo", "signature_left", "signature_right"]:
        uploaded = save_upload(request.files.get(field))
        if uploaded:
            setattr(tpl, field, uploaded)


def unique_verify_code():
    from models import CertificateRecipient
    while True:
        code = uuid.uuid4().hex[:12].upper()
        if not CertificateRecipient.query.filter_by(verify_code=code).first():
            return code


def create_certificate_recipient(tpl, full_name, recipient_type="manual", team_id=None, athlete_id=None, coach_id=None, role_text="", award_text="", sport_text=""):
    from models import CertificateRecipient
    full_name = (full_name or "").strip()
    if not full_name:
        return None
    exists = CertificateRecipient.query.filter_by(template_id=tpl.id, full_name=full_name, recipient_type=recipient_type, athlete_id=athlete_id, coach_id=coach_id, award_text=award_text, sport_text=sport_text).first()
    if exists:
        return None
    cert = CertificateRecipient(
        template_id=tpl.id,
        event_id=tpl.event_id,
        team_id=team_id,
        athlete_id=athlete_id,
        coach_id=coach_id,
        recipient_type=recipient_type,
        full_name=full_name,
        role_text=role_text,
        award_text=award_text,
        sport_text=sport_text,
        verify_code=unique_verify_code(),
    )
    db.session.add(cert)
    return cert


def generate_certificates_for_template(tpl, mode):
    from models import Athlete, Coach, CertificateRecipient, Team
    event = tpl.event
    created = 0
    if mode in ("participant", "athlete"):
        athletes = Athlete.query.join(Team).filter(Team.event_id == event.id).all()
        for a in athletes:
            sport_text = ", ".join([f"{r.sport_name} {r.category_name or ''} {r.gender or ''}".strip() for r in a.registrations])
            if create_certificate_recipient(tpl, a.full_name, "athlete", team_id=a.team_id, athlete_id=a.id, role_text="นักกีฬา", award_text="ผู้เข้าร่วม", sport_text=sport_text):
                created += 1
    elif mode == "coach":
        coaches = Coach.query.join(Team).filter(Team.event_id == event.id).all()
        for c in coaches:
            if create_certificate_recipient(tpl, c.full_name, "coach", team_id=c.team_id, coach_id=c.id, role_text="ผู้ฝึกสอน", award_text="ผู้ฝึกสอน", sport_text=c.sport_responsibility or ""):
                created += 1
    elif mode == "winner":
        entries = collect_medal_entries(event)
        for e in entries:
            team = e.get("team")
            name = e.get("detail") or (team.name if team else "")
            award = f"{medal_label(e.get('medal'))} อันดับ {e.get('rank')}"
            sport = f"{e.get('competition')} / {e.get('sport')} / {e.get('class_name')} / {e.get('gender')}"
            if create_certificate_recipient(tpl, name, "winner", team_id=e.get("team_id"), role_text="ผู้ได้รับรางวัล", award_text=award, sport_text=sport):
                created += 1
    elif mode in ("committee", "judge", "sponsor"):
        # ใช้การเพิ่มรายชื่อเองในหน้า Certificate สำหรับกลุ่มนี้ เพื่อไม่บังคับรูปแบบข้อมูล
        pass
    db.session.commit()
    return created


def render_certificate_body(cert):
    tpl = cert.template
    text = tpl.body or "ได้เข้าร่วมกิจกรรม/การแข่งขันในงาน {event_name}"
    replacements = {
        "{event_name}": cert.event.name if cert.event else "",
        "{name}": cert.full_name,
        "{team}": cert.team.name if cert.team else "",
        "{role}": cert.role_text or "",
        "{award}": cert.award_text or "",
        "{sport}": cert.sport_text or "",
        "{date}": cert.issued_at.strftime("%d/%m/%Y") if cert.issued_at else "",
    }
    for key, val in replacements.items():
        text = text.replace(key, val)
    return text


def make_qr_data_uri(url):
    try:
        import qrcode
        img = qrcode.make(url)
        buf = BytesIO()
        img.save(buf, format="PNG")
        encoded = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/png;base64,{encoded}"
    except Exception:
        return ""


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


app = create_app()

if __name__ == "__main__":
    # Railway/production must bind to 0.0.0.0 and use the PORT env var.
    # Local run still works with: python app.py
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
